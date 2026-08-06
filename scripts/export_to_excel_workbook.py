import os
import sys
import re
import csv
import pandas as pd
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference


def clean_phone_str(val):
    if pd.isna(val) or val is None:
        return ""
    val_s = str(val).strip()
    if not val_s or val_s.lower() == "nan":
        return ""
    # Quitar decimales si viene como float
    val_s = re.sub(r"\.0$", "", val_s)
    # Si viene en notación científica ej 5.49223e+12
    if "e+" in val_s.lower():
        try:
            val_s = str(int(float(val_s)))
        except Exception:
            pass
    return val_s


def create_excel_workbook():
    data_dir = "data"
    prospects_csv = os.path.join(data_dir, "prospectos_agencia_ia_mdp.csv")
    master_csv = os.path.join(data_dir, "master_comercios_mdp.csv")

    out_excel1 = os.path.join(data_dir, "prospeccion_agencia_ia_mdp.xlsx")
    out_excel2 = os.path.join(data_dir, "PROSPECCION_MAR_DEL_PLATA_COMPLETA.xlsx")
    out_excel3 = "PROSPECCION_MAR_DEL_PLATA_COMPLETA.xlsx"

    print("Leyendo archivos CSV...")
    df_prospects = pd.read_csv(prospects_csv, encoding="utf-8-sig", dtype=str)
    df_master = pd.read_csv(master_csv, encoding="utf-8-sig", dtype=str)

    # Limpiar teléfonos y whatsapp en los DataFrames
    for df in [df_prospects, df_master]:
        for col in ["telefono", "whatsapp"]:
            if col in df.columns:
                df[col] = df[col].apply(clean_phone_str)

    # Convertir puntaje a int
    df_prospects["puntaje_oportunidad"] = pd.to_numeric(df_prospects["puntaje_oportunidad"], errors="coerce").fillna(50).astype(int)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # Estilos reutilizables
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    font_section = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    fill_section = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    font_card_title = Font(name="Calibri", size=9, bold=True, color="595959")
    font_card_val = Font(name="Calibri", size=20, bold=True, color="1F4E78")
    font_card_hot_val = Font(name="Calibri", size=20, bold=True, color="C00000")
    fill_card = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")

    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    font_hot_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_hot_header = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    font_data = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_link = Font(name="Calibri", size=10, color="0563C1", underline="single")

    font_insight_title = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    font_insight_body = Font(name="Calibri", size=10, italic=True)

    fill_even = PatternFill(start_color="F9FAFC", end_color="F9FAFC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # =========================================================
    # 1. CREACIÓN DE LA HOJA: 📊 Dashboard & Conclusiones
    # =========================================================
    print("Generando hoja ejecutiva con gráficos: 📊 Dashboard & Conclusiones...")
    ws_dash = wb.create_sheet(title="📊 Dashboard & Conclusiones")
    ws_dash.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_dash.merge_cells("A1:F2")
    cell_t = ws_dash["A1"]
    cell_t.value = "📊 DASHBOARD EJECUTIVO Y GRÁFICOS DE PROSPECCIÓN - AGENCIA DE IA"
    cell_t.font = font_title
    cell_t.fill = fill_title
    cell_t.alignment = align_center

    # Métricas Generales (KPI Cards)
    total_leads = len(df_prospects)
    hot_leads = len(df_prospects[df_prospects["prioridad"].str.contains("ALTA", na=False)])
    wa_leads = len(df_prospects[df_prospects["whatsapp"].notna() & (df_prospects["whatsapp"] != "")])
    noweb_leads = len(df_prospects[df_prospects["tiene_web"] == "NO"])

    kpis = [
        ("A4:A5", "A4", "A5", "TOTAL PROSPECTOS", f"{total_leads:,}", font_card_val),
        ("B4:B5", "B4", "B5", "LEADS CALIENTES (ALTA 🔥)", f"{hot_leads:,}", font_card_hot_val),
        ("C4:C5", "C4", "C5", "CONTACTOS WHATSAPP DIRECTO", f"{wa_leads:,}", font_card_val),
        ("D4:D5", "D4", "D5", "COMERCIOS SIN WEB (71%)", f"{noweb_leads:,}", font_card_val),
    ]

    for merge_range, top_cell, bot_cell, label, val, font_v in kpis:
        ws_dash[top_cell] = label
        ws_dash[top_cell].font = font_card_title
        ws_dash[top_cell].fill = fill_card
        ws_dash[top_cell].alignment = align_center
        ws_dash[top_cell].border = thin_border

        ws_dash[bot_cell] = val
        ws_dash[bot_cell].font = font_v
        ws_dash[bot_cell].fill = fill_card
        ws_dash[bot_cell].alignment = align_center
        ws_dash[bot_cell].border = thin_border

    # Tabla Cruzada 1: Matriz por Rubro
    ws_dash.merge_cells("A7:F7")
    ws_dash["A7"] = "MATRIZ DE OPORTUNIDAD COMERCIAL POR RUBRO TARGET"
    ws_dash["A7"].font = font_section
    ws_dash["A7"].fill = fill_section
    ws_dash["A7"].alignment = align_center

    tbl1_headers = ["Rubro Comercial Target", "Total Prospects", "Leads Calientes 🔥", "% Oportunidad Alta", "Con WhatsApp Directo", "Servicio Principal de IA Recomendado"]
    ws_dash.append(tbl1_headers)

    for c_idx, h in enumerate(tbl1_headers, 1):
        cell = ws_dash.cell(row=8, column=c_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    sectors_summary = []
    for sector_name, group in df_prospects.groupby("sector"):
        t_sec = len(group)
        h_sec = len(group[group["prioridad"].str.contains("ALTA", na=False)])
        wa_sec = len(group[group["whatsapp"].notna() & (group["whatsapp"] != "")])
        pct_h = (h_sec / t_sec) * 100 if t_sec > 0 else 0
        top_pitch = group["servicio_pitch_sugerido"].mode()[0] if not group["servicio_pitch_sugerido"].empty else "Servicio IA Personalizado"
        sectors_summary.append((sector_name, t_sec, h_sec, pct_h, wa_sec, top_pitch))

    sectors_summary.sort(key=lambda x: x[2], reverse=True)

    r_idx = 9
    for s_name, t_sec, h_sec, pct_h, wa_sec, pitch_name in sectors_summary:
        row_vals = [s_name, t_sec, h_sec, f"{pct_h:.1f}%", wa_sec, pitch_name]
        ws_dash.append(row_vals)
        r_fill = fill_even if r_idx % 2 == 0 else fill_odd

        for c_idx in range(1, 7):
            cell = ws_dash.cell(row=r_idx, column=c_idx)
            cell.font = font_data
            cell.fill = r_fill
            cell.border = thin_border
            cell.alignment = align_left if c_idx in [1, 6] else align_center

        r_idx += 1

    # Fila de Totales
    ws_dash.append(["TOTAL CONSOLIDADO", total_leads, hot_leads, f"{(hot_leads/total_leads*100):.1f}%", wa_leads, "Estrategia Multicanal"])
    for c_idx in range(1, 7):
        cell = ws_dash.cell(row=r_idx, column=c_idx)
        cell.font = font_bold
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.border = thin_border
        cell.alignment = align_left if c_idx in [1, 6] else align_center

    # Tabla Cruzada 2: Presencia Web vs WhatsApp
    start_r2 = r_idx + 2
    ws_dash.merge_cells(f"A{start_r2}:F{start_r2}")
    ws_dash[f"A{start_r2}"] = "ANÁLISIS DE MADUREZ DIGITAL Y OPORTUNIDAD DE VENTA"
    ws_dash[f"A{start_r2}"].font = font_section
    ws_dash[f"A{start_r2}"].fill = fill_section
    ws_dash[f"A{start_r2}"].alignment = align_center

    tbl2_headers = ["Estado de Presencia Digital", "Con WhatsApp Directo", "Sin WhatsApp Directo", "Total Leads", "% del Mercado", "Servicio Sugerido de Entrada"]
    r_h2 = start_r2 + 1
    ws_dash.append(tbl2_headers)

    for c_idx, h in enumerate(tbl2_headers, 1):
        cell = ws_dash.cell(row=r_h2, column=c_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    noweb_wa = len(df_prospects[(df_prospects["tiene_web"] == "NO") & df_prospects["whatsapp"].notna() & (df_prospects["whatsapp"] != "")])
    noweb_nowa = len(df_prospects[(df_prospects["tiene_web"] == "NO") & (df_prospects["whatsapp"].isna() | (df_prospects["whatsapp"] == ""))])
    noweb_tot = noweb_wa + noweb_nowa

    web_wa = len(df_prospects[(df_prospects["tiene_web"] == "SI") & df_prospects["whatsapp"].notna() & (df_prospects["whatsapp"] != "")])
    web_nowa = len(df_prospects[(df_prospects["tiene_web"] == "SI") & (df_prospects["whatsapp"].isna() | (df_prospects["whatsapp"] == ""))])
    web_tot = web_wa + web_nowa

    cross_matrix = [
        ("Sin Sitio Web (Solo Redes/WhatsApp)", noweb_wa, noweb_nowa, noweb_tot, f"{(noweb_tot/total_leads*100):.1f}%", "Venta de Sitio Web Responsivo + Bot de Atención IA 24/7"),
        ("Con Sitio Web (Sin Bot de IA)", web_wa, web_nowa, web_tot, f"{(web_tot/total_leads*100):.1f}%", "Integración de Bot de WhatsApp 24/7 + Agendador de Turnos"),
    ]

    r_idx2 = r_h2 + 1
    for row_v in cross_matrix:
        ws_dash.append(list(row_v))
        r_fill = fill_even if r_idx2 % 2 == 0 else fill_odd
        for c_idx in range(1, 7):
            cell = ws_dash.cell(row=r_idx2, column=c_idx)
            cell.font = font_data
            cell.fill = r_fill
            cell.border = thin_border
            cell.alignment = align_left if c_idx in [1, 6] else align_center
        r_idx2 += 1

    # Gráficos
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Leads Calientes 🔥 por Rubro Target"
    chart1.y_axis.title = "Leads Calientes"
    chart1.x_axis.title = "Rubro Comercial"

    data1 = Reference(ws_dash, min_col=3, min_row=8, max_row=8 + len(sectors_summary))
    cats1 = Reference(ws_dash, min_col=1, min_row=9, max_row=8 + len(sectors_summary))
    chart1.add_data(data1, titles_from_data=False)
    chart1.set_categories(cats1)
    chart1.legend = None
    chart1.width = 16
    chart1.height = 10
    ws_dash.add_chart(chart1, "H4")

    chart2 = PieChart()
    chart2.title = "Distribución Digital (% Sin Web vs Con Web)"
    data2 = Reference(ws_dash, min_col=4, min_row=r_h2 + 1, max_row=r_h2 + 2)
    cats2 = Reference(ws_dash, min_col=1, min_row=r_h2 + 1, max_row=r_h2 + 2)
    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(cats2)
    chart2.width = 15
    chart2.height = 10
    ws_dash.add_chart(chart2, "H19")

    # Helper para agregar DataFrames estándar formateando números de WhatsApp y teléfono como TEXTO EXPLÍCITO
    def add_df_to_sheet(wb, title, df, is_hot=False):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True

        headers = list(df.columns)
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_hot_header if is_hot else font_header
            cell.fill = fill_hot_header if is_hot else fill_header
            cell.alignment = align_center
            cell.border = thin_border

        for row_idx, row_data in enumerate(df.values, 2):
            ws.append(list(row_data))
            row_fill = fill_even if row_idx % 2 == 0 else fill_odd

            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = align_left

                header_name = headers[col_idx - 1]
                val_str = clean_phone_str(cell.value)

                # Formatear WhatsApp y Teléfono explícitamente como TEXTO para evitar e+12
                if header_name in ["whatsapp", "telefono"]:
                    cell.number_format = "@"
                    cell.value = val_str
                    cell.alignment = align_center

                # Detectar links de WhatsApp o URL
                if "wa.me" in str(cell.value or "") or "http://" in str(cell.value or "") or "https://" in str(cell.value or ""):
                    cell.font = font_link
                    cell.hyperlink = str(cell.value)

                if header_name in ["id_lead", "id", "prioridad", "puntaje_oportunidad", "tiene_web", "zona", "sector"]:
                    cell.alignment = align_center

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    # 2. Hoja: 🔥 Leads Calientes
    df_hot = df_prospects[df_prospects["prioridad"].str.contains("ALTA", na=False)].copy()
    print(f"Generando hoja: 🔥 Leads Calientes ({len(df_hot)} registros)...")
    add_df_to_sheet(wb, "🔥 Leads Calientes", df_hot, is_hot=True)

    # 2b. Hoja dedicada: 🖥️ Venta Gestión Comercial (POS, Stock, Cuentas Corrientes y Cierre de Caja)
    if "apto_gestion_comercial" in df_prospects.columns:
        df_gestion = df_prospects[df_prospects["apto_gestion_comercial"].str.contains("SI", na=False)].copy()
    else:
        df_gestion = df_prospects[df_prospects["sector"].isin(["Comercio General", "Showrooms e Indumentaria (Instagram)", "Comida a Domicilio y Delivery", "Automotriz y Servicios"])].copy()
    
    print(f"Generando hoja dedicada: 🖥️ Venta Gestión Comercial ({len(df_gestion)} registros)...")
    add_df_to_sheet(wb, "🖥️ Venta Gestión Comercial", df_gestion, is_hot=True)

    # 3. Hoja: 🛒 Showrooms e Instagram
    df_inst = df_prospects[df_prospects["sector"].isin(["Showrooms e Indumentaria (Instagram)", "Comida a Domicilio y Delivery"])].copy()
    print(f"Generando hoja: 🛒 Showrooms e Instagram ({len(df_inst)} registros)...")
    add_df_to_sheet(wb, "🛒 Showrooms e Instagram", df_inst)

    # 4. Hoja: 🏥 Salud y Estética
    df_salud = df_prospects[df_prospects["sector"] == "Salud y Estética"].copy()
    print(f"Generando hoja: 🏥 Salud y Estética ({len(df_salud)} registros)...")
    add_df_to_sheet(wb, "🏥 Salud y Estética", df_salud)

    # 5. Hoja: 🏠 Inmobiliarias
    df_inmo = df_prospects[df_prospects["sector"] == "Inmobiliarias"].copy()
    print(f"Generando hoja: 🏠 Inmobiliarias ({len(df_inmo)} registros)...")
    add_df_to_sheet(wb, "🏠 Inmobiliarias", df_inmo)

    # 6. Hoja: 🏨 Turismo y Hoteles
    df_turi = df_prospects[df_prospects["sector"] == "Turismo y Alojamiento"].copy()
    print(f"Generando hoja: 🏨 Turismo y Hoteles ({len(df_turi)} registros)...")
    add_df_to_sheet(wb, "🏨 Turismo y Hoteles", df_turi)

    # 7. Hoja: 🚗 Lavaderos y Automotriz
    df_auto = df_prospects[df_prospects["sector"] == "Automotriz y Servicios"].copy()
    print(f"Generando hoja: 🚗 Lavaderos y Automotriz ({len(df_auto)} registros)...")
    add_df_to_sheet(wb, "🚗 Lavaderos y Automotriz", df_auto)

    # 8. Hoja: 🏬 Comercio General & Minorista
    df_comercio = df_prospects[df_prospects["sector"] == "Comercio General"].copy()
    print(f"Generando hoja: 🏬 Comercio General & Minorista ({len(df_comercio)} registros)...")
    add_df_to_sheet(wb, "🏬 Comercio General & Minorista", df_comercio)

    # 9. Hoja: 📊 Todos los Prospectos
    print(f"Generando hoja: 📊 Todos los Prospectos ({len(df_prospects)} registros)...")
    add_df_to_sheet(wb, "📊 Todos los Prospectos", df_prospects)

    # 10. Hoja: 🏢 Master Comercios MDP (con Puntaje y Prioridad incorporados)
    df_master_scored = df_master.copy()
    score_map = dict(zip(df_prospects["nombre"].str.lower(), df_prospects["puntaje_oportunidad"]))
    priority_map = dict(zip(df_prospects["nombre"].str.lower(), df_prospects["prioridad"]))
    pitch_map = dict(zip(df_prospects["nombre"].str.lower(), df_prospects["servicio_pitch_sugerido"]))

    df_master_scored["prioridad"] = df_master_scored["nombre"].str.lower().map(priority_map).fillna("MEDIA (Lead Templado ⚡)")
    df_master_scored["puntaje_oportunidad"] = df_master_scored["nombre"].str.lower().map(score_map).fillna(50)
    df_master_scored["servicio_pitch_sugerido"] = df_master_scored["nombre"].str.lower().map(pitch_map).fillna("Creación de Página Web Profesional + Bot de Atención Automatizado 24/7")

    print(f"Generando hoja: 🏢 Master Comercios MDP con Puntajes ({len(df_master_scored)} registros)...")
    add_df_to_sheet(wb, "🏢 Master Comercios MDP", df_master_scored)

    # Guardar en las 3 ubicaciones para evitar bloqueos de archivo en Excel
    for path in [out_excel1, out_excel2, out_excel3]:
        try:
            wb.save(path)
            print(f"[GUARDADO EXITO] {path}")
        except Exception as e:
            print(f"[WARNING] No se pudo guardar en {path}: {e}")


if __name__ == "__main__":
    create_excel_workbook()
