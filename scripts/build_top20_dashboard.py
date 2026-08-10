# -*- coding: utf-8 -*-
"""
build_top20_dashboard.py
Genera outreach/top20_mdp_dashboard.html a partir de PROSPECCION_MAR_DEL_PLATA_COMPLETA.xlsx
con los 20 leads top (balanceados por rubro) y su plan de mensajeria integrado.
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "PROSPECCION_MAR_DEL_PLATA_COMPLETA.xlsx"
OUT = ROOT / "outreach" / "top20_mdp_dashboard.html"

LEADS = [
    {"id": "LEAD-IA-1313", "slug": "clinica-luro", "rubro": "Salud y Estetica",
     "local_html": "webs/leads/LEAD-IA-1313-clinica-luro/index.html", "demo": "salud-estetica"},
    {"id": "LEAD-IA-1301", "slug": "sanatorio-avenida", "rubro": "Salud y Estetica",
     "local_html": "webs/leads/LEAD-IA-1301-sanatorio-avenida/index.html", "demo": "salud-estetica"},
    {"id": "LEAD-IA-1304", "slug": "cardiologia-dra-ana-gracia-alonso", "rubro": "Salud y Estetica",
     "local_html": "webs/leads/LEAD-IA-1304-cardiologia-dra-ana-gracia-alonso/index.html", "demo": "salud-estetica"},
    {"id": "LEAD-IA-1279", "slug": "consultorios-integrales-colinas", "rubro": "Salud y Estetica",
     "local_html": "webs/leads/LEAD-IA-1279-consultorios-integrales-colinas/index.html", "demo": "salud-estetica"},
    {"id": "LEAD-IA-1300", "slug": "oceano-mar-consultorios", "rubro": "Salud y Estetica",
     "local_html": "webs/leads/LEAD-IA-1300-oceano-mar-consultorios/index.html", "demo": "salud-estetica"},
    {"id": "LEAD-IA-1275", "slug": "centro-medico-edison", "rubro": "Salud y Estetica",
     "local_html": "webs/leads/LEAD-IA-1275-centro-medico-edison/index.html", "demo": "salud-estetica"},
    {"id": "LEAD-IA-1290", "slug": "kytos-salud-integral", "rubro": "Salud y Estetica",
     "local_html": "webs/leads/LEAD-IA-1290-kytos-salud-integral/index.html", "demo": "salud-estetica"},
    {"id": "LEAD-IA-1294", "slug": "san-lorenzo-instituto-medico", "rubro": "Salud y Estetica",
     "local_html": "webs/leads/LEAD-IA-1294-san-lorenzo-instituto-medico/index.html", "demo": "salud-estetica"},
    {"id": "LEAD-IA-0487", "slug": "molduras-aguilera", "rubro": "Comercio General",
     "local_html": "webs/leads/LEAD-IA-0487-molduras-aguilera/index.html", "demo": "comercio-minorista"},
    {"id": "LEAD-IA-0496", "slug": "humus-mar-del-plata", "rubro": "Comercio General",
     "local_html": "webs/leads/LEAD-IA-0496-humus-mar-del-plata/index.html", "demo": "comercio-minorista"},
    {"id": "LEAD-IA-0497", "slug": "all-design", "rubro": "Comercio General",
     "local_html": "webs/leads/LEAD-IA-0497-all-design/index.html", "demo": "comercio-minorista"},
    {"id": "LEAD-IA-0512", "slug": "krearte", "rubro": "Comercio General",
     "local_html": "webs/leads/LEAD-IA-0512-krearte/index.html", "demo": "comercio-minorista"},
    {"id": "LEAD-IA-0509", "slug": "monograf", "rubro": "Comercio General",
     "local_html": "webs/leads/LEAD-IA-0509-monograf/index.html", "demo": "comercio-minorista"},
    {"id": "LEAD-IA-0251", "slug": "terca", "rubro": "Delivery / Gastronomia",
     "local_html": "webs/terca/index.html", "demo": "delivery-gastronomia"},
    {"id": "LEAD-IA-0267", "slug": "verde-limon", "rubro": "Delivery / Gastronomia",
     "local_html": "webs/verde-limon/index.html", "demo": "delivery-gastronomia"},
    {"id": "LEAD-IA-0838", "slug": "restaurante-la-marina", "rubro": "Delivery / Gastronomia",
     "local_html": "sites/restaurante-la-marina/index.html", "demo": "delivery-gastronomia"},
    {"id": "LEAD-IA-0002", "slug": "antes-muerta-que-sencilla-showroom", "rubro": "Showrooms",
     "local_html": "previews/Showrooms_e_Indumentaria_(Instagram)/preview_Antes_Muerta_que_Sencilla_Showroom.html",
     "demo": "showroom-indumentaria"},
    {"id": "LEAD-IA-0029", "slug": "bulgarie", "rubro": "Showrooms",
     "local_html": "previews/Showrooms_e_Indumentaria_(Instagram)/preview_Bulgarie.html",
     "demo": "showroom-indumentaria"},
    {"id": "LEAD-IA-1364", "slug": "urgencias-odontologicas", "rubro": "Salud y Estetica",
     "local_html": "previews/Salud_y_Est\u00e9tica___Odontolog\u00eda/preview_Urgencias_Odontologicas.html",
     "demo": "salud-estetica"},
    {"id": "LEAD-IA-0654", "slug": "cabanas-entre-los-arboles", "rubro": "Turismo y Cabanas",
     "local_html": "previews/Turismo_y_Alojamiento___Caba\u00f1as_y_Complejos/preview_Caba\u00f1as_Entre_Los_Arboles.html",
     "demo": "generica-premium"},
]

SCRIPTS_RUBRO = {
    "Salud y Estetica": (
        "Hola {nombre}, \u00bfc\u00f3mo est\u00e1n? Te escribo desde nuestra agencia de tecnolog\u00eda en Mar del Plata. "
        "Estuvimos viendo su local y notamos que toman todos los turnos manualmente por WhatsApp. Creamos un "
        "Agendador Automatizado que permite reservar turnos 24hs, cobrar la se\u00f1a sola y enviar recordatorios "
        "para eliminar el ausentismo. \u00bfTe gustar\u00eda ver un video de 30 segundos de c\u00f3mo funcionar\u00eda?"
    ),
    "Comercio General": (
        "Hola {nombre}, \u00bfc\u00f3mo est\u00e1n? Les escribo desde nuestra agencia de tecnolog\u00eda en Mar del Plata. "
        "Estuvimos viendo su local y notamos que no cuentan con sitio web propio ni un sistema digitalizado de "
        "inventario y caja. Desarrollamos un combo para comercios de MDP: P\u00e1gina Web + Software de Gesti\u00f3n "
        "Comercial (POS, stock, caja). \u00bfTienen 2 minutos para que les env\u00ede un breve demo en video?"
    ),
    "Delivery / Gastronomia": (
        "Hola {nombre}, \u00bfc\u00f3mo va? Te escribo desde nuestra agencia en Mar del Plata. Vemos que tienen muy buena "
        "salida en delivery y quer\u00edamos mostrarles una herramienta para no perder ventas en horas pico: un Men\u00fa "
        "Digital por WhatsApp que toma el pedido completo y lo env\u00eda a cocina sin responder mensaje por mensaje. "
        "\u00bfTienen 2 minutos para que les muestre c\u00f3mo funciona?"
    ),
    "Showrooms": (
        "Hola {nombre}, \u00bfc\u00f3mo andan? Los contacto desde nuestra agencia en Mar del Plata. Vemos que mueven "
        "bastantes consultas por Instagram y quer\u00edamos mostrarles c\u00f3mo automatizar las ventas: un Bot de "
        "WhatsApp 24/7 que muestra cat\u00e1logo, talles/colores en tiempo real y toma el pedido con cobro por "
        "MercadoPago. \u00bfLes gustar\u00eda ver una demo de 1 minuto sin compromiso?"
    ),
    "Turismo y Cabanas": (
        "Hola {nombre}, \u00bfc\u00f3mo est\u00e1n? Te escribo desde nuestra agencia en Mar del Plata. Vemos que reciben "
        "muchas consultas de reservas y quer\u00edamos mostrarles c\u00f3mo automatizarlas: una web con disponibilidad en "
        "vivo y reservas directas por WhatsApp con se\u00f1a autom\u00e1tica. \u00bfTe interesar\u00eda ver una prueba r\u00e1pida?"
    ),
}

ESTADOS = ["Pendiente", "Enviado", "Respondi\u00f3", "Demo enviada", "Interesado", "Cerrado", "No interesado"]


def normalize_sector(s):
    if not s:
        return "Comercio General"
    s = s.strip()
    if "Salud" in s:
        return "Salud y Estetica"
    if "Showroom" in s or "Indumentaria" in s:
        return "Showrooms"
    if "Comida" in s or "Delivery" in s or "Gastronom" in s:
        return "Delivery / Gastronomia"
    if "Turismo" in s or "Caba" in s or "Hotel" in s:
        return "Turismo y Cabanas"
    return "Comercio General"


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["\U0001F4CA Todos los Prospectos"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {h: i for i, h in enumerate(hdr[:17])}
    by_id = {}
    for r in rows[1:]:
        if r[idx["id_lead"]]:
            by_id[r[idx["id_lead"]]] = r

    # scripts personalizados (solapa primeros 100)
    ws2 = wb["\U0001F680 PRIMEROS 100 LEADS"]
    rows2 = list(ws2.iter_rows(values_only=True))
    hdr2 = rows2[0]
    idx2 = {h: i for i, h in enumerate(hdr2)}
    scripts = {}
    for r in rows2[1:]:
        if r[idx2["id_lead"]]:
            scripts[r[idx2["id_lead"]]] = r[idx2["script_mensajeria_personalizado"]]

    out = []
    for lead in LEADS:
        r = by_id.get(lead["id"])
        nombre = r[idx["nombre"]] if r else lead["slug"].replace("-", " ").title()
        telefono = r[idx["telefono"]] if r else ""
        direccion = r[idx["direccion"]] if r and idx["direccion"] < len(r) else None
        score = r[idx["puntaje_oportunidad"]] if r else 65
        wa = r[idx["link_whatsapp_outreach"]] if r else ""
        script = scripts.get(lead["id"])
        if not script:
            template = SCRIPTS_RUBRO.get(lead["rubro"], SCRIPTS_RUBRO["Comercio General"])
            script = template.format(nombre=nombre)
        lead["nombre"] = nombre
        lead["telefono"] = telefono or ""
        lead["direccion"] = direccion or ""
        lead["score"] = int(score or 65)
        lead["wa_link"] = wa
        lead["script"] = script
        out.append(lead)

    # orden: score desc, luego rubro
    out.sort(key=lambda l: (-l["score"], l["rubro"], l["nombre"]))

    data_json = json.dumps(out, ensure_ascii=False)
    est_js = json.dumps(ESTADOS, ensure_ascii=False)

    html = TEMPLATE.replace("__DATA_JSON__", data_json).replace("__ESTADOS_JSON__", est_js)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(html, encoding="utf-8")
    print(f"Dashboard generado: {OUT} ({len(out)} leads)")


TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Top 20 MDP &mdash; Dashboard de Prospecci&oacute;n</title>
<style>
  :root{
    --bg:#0b0f1a; --panel:#121828; --panel2:#0f1422; --line:#1e2740;
    --txt:#e8ecf8; --muted:#8b95b5; --accent:#4f7cff; --accent2:#7c4fff;
    --green:#2fd07a; --amber:#f5a623; --red:#f05d5d;
    --wa:#25d366; --salud:#4f7cff; --comercio:#f5a623; --delivery:#ff6b4a;
    --showroom:#e6459e; --turismo:#2fd07a;
  }
  *{margin:0;padding:0;box-sizing:border-box}
  body{background:var(--bg);color:var(--txt);font-family:'Segoe UI',system-ui,sans-serif;min-height:100vh}
  .wrap{max-width:1180px;margin:0 auto;padding:28px 20px 80px}
  header.hero{background:linear-gradient(135deg,#101b36 0%,#171033 55%,#231044 100%);border:1px solid var(--line);border-radius:20px;padding:32px 28px;position:relative;overflow:hidden;margin-bottom:22px}
  header.hero::after{content:"";position:absolute;inset:0;background:radial-gradient(600px 200px at 85% -10%,rgba(124,79,255,.35),transparent),radial-gradient(500px 180px at 10% 110%,rgba(79,124,255,.28),transparent);pointer-events:none}
  .hero h1{font-size:clamp(22px,3.4vw,34px);letter-spacing:.3px;position:relative}
  .hero p{color:var(--muted);margin-top:8px;font-size:14px;position:relative}
  .hero .chips{display:flex;gap:8px;flex-wrap:wrap;margin-top:16px;position:relative}
  .chip{font-size:11.5px;padding:5px 12px;border-radius:999px;border:1px solid var(--line);background:rgba(255,255,255,.04);color:var(--muted)}
  .chip b{color:var(--txt)}
  .bar{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin:18px 0 14px;position:sticky;top:10px;z-index:40;background:rgba(11,15,26,.92);backdrop-filter:blur(8px);padding:10px 12px;border:1px solid var(--line);border-radius:14px}
  .bar input{flex:1;min-width:160px;background:var(--panel);border:1px solid var(--line);color:var(--txt);border-radius:10px;padding:9px 12px;font-size:13px;outline:none}
  .bar input:focus{border-color:var(--accent)}
  .pill{cursor:pointer;border:1px solid var(--line);background:var(--panel);color:var(--muted);padding:7px 13px;border-radius:999px;font-size:12px;transition:.15s}
  .pill:hover{color:var(--txt);border-color:var(--accent)}
  .pill.on{background:var(--accent);border-color:var(--accent);color:#fff;font-weight:600}
  .prog{display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:10px;margin-bottom:22px}
  .prog-card{background:var(--panel);border:1px solid var(--line);border-radius:14px;padding:12px 14px}
  .prog-card .num{font-size:22px;font-weight:700}
  .prog-card .lbl{font-size:11.5px;color:var(--muted);margin-top:2px}
  .grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(345px,1fr));gap:14px}
  .card{background:var(--panel);border:1px solid var(--line);border-radius:16px;padding:18px;display:flex;flex-direction:column;gap:10px;transition:.2s;position:relative;overflow:hidden}
  .card:hover{transform:translateY(-2px);border-color:#2c3a63;box-shadow:0 10px 30px rgba(0,0,0,.35)}
  .card::before{content:"";position:absolute;top:0;left:0;right:0;height:3px;background:var(--rc,var(--accent))}
  .card .top{display:flex;justify-content:space-between;align-items:flex-start;gap:8px}
  .card h3{font-size:16px;line-height:1.25}
  .score{flex-shrink:0;font-size:13px;font-weight:700;padding:4px 10px;border-radius:8px;background:rgba(47,208,122,.12);color:var(--green);border:1px solid rgba(47,208,122,.35)}
  .score.low{background:rgba(245,166,35,.12);color:var(--amber);border-color:rgba(245,166,35,.35)}
  .meta{font-size:12.5px;color:var(--muted);display:flex;flex-direction:column;gap:3px}
  .rubro{display:inline-block;font-size:11px;font-weight:600;padding:3px 10px;border-radius:999px;color:#fff;background:var(--rc,var(--accent));width:fit-content}
  .acts{display:flex;gap:8px;flex-wrap:wrap;margin-top:2px}
  .btn{display:inline-flex;align-items:center;gap:6px;font-size:12.5px;font-weight:600;padding:9px 13px;border-radius:10px;text-decoration:none;transition:.15s;border:1px solid transparent}
  .btn.wa{background:var(--wa);color:#04331a}
  .btn.wa:hover{filter:brightness(1.08)}
  .btn.ghost{background:transparent;color:var(--txt);border-color:var(--line)}
  .btn.ghost:hover{border-color:var(--accent)}
  .btn.dark{background:#1b2440;color:var(--txt);border-color:#26325a}
  .btn.dark:hover{border-color:var(--accent2)}
  .track{border-top:1px dashed var(--line);padding-top:10px;display:flex;gap:8px;align-items:center;flex-wrap:wrap}
  .track select{background:var(--panel2);border:1px solid var(--line);color:var(--txt);border-radius:8px;padding:6px 8px;font-size:12px;outline:none;flex:1;min-width:110px}
  .track input{flex:1.4;min-width:120px;background:var(--panel2);border:1px solid var(--line);color:var(--txt);border-radius:8px;padding:6px 8px;font-size:12px;outline:none}
  .track select:focus,.track input:focus{border-color:var(--accent)}
  section.block{background:var(--panel);border:1px solid var(--line);border-radius:16px;padding:22px;margin-top:26px}
  section.block h2{font-size:17px;margin-bottom:14px}
  .plan-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:14px}
  .plan-item{background:var(--panel2);border:1px solid var(--line);border-radius:12px;padding:14px}
  .plan-item h4{font-size:13px;color:var(--accent);margin-bottom:8px}
  .plan-item p,.plan-item li{font-size:12.5px;color:var(--muted);line-height:1.55}
  .plan-item ul{padding-left:16px}
  .day{font-size:12.5px;background:var(--panel2);border:1px solid var(--line);border-radius:10px;padding:10px 14px;margin-bottom:8px}
  .day b{color:var(--txt)}
  .script-block{background:var(--panel2);border:1px solid var(--line);border-left:3px solid var(--accent);border-radius:10px;padding:12px 14px;font-size:12.5px;color:var(--muted);line-height:1.6;white-space:pre-wrap;margin-bottom:10px}
  .tabs{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px}
  .tab{cursor:pointer;padding:7px 14px;border-radius:9px;border:1px solid var(--line);background:var(--panel2);color:var(--muted);font-size:12.5px}
  .tab.on{background:var(--accent);color:#fff;border-color:var(--accent)}
  .tabpane{display:none}
  .tabpane.on{display:block}
  footer{margin-top:30px;text-align:center;color:var(--muted);font-size:12px}
  .badge-inline{font-size:11px;color:var(--muted)}
  details{background:var(--panel2);border:1px solid var(--line);border-radius:10px;padding:0;margin-bottom:8px}
  summary{cursor:pointer;padding:11px 14px;font-size:13px;font-weight:600;color:var(--txt);list-style:none;display:flex;justify-content:space-between;align-items:center}
  summary::after{content:"+";color:var(--muted);font-size:15px}
  details[open] summary::after{content:"\2212"}
  details .inner{padding:0 14px 12px;font-size:12.5px;color:var(--muted);line-height:1.6;white-space:pre-wrap}
</style>
</head>
<body>
<div class="wrap">

  <header class="hero">
    <h1>&#128202; Top 20 MDP &mdash; Dashboard de Prospecci&oacute;n</h1>
    <p>20 locales priorizados de Mar del Plata &middot; 6 rubros &middot; Todos con web, preview o demo construida &middot; WhatsApp precargado con mensaje personalizado</p>
    <div class="chips">
      <span class="chip">&#129534; <b>12</b> leads score 80</span>
      <span class="chip">&#128187; <b>16</b> con web completa</span>
      <span class="chip">&#128274; <b>4</b> con preview</span>
      <span class="chip">&#128221; Plan de env&iacute;o 2 semanas</span>
    </div>
  </header>

  <div class="bar">
    <input id="search" type="text" placeholder="Buscar local o direcci&oacute;n...">
    <button class="pill on" data-rubro="">Todos</button>
    <button class="pill" data-rubro="Salud y Estetica">Salud</button>
    <button class="pill" data-rubro="Comercio General">Comercio</button>
    <button class="pill" data-rubro="Delivery / Gastronomia">Delivery</button>
    <button class="pill" data-rubro="Showrooms">Showrooms</button>
    <button class="pill" data-rubro="Turismo y Cabanas">Turismo</button>
  </div>

  <div class="prog" id="prog"></div>
  <div class="grid" id="grid"></div>

  <section class="block">
    <h2>&#128640; Plan de Mensajer&iacute;a &mdash; 2 semanas</h2>
    <div class="tabs" id="planTabs">
      <button class="tab on" data-tab="cadencia">Cadencia diaria</button>
      <button class="tab" data-tab="scripts">Scripts por rubro</button>
      <button class="tab" data-tab="reglas">Reglas anti-bloqueo</button>
      <button class="tab" data-tab="seguimiento">Seguimiento &amp; cierre</button>
    </div>

    <div class="tabpane on" id="pane-cadencia">
      <div class="day"><b>D&iacute;as 1-2 &mdash; Ronda 1 (20 mensajes):</b><br>10 mensajes/d&iacute;a en dos tandas: 09:30-11:30 (5) y 15:30-17:30 (5), con intervalo de 45-60 segundos entre env&iacute;os. Empezar por los score 80 (Salud + La Marina + Urgencias Odontol&oacute;gicas + Caba&ntilde;as).</div>
      <div class="day"><b>D&iacute;a 3 &mdash; Atenci&oacute;n inmediata:</b><br>Responder todos los interesados. A quien responda, mandarle el <b>HTML personalizado ya construido</b> (bot&oacute;n "Ver web") + video demo. Agendar llamada si pide m&aacute;s info.</div>
      <div class="day"><b>D&iacute;as 4-5 &mdash; Re-contacto 48h:</b><br>A los que no respondieron, enviar el script de re-contacto (distinto al primero, ver "Seguimiento"). M&aacute;ximo 1 re-contacto por lead.</div>
      <div class="day"><b>Semana 2 &mdash; Escalada telef&oacute;nica:</b><br>Llamar a los que no contestaron WhatsApp (tel en la tarjeta). A los "Interesado" sin cerrar: ofrecer demo en vivo de 15 min.</div>
    </div>

    <div class="tabpane" id="pane-scripts"></div>

    <div class="tabpane" id="pane-reglas">
      <div class="plan-grid">
        <div class="plan-item"><h4>Horarios seguros</h4><p>Comercios: 10:00-12:00 / 16:00-18:00. Salud y consultorios: 09:30-11:30 / 15:00-17:00. Evitar lunes temprano, fines de semana y horario de almuerzo (12:30-14:30).</p></div>
        <div class="plan-item"><h4>Volumen</h4><p>M&aacute;ximo 10-15 mensajes/d&iacute;a con n&uacute;mero nuevo. No re-enviar si no hubo respuesta (solo 1 re-contacto). Borrar mensajes fallidos del historial.</p></div>
        <div class="plan-item"><h4>Interacci&oacute;n previa</h4><p>Antes de cada tanda, interactuar 3-5 min en WhatsApp/Instagram (estados, historias) para "calentar" el n&uacute;mero. Nunca mandar links en el primer mensaje.</p></div>
        <div class="plan-item"><h4>Identificaci&oacute;n</h4><p>Siempre presentarse como agencia de Mar del Plata, mencionar la direcci&oacute;n del local (da cercan&iacute;a y prueba de que fue visto) y pedir permiso para enviar demo.</p></div>
      </div>
    </div>

    <div class="tabpane" id="pane-seguimiento">
      <div class="plan-grid">
        <div class="plan-item"><h4>Re-contacto 48h (no respondieron)</h4><p>"Hola {nombre}, te escrib&iacute; hace unos d&iacute;as y vi que qued&oacute; pendiente. Te dejo el dato: preparamos una demo corta del sistema para tu rubro. &iquest;Te la mando por ac&aacute;?"</p></div>
        <div class="plan-item"><h4>Al que respondi&oacute;</h4><p>Enviar el HTML personalizado ya construido + 1 video demo de 30 seg. Preguntar: "&iquest;Te gustar&iacute;a que lo personalicemos con tu logo y productos?"</p></div>
        <div class="plan-item"><h4>Cierre</h4><p>50% se&ntilde;a para arrancar &middot; Demo funcional en 5 d&iacute;as &middot; 50% al activar. Combo 360&deg;: Web + Bot + Gesti&oacute;n desde $340.000 ARS.</p></div>
        <div class="plan-item"><h4>Registro</h4><p>Actualizar el estado de cada tarjeta arriba (se guarda en el navegador). Criterios: "Interesado" = pidi&oacute; m&aacute;s info; "Cerrado" = pag&oacute; se&ntilde;a.</p></div>
      </div>
    </div>
  </section>

  <footer>Dashboard de prospecci&oacute;n &mdash; Generado desde PROSPECCION_MAR_DEL_PLATA_COMPLETA.xlsx &middot; Estados guardados localmente en tu navegador</footer>
</div>

<script>
const DATA = __DATA_JSON__;
const ESTADOS = __ESTADOS_JSON__;
const LS_KEY = "top20_mdp_estados_v1";
let filtro = "";
let estados = {};
try { estados = JSON.parse(localStorage.getItem(LS_KEY) || "{}"); } catch(e) { estados = {}; }

const RUBRO_COLOR = {
  "Salud y Estetica":"var(--salud)","Comercio General":"var(--comercio)",
  "Delivery / Gastronomia":"var(--delivery)","Showrooms":"var(--showroom)","Turismo y Cabanas":"var(--turismo)"
};

function guardar(){
  localStorage.setItem(LS_KEY, JSON.stringify(estados));
  renderProg();
}

function renderProg(){
  const conteo = {};
  ESTADOS.forEach(e => conteo[e] = 0);
  DATA.forEach(l => {
    const est = (estados[l.id] || {}).estado || "Pendiente";
    conteo[est]++;
  });
  document.getElementById("prog").innerHTML = ESTADOS.map(e =>
    `<div class="prog-card"><div class="num">${conteo[e]}</div><div class="lbl">${e}</div></div>`
  ).join("");
}

function render(){
  const q = document.getElementById("search").value.toLowerCase();
  const grid = document.getElementById("grid");
  const lista = DATA.filter(l =>
    (!filtro || l.rubro === filtro) &&
    (!q || (l.nombre + " " + (l.direccion||"") + " " + l.telefono).toLowerCase().includes(q))
  );
  if(!lista.length){ grid.innerHTML = '<p style="color:var(--muted);grid-column:1/-1">Sin resultados.</p>'; return; }
  grid.innerHTML = lista.map(l => {
    const est = (estados[l.id] || {}).estado || "Pendiente";
    const nota = (estados[l.id] || {}).nota || "";
    const scoreCls = l.score >= 80 ? "" : "low";
    const rc = RUBRO_COLOR[l.rubro] || "var(--accent)";
    const local = encodeURIComponent(l.local_html);
    return `<div class="card" style="--rc:${rc}">
      <div class="top"><h3>${l.nombre}</h3><span class="score ${scoreCls}">${l.score}</span></div>
      <span class="rubro" style="background:${rc}">${l.rubro}</span>
      <div class="meta"><span>&#128205; ${l.direccion || "&mdash;"}</span><span>&#128222; ${l.telefono || "&mdash;"}</span></div>
      <div class="acts">
        <a class="btn wa" target="_blank" rel="noopener" href="${l.wa_link}">&#128172; WhatsApp</a>
        <a class="btn ghost" href="../${local}" target="_blank" rel="noopener" title="Abrir HTML ya construido">&#128187; Ver web</a>
        <a class="btn dark" href="../demos/${l.demo}/index.html" target="_blank" rel="noopener">&#127912; Demo gen&eacute;rica</a>
      </div>
      <div class="track">
        <select onchange="setEst('${l.id}', this.value)">
          ${ESTADOS.map(e => `<option ${e===est?"selected":""}>${e}</option>`).join("")}
        </select>
        <input placeholder="Nota (ej: llamar 15hs)" value="${nota.replace(/"/g,"&quot;")}" onchange="setNota('${l.id}', this.value)">
      </div>
    </div>`;
  }).join("");
}

function setEst(id, v){ estados[id] = estados[id] || {}; estados[id].estado = v; guardar(); }
function setNota(id, v){ estados[id] = estados[id] || {}; estados[id].nota = v; guardar(); }

document.getElementById("search").addEventListener("input", render);
document.querySelectorAll(".pill").forEach(p => p.addEventListener("click", () => {
  filtro = p.dataset.rubro;
  document.querySelectorAll(".pill").forEach(x => x.classList.toggle("on", x === p));
  render();
}));

document.getElementById("planTabs").addEventListener("click", e => {
  const b = e.target.closest(".tab"); if(!b) return;
  document.querySelectorAll(".tab").forEach(x => x.classList.toggle("on", x === b));
  document.querySelectorAll(".tabpane").forEach(x => x.classList.toggle("on", x.id === "pane-" + b.dataset.tab));
});

document.getElementById("pane-scripts").innerHTML = DATA.filter((l,i,arr) =>
  arr.findIndex(x => x.rubro === l.rubro) === i
).map(l => `<details><summary>${l.rubro}</summary><div class="inner">${l.script}</div></details>`).join("");

renderProg();
render();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    main()
