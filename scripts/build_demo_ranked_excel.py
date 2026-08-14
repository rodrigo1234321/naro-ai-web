import os
import sys
import re
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")


def clean_phone_str(val):
    if pd.isna(val) or val is None:
        return ""
    val_s = str(val).strip()
    if not val_s or val_s.lower() == "nan":
        return ""
    val_s = re.sub(r"\.0$", "", val_s)
    if "e+" in val_s.lower():
        try:
            val_s = str(int(float(val_s)))
        except Exception:
            pass
    return val_s


def is_valid_wa(val):
    s = clean_phone_str(val)
    s_digits = re.sub(r"\D", "", s)
    return len(s_digits) >= 10 and s_digits.startswith("549")


def get_demo_info(row):
    nombre = str(row.get("nombre", "")).lower()
    sector = str(row.get("sector", "")).lower()

    if "la marina" in nombre or "marina" in nombre:
        return "SI - Demo Personalizada", "sites/restaurante-la-marina/index.html (Restaurante La Marina)"
    elif "stella maris" in nombre:
        return "SI - Demo Personalizada", "stella-maris-pastas/index.html (Stella Maris Pastas)"
    elif "eufforia" in nombre or "serena" in nombre:
        return "SI - Demo Personalizada", "demos/salud-estetica/index.html (Eufforia Estética - Serena García)"
    elif "sabuesos" in nombre:
        return "SI - Demo Personalizada", "demos/comercio-minorista/index.html (Sabuesos Pet Shop - Dorrego 2662)"
    elif "francesca" in nombre:
        return "SI - Demo Personalizada", "demos/comercio-minorista/index.html (Francesca Uniformes - Dorrego 2752)"
    elif "ms refrigeracion" in nombre or "ms refrigeración" in nombre:
        return "SI - Demo Personalizada", "ms-refrigeracion-web/index.html (MS Refrigeración)"
    elif "salud" in sector or "estética" in sector or "estetica" in sector:
        return "SI - MVP Rubro", "demos/salud-estetica/index.html (Agendador Turnos 24/7 + Señas MP)"
    elif "showroom" in sector or "indumentaria" in sector:
        return "SI - MVP Rubro", "demos/showroom-indumentaria/index.html (Tienda Autónoma + Stock Talle/Color)"
    elif "delivery" in sector or "gastronomía" in sector or "gastronomia" in sector:
        return "SI - MVP Rubro", "demos/delivery-gastronomia/index.html (Menú Digital + Comandera Cocina)"
    elif "automotriz" in sector or "lavadero" in sector:
        return "SI - MVP Rubro", "demos/comercio-minorista/index.html (Web Servicio + POS Taller)"
    else:
        return "SI - MVP Rubro", "demos/comercio-minorista/index.html (Web Catálogo + Software Gestión POS)"


def main():
    data_dir = "data"
    prospects_csv = os.path.join(data_dir, "prospectos_agencia_ia_mdp.csv")

    out_csv = os.path.join(data_dir, "prospectos_seleccionados_con_demo.csv")
    out_excel_data = os.path.join(data_dir, "PROSPECCION_LEADS_CON_DEMO_WHATSAPP.xlsx")
    out_excel_root = "PROSPECCION_LEADS_CON_DEMO_WHATSAPP.xlsx"

    print("Leyendo base de prospectos...")
    df = pd.read_csv(prospects_csv, encoding="utf-8-sig", dtype=str)

    # 1. Filtrar estrictamente SIN WEB
    df["tiene_web_clean"] = df["tiene_web"].fillna("NO").str.upper()
    df_sin_web = df[(df["tiene_web_clean"] == "NO") | (df["web"].isna()) | (df["web"] == "")].copy()

    # 2. Filtrar estrictamente CON WHATSAPP VALIDO (549...)
    df_sin_web["whatsapp_clean"] = df_sin_web["whatsapp"].apply(clean_phone_str)
    df_valid = df_sin_web[df_sin_web["whatsapp_clean"].apply(is_valid_wa)].copy()

    # 3. Asignar Demo HTML
    res = df_valid.apply(get_demo_info, axis=1)
    df_valid["demo_html_tipo"] = [r[0] for r in res]
    df_valid["demo_html_path"] = [r[1] for r in res]

    # 4. Ordenar DE MAYOR A PEOR
    df_valid["puntaje_num"] = pd.to_numeric(df_valid["puntaje_oportunidad"], errors="coerce").fillna(0)

    def calculate_rank(row):
        score = row["puntaje_num"]
        if "Personalizada" in row["demo_html_tipo"]:
            return score + 2000
        elif "ALTA" in str(row.get("prioridad", "")):
            return score + 1000
        elif "MEDIA" in str(row.get("prioridad", "")):
            return score + 500
        return score

    df_valid["rank_score"] = df_valid.apply(calculate_rank, axis=1)
    df_ranked = df_valid.sort_values(by="rank_score", ascending=False).reset_index(drop=True)
    df_ranked["ranking"] = [f"#{i+1}" for i in range(len(df_ranked))]

    cols_order = [
        "ranking",
        "nombre",
        "sector",
        "prioridad",
        "puntaje_oportunidad",
        "demo_html_tipo",
        "demo_html_path",
        "servicio_pitch_sugerido",
        "apto_gestion_comercial",
        "pitch_gestion_comercial",
        "whatsapp",
        "link_whatsapp_outreach",
        "direccion",
        "zona",
        "url_google_maps",
    ]

    # Guardar CSV filtrado
    df_ranked[cols_order].to_csv(out_csv, index=False, encoding="utf-8-sig")
    print(f"[CSV EXITO] Guardado dataset de {len(df_ranked)} prospectos seleccionados con Demo y WA -> {out_csv}")

    # =========================================================
    # GENERAR LIBRO EXCEL MAESTRO NUEVO
    # =========================================================
    print("Creando libro de Excel nuevo: PROSPECCION_LEADS_CON_DEMO_WHATSAPP.xlsx...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Estilos
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_hot_header = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    font_data = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_link = Font(name="Calibri", size=10, color="0563C1", underline="single")

    fill_even = PatternFill(start_color="F9FAFC", end_color="F9FAFC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    def add_df_sheet(sheet_name, dataframe, is_hot=False):
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        # Header
        ws.append(list(dataframe.columns))
        header_fill = fill_hot_header if is_hot else fill_header

        for col_num in range(1, len(dataframe.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[1].height = 26

        # Fill Data
        for row_idx, row_data in enumerate(dataframe.values, start=2):
            ws.append(list(row_data))
            ws.row_dimensions[row_idx].height = 20
            row_fill = fill_even if row_idx % 2 == 0 else fill_odd

            for col_idx in range(1, len(dataframe.columns) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                col_name = dataframe.columns[col_idx - 1]
                c.fill = row_fill
                c.border = thin_border
                c.font = font_data
                c.alignment = align_left

                # Formato texto para whatsapp y teléfono
                if col_name in ["telefono", "whatsapp"]:
                    c.number_format = "@"
                    c.alignment = align_center

                if col_name in ["ranking", "puntaje_oportunidad", "prioridad", "demo_html_tipo"]:
                    c.alignment = align_center
                    if col_name == "ranking":
                        c.font = font_bold

                # Formato Hyperlink
                if col_name in ["link_whatsapp_outreach", "url_google_maps", "web"] and c.value:
                    val_str = str(c.value)
                    if val_str.startswith("http"):
                        c.hyperlink = val_str
                        c.font = font_link

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 55)

    # 1. Pestaña Principal: 🚀 LEADS SELECCIONADOS (DEMO + WA)
    print(f"Generando hoja #1: 🚀 LEADS SELECCIONADOS ({len(df_ranked)} registros)...")
    add_df_sheet("🚀 LEADS SELECCIONADOS", df_ranked[cols_order], is_hot=True)

    # 2. Pestaña: 🌟 CON DEMO PERSONALIZADA
    df_pers = df_ranked[df_ranked["demo_html_tipo"].str.contains("Personalizada", na=False)].copy()
    print(f"Generando hoja #2: 🌟 CON DEMO PERSONALIZADA ({len(df_pers)} registros)...")
    add_df_sheet("🌟 DEMOS PERSONALIZADAS", df_pers[cols_order], is_hot=True)

    # 3. Pestaña: 🔥 TOP 100 CON DEMO & WA
    df_top100 = df_ranked.head(100).copy()
    print(f"Generando hoja #3: 🔥 TOP 100 CON DEMO & WA ({len(df_top100)} registros)...")
    add_df_sheet("🔥 TOP 100 CON DEMO & WA", df_top100, is_hot=True)

    # 4. Pestaña: 🖥️ APTO GESTION COMERCIAL
    df_gestion = df_ranked[df_ranked["apto_gestion_comercial"].str.contains("SI", na=False)].copy()
    print(f"Generando hoja #4: 🖥️ APTO GESTION COMERCIAL ({len(df_gestion)} registros)...")
    add_df_sheet("🖥️ APTO GESTION COMERCIAL", df_gestion[cols_order])

    # Guardar en data/ y en la raíz
    for path in [out_excel_data, out_excel_root]:
        try:
            wb.save(path)
            print(f"[EXCEL GUARDADO] {path}")
        except Exception as e:
            print(f"[WARNING] No se pudo guardar en {path}: {e}")


if __name__ == "__main__":
    main()
