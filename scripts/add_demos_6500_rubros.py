# -*- coding: utf-8 -*-
"""Conecta los 6500 prospectos con el demo del Demo Factory más cercano.

Genera PROSPECCION_6500_RUBROS_DEMOS.xlsx con:
- Una hoja por rubro de demo (28 hojas) + hoja resumen.
- Cada fila: negocio, sector, zona, teléfono, URL (web real si existe,
  si no demo personalizada con ?n=&t=&d=), mensaje de prospección y
  link wa.me con el mensaje precargado.

Uso: python scripts/add_demos_6500_rubros.py
"""
import re
import sys
from pathlib import Path
from urllib.parse import quote

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = ROOT / "PROSPECCION_MAR_DEL_PLATA_COMPLETA.xlsx"
OUT_XLSX = ROOT / "PROSPECCION_6500_RUBROS_DEMOS.xlsx"
SITE = "https://mis-clientes-html.pages.dev"

# slug demo -> (nombre hoja, keywords, rubro para mensaje, emoji)
DEMOS = [
    ("clinica-aura", "Clinica y consultorios", "clinica,consultorio medico,centro medico,salud,medicina,analisis clinico,cardiologia,ginecologia,pediatra,fisiatria,medico", "consultorio medico", "🩺"),
    ("dental-sonrisa", "Odontologia", "dental,odontologia,dentista,ortodoncia,odontologico,clinica dental", "consultorio odontologico", "🦷"),
    ("estetica-lumiere", "Estetica y belleza", "estetica,depilacion,laser,belleza,cosmiatra,cosmetologia,spa,masaje estetico,manicuria,cejas", "centro de estetica", "✨"),
    ("peluqueria-ambar", "Peluqueria y barberia", "peluqueria,peluquero,barberia,barbero,colorimetria,estilista,unisex", "peluqueria", "💇"),
    ("kinesio-movere", "Kinesiologia y salud", "kinesio,fisioterapia,rehabilitacion,traumatologia,osteopatia,massoterapia,kinesiologo", "centro de kinesiologia", "🦵"),
    ("restaurante-rias", "Restaurantes", "restaurante,parrilla,bistro,comida de mar,marisqueria,gastronomia,comida gourmet,chef,menu,sushi,minutas,bodegon,tenedor libre", "restaurante", "🍽️"),
    ("cafe-verde-alba", "Cafeterias", "cafe,cafeteria,coffee,confiteria,brunch,panaderia,pasteleria,factura,tortas,reposteria", "cafeteria", "☕"),
    ("viandas-sabores", "Viandas y saludable", "vianda,viandas,catering,comida casera,dietetica,saludable,vegano,vegetariano,sin tacc,verduleria,fruteria,carniceria,pescaderia,frutas,verduras", "viandas", "🥗"),
    ("rotiseria-don-gino", "Rotiserias y delivery", "rotiseria,pollo a la spiedo,empanada,pizza,pastas,take away,heladeria,hamburgueseria,burger,parripollo,sandwicheria,lomiteria,pizzeria,delivery", "rotiseria", "🍗"),
    ("cerveceria-punto-cebada", "Cervecerias y bares", "cerveceria,cerveza,beer bar,bar,brewery,chopp", "cerveceria", "🍺"),
    ("vinoteca-cava-puerto", "Vinotecas", "vinoteca,vino,licor,espirituosa,whisky,fernet", "vinoteca", "🍷"),
    ("showroom-nube", "Showrooms e indumentaria", "showroom,indumentaria,ropa,moda,vestimenta,lenceria,bikini,outlet,feria americana,marcas", "showroom de indumentaria", "👗"),
    ("sport-base9", "Deportes", "deporte,deportivo,camiseta,sport,indumentaria deportiva,calzado deportivo,planet fitness,gimnasio", "local deportivo", "⚽"),
    ("calzado-paso-norte", "Calzado", "calzado,zapatilla,zapateria,zapatos,zapatillas,botines", "local de calzado", "👟"),
    ("inmobiliaria-costa-real", "Inmobiliarias", "inmobiliaria,propiedades,bienes raices,real estate,corretaje,venta de propiedades", "inmobiliaria", "🏠"),
    ("temporarios-dunas", "Alquileres temporarios", "temporario,temporada,alquiler temporario,apart temporario,departamento temporario", "alquiler temporario", "🏖️"),
    ("cabanas-aires-faro", "Cabanas y complejos", "cabana,cabinas,complejo turistico,camping,cabana en la sierra,cabana a la montana", "cabanas", "🌲"),
    ("hotel-olas-sur", "Hoteles y posadas", "hotel,apart hotel,hostel,posada,hospedaje,hosteria,boutique", "hotel", "🏨"),
    ("lavadero-aquashine", "Lavaderos de autos", "lavadero,lavado de autos,car wash,detailing,encerado,pulido de auto", "lavadero de autos", "🚿"),
    ("gomeria-rodado-sur", "Gomerias", "gomeria,neumatico,gomas,cubiertas,balanceo,alineacion", "gomeria", "🛞"),
    ("taller-motorbox", "Talleres y repuestos", "taller mecanico,mecanica,repuestos,autopartes,accesorios de auto,chapa y pintura,service,escaner,automotriz,hidraulico", "taller automotriz", "🔧"),
    ("regaleria-dulce-detalle", "Regaleria", "regaleria,regalos,peluche,souvenir,cotillon,gift shop,articulos de regalo,jugueteria,joyeria,bijouterie,accesorios", "regaleria", "🎁"),
    ("imprenta-estampa", "Imprentas y graficas", "imprenta,grafica,graf,estampado,sublimacion,ploteo,gigantografia,rotulacion,identidad visual,libreria,articulos de libreria", "imprenta", "🖨️"),
    ("distribuidora-mdp", "Distribuidoras y mayorista", "distribuidora,mayorista,wholesale,deposito,bebidas,mercaderia,abarrotes,minimercado,almacen,supermercado,autoservicio,kiosco", "distribuidora", "📦"),
    ("ferreteria-ferretodo", "Ferreterias y pintureria", "ferreteria,ferretero,pintureria,pintura,herramienta,sanitarios,electricidad,materiales,iluminacion,vidrieria,cerrajeria,construccion", "ferreteria", "🔨"),
    ("petshop-patitas", "Pet shops y veterinarias", "pet shop,pet,perro,gato,mascota,mascoteria,veterinaria,veterinario,balanceado,acuario", "pet shop", "🐾"),
    ("contable-conta-co", "Estudios contables", "contable,contador,estudio contable,impuestos,monotributo,liquidacion,administracion contable,tributario", "estudio contable", "📊"),
    ("flores-jardin-puerto", "Florerias", "flores,floreria,floristeria,decoracion floral", "floreria", "🌸"),
]

# Fallback: sector del Excel -> demo por defecto
SECTOR_FALLBACK = {
    "Comercio General": "showroom-nube",
    "Salud y Estética": "clinica-aura",
    "Showrooms e Indumentaria (Instagram)": "showroom-nube",
    "Comida a Domicilio y Delivery": "viandas-sabores",
    "Inmobiliarias": "inmobiliaria-costa-real",
    "Turismo y Alojamiento": "cabanas-aires-faro",
    "Industrial y Puerto": "distribuidora-mdp",
    "Gastronomía": "restaurante-rias",
    "Automotriz y Servicios": "taller-motorbox",
    "Servicios Profesionales": "contable-conta-co",
}

MSG = (
    "Hola {nombre}! Te escribo de Naro AI, hacemos webs para negocios de Mar del Plata \U0001F30A\n"
    "Vi tu {rubro} en {zona} y te prepare una demo en vivo de como se veria tu web, con tu marca y tu numero de WhatsApp:\n"
    "{url}\n"
    "Abrela en el celular y fijate que todo funciona: boton de WhatsApp, direccion y servicios.\n"
    "La activamos hoy? Tiene dominio propio y la podes actualizar cuando quieras. Cualquier cosa me escribis por aca \U0001F64C"
)

HEADERS = ["#", "Nombre", "Sector", "Zona", "Prioridad", "Teléfono", "Tiene Web", "URL (Web o Demo)", "Mensaje WhatsApp", "Link WhatsApp"]


def norm(s):
    return (
        s.lower().strip()
        .replace("á", "a").replace("é", "e").replace("í", "i")
        .replace("ó", "o").replace("ú", "u").replace("ñ", "n")
    )


def norm_tel(raw) -> str:
    d = re.sub(r"\D", "", str(raw or ""))
    if not d:
        return ""
    d = d.lstrip("0")
    if not d.startswith("54"):
        d = "54" + d
    return d


def main():
    print("Leyendo webs construidas...")
    wbs = openpyxl.load_workbook(SRC_XLSX, read_only=True, data_only=True)
    ws_webs = wbs["🌐 Webs Construidas"]
    webs_norm = {}
    for r in ws_webs.iter_rows(min_row=2, values_only=True):
        if r[2] and r[7]:
            webs_norm[norm(str(r[2]))] = str(r[7])

    print("Leyendo los 6500 prospectos...")
    ws = wbs["📊 Todos los Prospectos"]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows))
    idx = {h: i for i, h in enumerate(hdr)}

    rubro_by_demo = {slug: (sheet, keywords, frase, emoji) for slug, sheet, keywords, frase, emoji in DEMOS}
    # rules ordenadas por longitud de keyword (mas especificas primero)
    rules = []
    for slug, sheet, keywords, frase, emoji in DEMOS:
        for k in keywords.split(","):
            if k:
                rules.append((k, slug, frase))
    rules.sort(key=lambda r: -len(r[0]))

    data = {}  # slug -> lista de filas
    sin_demo = 0
    for n, r in enumerate(rows, 2):
        nombre = r[idx["nombre"]]
        if not nombre:
            continue
        nombre = str(nombre).strip()
        sector = str(r[idx["sector"]] or "").strip()
        zona = str(r[idx["zona"]] or "").strip()
        prioridad = str(r[idx["prioridad"]] or "").strip()
        tel = norm_tel(r[idx["telefono"]])
        tiene_web = str(r[idx["tiene_web"]] or "").strip()
        url = ""
        tipo = ""

        low = norm(nombre)
        if low in webs_norm:
            url = webs_norm[low]
            tipo = "WEB REAL"
            slug_dest = "WEB_REAL"
        elif low:
            match = next((m for m in rules if m[0] in low), None)
            if match:
                kw, slug, frase = match
                slug_dest = slug
                url = f"{SITE}/demos/{slug}.html?n={quote(nombre)}&t={tel}&d={quote(nombre)}"
                tipo = "DEMO"
            else:
                slug_fb = SECTOR_FALLBACK.get(sector)
                if slug_fb:
                    slug_dest = slug_fb
                    url = f"{SITE}/demos/{slug_fb}.html?n={quote(nombre)}&t={tel}&d={quote(nombre)}"
                    frase = rubro_by_demo[slug_fb][2]
                    tipo = "DEMO (genérico)"
                else:
                    sin_demo += 1
                    continue
        else:
            continue

        zona_msg = zona if zona else "la zona"
        mensaje = MSG.format(nombre=nombre, rubro=frase if tipo != "WEB REAL" else "negocio", zona=zona_msg, url=url)
        link = f"https://wa.me/{tel}?text={quote(mensaje)}" if tel else ""

        data.setdefault(slug_dest, []).append(
            (n, nombre, sector, zona, prioridad, tel, tiene_web, url, mensaje, link, frase, emoji)
        )
        if n % 1000 == 0:
            print(f"  {n} filas...")

    print("Escribiendo Excel...")
    wbo = openpyxl.Workbook()
    ws_res = wbo.active
    ws_res.title = "📊 Resumen"
    ws_res.append(["Rubro (Demo)", "Negocios", "Con web real", "Con demo"])
    for slug, sheet, keywords, frase, emoji in DEMOS:
        filas = data.get(slug, [])
        reales = [f for f in data.get("WEB_REAL", []) if f[10] == frase]
        ws_res.append([f"{emoji} {sheet}", len(filas), len(reales), len(filas) - len(reales)])
    ws_res.append(["WEB REAL (ya construida)", len(data.get("WEB_REAL", []))])

    for slug, sheet, keywords, frase, emoji in DEMOS:
        filas = data.get(slug, [])
        if not filas:
            continue
        wsn = wbo.create_sheet(f"{emoji} {sheet}")
        wsn.append(HEADERS)
        for f in filas:
            wsn.append(f[:10])

    if data.get("WEB_REAL"):
        wsn = wbo.create_sheet("🌐 Webs Reales")
        wsn.append(HEADERS)
        for f in data["WEB_REAL"]:
            wsn.append(f[:10])

    # Escribir tambien la columna de demo en el Excel original? no: salida separada
    wbo.save(OUT_XLSX)
    print(f"OK -> {OUT_XLSX.name} | {sum(len(v) for v in data.values())} negocios | sin demo: {sin_demo}")
    for slug, sheet, keywords, frase, emoji in DEMOS:
        filas = data.get(slug, [])
        if filas:
            print(f"  {emoji} {sheet}: {len(filas)}")


if __name__ == "__main__":
    sys.exit(main())
