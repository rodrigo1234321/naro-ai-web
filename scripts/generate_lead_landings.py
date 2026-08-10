# -*- coding: utf-8 -*-
"""
Generador de landings demo para prospectos de Naro AI (Mar del Plata).
Uso: python scripts/generate_lead_landings.py [--ids LEAD-IA-1313,LEAD-IA-1301] [--all-first-100]
Sector Salud y Estetica por ahora. Datos reales desde el Excel; copy de demo configurable en OVERRIDES.
"""
import openpyxl, re, sys, os, unicodedata, urllib.parse, json, argparse

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "PROSPECCION_MAR_DEL_PLATA_COMPLETA.xlsx")
OUT_ROOT = os.path.join(BASE, "webs", "leads")

# ---------------------------------------------------------------- helpers

def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s

def wa_link(num, text):
    return "https://wa.me/%s?text=%s" % (num, urllib.parse.quote(text))

def maps_geo(url):
    m = re.search(r"!3d(-?[\d.]+)!4d(-?[\d.]+)", url or "")
    return (m.group(1), m.group(2)) if m else ("", "")

def load_rows(max_row=None):
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = None
    for name in wb.sheetnames:
        if "100" in name:
            ws = wb[name]
            break
    assert ws is not None, "Solapa de primeros 100 no encontrada"
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if max_row and len(rows) >= max_row:
            break
        rows.append(dict(zip(["id","nombre","sector","zona","prioridad","puntaje","servicio_pitch","apto_gestion",
                              "pitch_gestion","tiene_web","web","telefono","whatsapp","link_outreach","direccion",
                              "razones","maps","puntaje_num","angulo","script"], r)))
    return rows

def phone_display(t):
    t = (t or "").strip()
    if re.match(r"^\d{4}\s?\d{6,7}$", t):
        t = re.sub(r"\s+", "", t)
        return "%s %s-%s" % (t[:4], t[4:7], t[7:])
    return t

# ---------------------------------------------------------------- iconos SVG

ICONS = {
"heart":   '<path d="M20.8 4.6a5.5 5.5 0 0 0-7.8 0L12 5.7l-1-1.1a5.5 5.5 0 0 0-7.8 7.8l1 1L12 21l7.8-7.6 1-1a5.5 5.5 0 0 0 0-7.8z"/>',
"activity":'<path d="M22 12h-4l-3 9L9 3l-3 9H2"/>',
"user":    '<path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/>',
"droplet": '<path d="M12 2.7S6 9 6 14a6 6 0 0 0 12 0c0-5-6-11.3-6-11.3z"/>',
"brain":   '<path d="M9.5 2A2.5 2.5 0 0 1 12 4.5v15a2.5 2.5 0 0 1-5 0V9.5"/><path d="M14.5 2A2.5 2.5 0 0 0 12 4.5v15a2.5 2.5 0 0 0 5 0V9.5"/><path d="M2 9.5h2M20 9.5h2M2 14h2M20 14h2M2 12h2M20 12h2"/>',
"leaf":    '<path d="M11 20A7 7 0 0 1 9.8 6.1C15.5 5 17 4.5 19 2c1 2 2 4.2 2 8 0 5.5-4.8 10-10 10z"/><path d="M2 21c0-3 1.9-5.6 4.5-6.9"/>',
"baby":    '<path d="M12 4a2 2 0 1 0 0-4 2 2 0 0 0 0 4z"/><path d="M9 8h6l1 8H8l1-8z"/><path d="M12 8v4M8 14h8"/>',
"clipboard":'<rect x="5" y="3" width="14" height="18" rx="2"/><path d="M9 3a2 2 0 0 1 6 0M9 12h6M9 16h4"/>',
"shield":  '<path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/>',
"clock":   '<circle cx="12" cy="12" r="9"/><path d="M12 7v5l3 2"/>',
"map-pin": '<path d="M20 10c0 6-8 12-8 12S4 16 4 10a8 8 0 1 1 16 0z"/><circle cx="12" cy="10" r="3"/>',
"phone":   '<path d="M22 16.9v3a2 2 0 0 1-2.2 2 19.8 19.8 0 0 1-8.6-3.1 19.5 19.5 0 0 1-6-6A19.8 19.8 0 0 1 2.1 4.2 2 2 0 0 1 4.1 2h3a2 2 0 0 1 2 1.7c.1 1 .4 2 .7 2.9a2 2 0 0 1-.5 2.1L8.1 9.9a16 16 0 0 0 6 6l1.2-1.2a2 2 0 0 1 2.1-.5c.9.3 1.9.6 2.9.7a2 2 0 0 1 1.7 2z"/>',
"check":   '<path d="M20 6L9 17l-5-5"/>',
"cross":   '<path d="M9 3h6v6h6v6h-6v6H9v-6H3V9h6V3z"/>',
"steth":   '<path d="M4 3v6a6 6 0 0 0 12 0V3"/><circle cx="4" cy="3" r="1"/><circle cx="16" cy="3" r="1"/><path d="M14 21a5 5 0 0 0 5-5v-3a3 3 0 0 1 6 0"/>',
"eye":     '<path d="M1 12s4-7 11-7 11 7 11 7-4 7-11 7S1 12 1 12z"/><circle cx="12" cy="12" r="3"/>',
"bone":    '<path d="M17 10c.7-.7 1.7-1 2.7-.8 1.4.3 2.4 1.6 2.3 3-.2 1.4-1.3 2.5-2.7 2.8"/><path d="M7 14c-.7.7-1.7 1-2.7.8-1.4-.3-2.4-1.6-2.3-3 .2-1.4 1.3-2.5 2.7-2.8"/><path d="M10 4c.7-.7 1-1.7.8-2.7A2.9 2.9 0 0 0 7.8-1c-1.4.2-2.5 1.3-2.8 2.7"/><path d="M14 20c-.7.7-1 1.7-.8 2.7.3 1.4 1.6 2.4 3 2.3 1.4-.2 2.5-1.3 2.8-2.7"/>',
"tooth":   '<path d="M12 5.5C10.5 3.5 8 3 6.5 4.5c-1.5 1.5-.5 4-1.5 6.5-.8 2-3 2-3 5 0 2 1.5 3.5 3 3.5 1 0 1.5-.5 2-2 .5-1.5 1.5-2.5 3-2.5s2.5 1 3 2.5c.5 1.5 1 2 2 2 1.5 0 3-1.5 3-3.5 0-3-2.2-3-3-5-1-2.5 0-5-1.5-6.5C16 3 13.5 3.5 12 5.5z"/>',
"microscope":'<path d="M6 18h8M7 18a9 9 0 0 1 9-9"/><circle cx="10" cy="5" r="2"/><path d="M10 3V1M13 5h2M6 22h12"/>',
"zap":     '<path d="M13 2L3 14h9l-1 8 10-12h-9l1-8z"/>',
"frame":   '<rect x="4" y="4" width="16" height="16" rx="2"/><rect x="9" y="9" width="6" height="6" rx="1"/>',
"palette": '<path d="M12 2a10 10 0 0 0 0 20c1.1 0 2-.9 2-2 0-.5-.2-1-.5-1.3-.3-.4-.5-.8-.5-1.2 0-1.1.9-2 2-2h2.7c2.4 0 4.3-2 4.3-4.4A10 10 0 0 0 12 2z"/><circle cx="7.5" cy="11" r="1.2"/><circle cx="10" cy="6.5" r="1.2"/><circle cx="15" cy="6.5" r="1.2"/><circle cx="17.5" cy="10.5" r="1.2"/>',
"truck":   '<path d="M1 5h13v11H1zM14 9h4l4 4v3h-8z"/><circle cx="6" cy="18.5" r="1.8"/><circle cx="17" cy="18.5" r="1.8"/>',
"home":    '<path d="M3 11l9-8 9 8"/><path d="M5 9.5V21h14V9.5M9 21v-7h6v7"/>',
"sparkles":'<path d="M12 3l1.9 5.1L19 10l-5.1 1.9L12 17l-1.9-5.1L5 10l5.1-1.9L12 3z"/><path d="M19 15l.9 2.1L22 18l-2.1.9L19 21l-.9-2.1L16 18l2.1-.9L19 15z"/>',
"tag":     '<path d="M20.6 13.4L11 3.8A2 2 0 0 0 9.6 3.2H4a1 1 0 0 0-1 1v5.6A2 2 0 0 0 3.8 11l9.6 9.6a2 2 0 0 0 2.8 0l4.4-4.4a2 2 0 0 0 0-2.8z"/><circle cx="7.5" cy="7.5" r="1.3"/>',
"flag":    '<path d="M5 21V3"/><path d="M5 3h13l-2 4 2 4H5"/>',
"print":   '<path d="M6 8V3h12v5M6 18h12v3H6zM6 14h12v4H6zM6 10h12"/>',
"puzzle":  '<path d="M12 3a2 2 0 0 1 4 0 2 2 0 0 1 2 2h2v5h-2a2 2 0 0 0-4 0h-2V7H9a2 2 0 0 0-2-2 2 2 0 0 1 4-2zM12 14v7h9v-5h-2a2 2 0 0 1-4 0H12z"/>',
"gem":     '<path d="M6 3h12l4 6-10 12L2 9l4-6z"/><path d="M2 9h20M12 21L8 9l3-6M12 21l4-12-3-6"/>',
"scissors":'<circle cx="6" cy="6" r="3"/><circle cx="6" cy="18" r="3"/><path d="M20 4L8.1 15.9M14.5 14.5L20 20M8.1 8.1L12 12"/>',
"shirt":   '<path d="M20 7l-4-3h-2a2 2 0 0 1-4 0H8L4 7l-1 5 4 1v9h10v-9l4-1-1-5z"/>',
"tree":    '<path d="M12 2l3 5H9l3-5z"/><path d="M5 17l2-5h10l2 5H5z"/><path d="M8 22l1-5h6l1 5H8z"/>',
"package": '<path d="M21 8v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V8"/><path d="M1 4h22v4H1zM10 13h4"/>',
"store":   '<path d="M3 9l1.5-5h15L21 9M3 9a2 2 0 0 0 4 0 2 2 0 0 0 4 0 2 2 0 0 0 4 0 2 2 0 0 0 4 0M5 11v9h14v-9M10 20v-6h4v6"/>',
}

def icon(name):
    return '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">' + ICONS.get(name, ICONS["cross"]) + "</svg>"

# ---------------------------------------------------------------- datos por lead

OVERRIDES = {
"LEAD-IA-1313": {
    "accent": "#0E7C7B", "accent_dark": "#0A5C5B", "short": "Clínica Luro", "badge": "Clínica médica · Mar del Plata",
    "tagline": "Atención médica integral en el corazón de Mar del Plata. Agendá tu turno por WhatsApp en menos de un minuto, sin llamadas ni esperas.",
    "sub": "En Clínica Luro reunimos especialistas de primer nivel con tecnología de gestión que te permite reservar, reprogramar y confirmar tu consulta desde el celular, las 24 horas.",
    "especialidades": [
        ("Clínica médica", "Consultas de rutina, control y seguimiento con médicos clínicos de cabecera.", "user"),
        ("Cardiología", "Evaluación cardiovascular, electrocardiograma y control de factores de riesgo.", "heart"),
        ("Dermatología", "Consulta dermatológica, control de lunares y tratamientos estéticos médicos.", "droplet"),
        ("Ginecología", "Controles anuales, estudios ginecológicos y seguimiento integral de la paciente.", "activity"),
        ("Pediatría", "Control de niño sano, vacunación y atención de consultas pediátricas.", "baby"),
        ("Nutrición", "Plan alimentario personalizado y seguimiento de hábitos saludables.", "leaf"),
    ],
    "horarios": [("Lunes a Viernes", "08:00 – 20:00"), ("Sábados", "08:00 – 14:00")],
},
"LEAD-IA-1301": {
    "accent": "#1D4E89", "accent_dark": "#153B6B", "short": "Sanatorio Avenida", "badge": "Sanatorio · Mar del Plata",
    "tagline": "Atención de confianza para toda la familia. Guardia y consultorios con turnos confirmados al instante por WhatsApp.",
    "sub": "Sanatorio Avenida combina más de dos décadas de atención sanitaria con un sistema de turnos digital que elimina las salas de espera: reservá, confirmá y recibí recordatorios automáticos.",
    "especialidades": [
        ("Guardia 24 hs", "Atención de urgencias y emergencias durante todo el día, todos los días.", "shield"),
        ("Clínica médica", "Consulta clínica general, control de patologías crónicas y certificados.", "steth"),
        ("Análisis clínicos", "Extracciones y laboratorio con resultados digitales y derivación por WhatsApp.", "microscope"),
        ("Ecografías", "Estudios ecográficos con turno programado y entrega inmediata de imágenes.", "clipboard"),
        ("Kinesiología", "Rehabilitación kinesiológica con seguimiento personalizado.", "activity"),
        ("Cirugía ambulatoria", "Procedimientos menores con internación de día y alta segura.", "cross"),
    ],
    "horarios": [("Guardia", "24 horas"), ("Consultorios", "Lun a Vie 08:00 – 20:00"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-1304": {
    "accent": "#B03A48", "accent_dark": "#8F2B38", "short": "Dra. Ana Gracia Alonso", "badge": "Cardiología · Mar del Plata",
    "tagline": "El cuidado de tu corazón, con estudios de precisión y un seguimiento que no te suelta. Turnos por WhatsApp.",
    "sub": "Consultorio de cardiología de la Dra. Ana Gracia Alonso en Av. Fortunato de la Plaza 3668. Electrocardiograma, ecocardiograma y ergometría con agenda digital: pedí tu turno sin llamar.",
    "especialidades": [
        ("Consulta cardiológica", "Evaluación cardiovascular completa y control periódico de pacientes.", "heart"),
        ("Electrocardiograma", "ECG de reposo con informe inmediato y seguimiento en consulta.", "activity"),
        ("Ecocardiograma Doppler", "Estudio ecográfico del corazón con evaluación de válvulas y función ventricular.", "microscope"),
        ("Ergometría", "Prueba de esfuerzo para evaluar la respuesta cardiovascular al ejercicio.", "zap"),
        ("Holter 24 hs", "Monitoreo ambulatorio de ritmo cardíaco con informe detallado.", "clock"),
        ("Cardioprevención", "Plan preventivo personalizado: riesgo cardiovascular, actividad y alimentación.", "shield"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 18:00"), ("Sábados", "con turno previo")],
},
"LEAD-IA-1279": {
    "accent": "#4A6B57", "accent_dark": "#3A5545", "short": "Consultorios Colinas", "badge": "Consultorios integrales · Mar del Plata",
    "tagline": "Salud integral para toda la familia: kinesiología, psicología, nutrición y más en un solo lugar.",
    "sub": "En Consultorios Integrales Colinas reunimos un equipo multidisciplinario bajo el mismo techo. Turnos coordinados por WhatsApp para que organices tus terapias sin vueltas.",
    "especialidades": [
        ("Kinesiología y fisioterapia", "Rehabilitación de lesiones, dolores crónicos y recuperación post operatoria.", "activity"),
        ("Psicología", "Terapia individual y de pareja con profesionales matriculados.", "brain"),
        ("Nutrición", "Plan alimentario personalizado, deporte y salud digestiva.", "leaf"),
        ("Fonoaudiología", "Evaluación y tratamiento del lenguaje, la voz y la deglución.", "eye"),
        ("Osteopatía", "Terapia manual para el alivio del dolor y la restauración del movimiento.", "bone"),
        ("Clínica médica", "Consulta clínica, controles anuales y certificados médicos.", "steth"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 19:00"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-1300": {
    "accent": "#176B87", "accent_dark": "#105368", "short": "Océano Mar", "badge": "Consultorios · Mar del Plata",
    "tagline": "Especialistas bajo el mismo techo, con atención humana y turnos digitales. Cerca del mar, cerca de vos.",
    "sub": "Océano Mar Consultorios en Rodríguez Peña 2675. Un espacio de salud moderno donde cada especialista cuenta con agenda digital conectada a WhatsApp.",
    "especialidades": [
        ("Medicina general", "Consulta de rutina, control de salud y derivación coordinada.", "user"),
        ("Dermatología", "Diagnóstico y tratamiento de afecciones de la piel, pelo y uñas.", "droplet"),
        ("Nutrición y deporte", "Planes de alimentación para salud, rendimiento y descenso de peso.", "leaf"),
        ("Psicología", "Atención psicológica individual y terapia breve.", "brain"),
        ("Ecografías", "Estudios ecográficos ginecológicos, abdominales y musculares.", "microscope"),
        ("Kinesiología", "Tratamiento kinésico y readaptación deportiva.", "activity"),
    ],
    "horarios": [("Lunes a Viernes", "08:30 – 19:30"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-1275": {
    "accent": "#5B5EA6", "accent_dark": "#474A87", "short": "Centro Médico Edison", "badge": "Centro médico · Mar del Plata",
    "tagline": "Medicina de cabecera y especialistas en un solo centro médico. Turnos por WhatsApp, sin llamadas ni esperas.",
    "sub": "Centro Médico Edison reúne clínica médica y especialidades con agenda digital: reservá, reprogramá y confirmá tu consulta desde el celular, las 24 horas.",
    "especialidades": [
        ("Clínica médica", "Consulta de cabecera, controles anuales y seguimiento de patologías crónicas.", "steth"),
        ("Cardiología", "Evaluación cardiovascular con electrocardiograma y control de riesgo.", "heart"),
        ("Kinesiología", "Rehabilitación y tratamiento del dolor con kinesiólogos matriculados.", "activity"),
        ("Ecografías", "Estudios ecográficos abdominales, musculares y de partes blandas.", "microscope"),
        ("Psicología", "Atención psicológica individual y orientación terapéutica.", "brain"),
        ("Dermatología", "Consulta dermatológica y control de la piel.", "droplet"),
    ],
    "horarios": [("Lunes a Viernes", "08:00 – 20:00"), ("Sábados", "08:00 – 13:00")],
},
"LEAD-IA-1294": {
    "accent": "#A66A2E", "accent_dark": "#855322", "short": "Instituto San Lorenzo", "badge": "Instituto médico · Mar del Plata",
    "tagline": "Instituto médico de confianza en el centro de Mar del Plata. Reservá tu consulta por WhatsApp al instante.",
    "sub": "San Lorenzo Instituto Médico en San Lorenzo 2076. Atención médica integral con turnos digitales, recordatorios automáticos y confirmación inmediata.",
    "especialidades": [
        ("Clínica médica", "Consulta general, controles de salud y certificados médicos.", "user"),
        ("Cardiología", "Consulta cardiológica, ECG y seguimiento de pacientes con riesgo cardiovascular.", "heart"),
        ("Traumatología", "Atención de lesiones, dolores articulares y seguimiento post quirúrgico.", "bone"),
        ("Análisis clínicos", "Laboratorio con extracción en sede y resultados digitales.", "microscope"),
        ("Ecografías", "Estudios ecográficos con turno programado.", "clipboard"),
        ("Nutrición", "Plan alimentario personalizado y control de peso.", "leaf"),
    ],
    "horarios": [("Lunes a Viernes", "08:00 – 20:00"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-1297": {
    "accent": "#C2418C", "accent_dark": "#9E3372", "short": "Estética N", "badge": "Centro de estética · Mar del Plata",
    "tagline": "Tu centro de estética y belleza en Mar del Plata. Agendá tus tratamientos por WhatsApp en 30 segundos.",
    "sub": "En Estética N cuidamos cada detalle de tu imagen. Reservá depilación, tratamientos faciales y corporales con agenda digital conectada a WhatsApp.",
    "ben_title": "Tu belleza, con menos vueltas",
    "especialidades": [
        ("Depilación láser", "Sesiones con tecnología de última generación y plan personalizado.", "zap"),
        ("Limpieza facial", "Tratamientos faciales para todo tipo de piel, con diagnóstico previo.", "droplet"),
        ("Manicuría y pedicuría", "Uñas esculpidas, semipermanente y cuidados de manos y pies.", "user"),
        ("Masajes", "Relajación, descontracturantes y drenaje linfático.", "heart"),
        ("Tratamientos corporales", "Reducción, radiofrecuencia y modelado corporal.", "activity"),
        ("Lifting de pestañas", "Pestañas perfectas con lifting, lash lifting y laminado.", "eye"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 20:00"), ("Sábados", "09:00 – 14:00")],
},
"LEAD-IA-1286": {
    "accent": "#3A7D44", "accent_dark": "#2E6340", "short": "Córdoba Consultorios", "badge": "Consultorios médicos · Mar del Plata",
    "tagline": "Consultorios médicos con especialistas que te acompañan. Turnos por WhatsApp confirmados al instante.",
    "sub": "Córdoba Consultorios Médicos en Mar del Plata. Un equipo de especialistas con agenda digital para que reservar tu consulta sea lo más simple del día.",
    "especialidades": [
        ("Clínica médica", "Consulta general, controles periódicos y seguimiento integral.", "steth"),
        ("Cardiología", "Evaluación del corazón con ECG y control de factores de riesgo.", "heart"),
        ("Traumatología", "Atención de lesiones, deportología y dolor articular.", "bone"),
        ("Ecografías", "Estudios ecográficos con informe inmediato.", "microscope"),
        ("Análisis clínicos", "Laboratorio y extracciones con resultados por WhatsApp.", "clipboard"),
        ("Nutrición", "Plan alimentario y hábitos saludables personalizados.", "leaf"),
    ],
    "horarios": [("Lunes a Viernes", "08:00 – 19:00"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-1284": {
    "accent": "#0E7490", "accent_dark": "#0A5C72", "short": "Espacio Edison", "badge": "Espacio de salud · Mar del Plata",
    "tagline": "Un espacio de salud y bienestar en Av. Edison 1736. Terapias y consultas con turno por WhatsApp.",
    "sub": "Espacio Edison es un lugar pensado para tu bienestar integral. Profesionales de distintas disciplinas, una sola agenda digital y atención humana.",
    "especialidades": [
        ("Kinesiología", "Rehabilitación, tratamiento del dolor y reeducación postural.", "activity"),
        ("Psicología", "Terapia individual, ansiedad y acompañamiento emocional.", "brain"),
        ("Osteopatía", "Terapia manual para restaurar el equilibrio del cuerpo.", "bone"),
        ("Fonoaudiología", "Evaluación y tratamiento del lenguaje y la voz.", "eye"),
        ("Terapia ocupacional", "Acompañamiento para la independencia en las actividades diarias.", "user"),
        ("Nutrición", "Alimentación saludable y planes personalizados.", "leaf"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 19:00"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-1281": {
    "accent": "#8A63B2", "accent_dark": "#6E4C93", "short": "Consultorio Edison", "badge": "Consultorio médico · Mar del Plata",
    "tagline": "Atención médica cercana y sin vueltas. Sacá tu turno por WhatsApp y evitate la espera en el consultorio.",
    "sub": "Consultorio Edison en Mar del Plata. Consultas y estudios con agenda digital: reservá, reprogramá y confirmá desde el celular, las 24 horas.",
    "especialidades": [
        ("Clínica médica", "Consulta de rutina, controles y certificados médicos.", "steth"),
        ("Cardiología", "Evaluación cardiovascular y electrocardiograma.", "heart"),
        ("Ginecología", "Controles ginecológicos y seguimiento integral de la paciente.", "activity"),
        ("Pediatría", "Atención pediátrica y control del niño sano.", "baby"),
        ("Análisis clínicos", "Extracciones y laboratorio con resultados digitales.", "microscope"),
        ("Ecografías", "Estudios ecográficos con turno programado.", "clipboard"),
    ],
    "horarios": [("Lunes a Viernes", "08:00 – 19:00"), ("Sábados", "08:00 – 13:00")],
},
"LEAD-IA-1290": {
    "accent": "#1F7A5C", "accent_dark": "#175F46", "short": "KYTOS Salud", "badge": "Salud integral · Mar del Plata",
    "tagline": "Salud integral en Mar del Plata: cuerpo, mente y movimiento bajo un mismo techo. Turnos por WhatsApp.",
    "sub": "KYTOS Salud Integral en Av. Edison 2044. Un equipo interdisciplinario que te acompaña con agenda digital y recordatorios automáticos.",
    "especialidades": [
        ("Kinesiología", "Rehabilitación, tratamiento del dolor y reeducación postural.", "activity"),
        ("Osteopatía", "Terapia manual para el equilibrio y la movilidad del cuerpo.", "bone"),
        ("Psicología", "Atención psicológica individual y acompañamiento emocional.", "brain"),
        ("Nutrición", "Plan alimentario personalizado y hábitos saludables.", "leaf"),
        ("Medicina general", "Consulta de rutina y control integral de salud.", "steth"),
        ("Terapias complementarias", "Relajación, respiración y bienestar integral.", "heart"),
    ],
    "horarios": [("Lunes a Viernes", "08:30 – 20:00"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-1315": {
    "accent": "#C2564E", "accent_dark": "#9C423C", "short": "Consultorios San Juan", "badge": "Consultorios médicos · Mar del Plata",
    "tagline": "Especialistas que te atienden con calma y puntualidad. Reservá tu consulta por WhatsApp en 30 segundos.",
    "sub": "Consultorios San Juan en Mar del Plata. Medicina de familia y especialidades con confirmación de turno inmediata por WhatsApp.",
    "especialidades": [
        ("Clínica médica", "Consulta de cabecera, controles y seguimiento.", "steth"),
        ("Cardiología", "Consulta cardiológica, ECG y control de riesgo.", "heart"),
        ("Dermatología", "Piel, pelo y uñas: diagnóstico y tratamiento.", "droplet"),
        ("Ginecología", "Controles ginecológicos y estudios de rutina.", "activity"),
        ("Pediatría", "Control del niño sano y vacunación.", "baby"),
        ("Kinesiología", "Rehabilitación y tratamiento del dolor.", "activity"),
    ],
    "horarios": [("Lunes a Viernes", "08:00 – 20:00"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-1311": {
    "accent": "#316DA0", "accent_dark": "#25567F", "short": "Consultorios Luro", "badge": "Consultorios médicos · Mar del Plata",
    "tagline": "Consultorios médicos sobre Av. Pedro Luro, con especialistas y turnos confirmados al instante por WhatsApp.",
    "sub": "Consultorios Médicos Luro en Av. Pedro Luro 3806. Atención médica de calidad con agenda digital: pedí, confirmá y recibí recordatorios automáticos.",
    "especialidades": [
        ("Clínica médica", "Consulta general, controles anuales y seguimiento.", "user"),
        ("Cardiología", "Evaluación cardiovascular y electrocardiograma.", "heart"),
        ("Traumatología", "Lesiones, dolor articular y seguimiento post quirúrgico.", "bone"),
        ("Nutrición", "Planes alimentarios y control de hábitos.", "leaf"),
        ("Psicología", "Atención psicológica individual.", "brain"),
        ("Ecografías", "Estudios ecográficos con informe inmediato.", "microscope"),
    ],
    "horarios": [("Lunes a Viernes", "08:00 – 20:00"), ("Sábados", "08:00 – 13:00")],
},
"LEAD-IA-1314": {
    "accent": "#C06B3E", "accent_dark": "#9C542F", "short": "Centro de Día Lackar", "badge": "Centro de día · Mar del Plata",
    "tagline": "Un lugar de cuidado y contención para tu familia. Talleres, rehabilitación y acompañamiento con turno por WhatsApp.",
    "sub": "Centro de Día Lackar del Sur en Figueroa Alcorta 2041. Acompañamiento diurno con actividades terapéuticas, profesionales capacitados y comunicación directa por WhatsApp con las familias.",
    "especialidades": [
        ("Talleres terapéuticos", "Actividades cognitivas, recreativas y de estimulación diaria.", "brain"),
        ("Kinesiología", "Rehabilitación motora y mantenimiento funcional.", "activity"),
        ("Psicología", "Acompañamiento psicológico individual y familiar.", "user"),
        ("Enfermería", "Cuidados de enfermería y administración de medicación.", "shield"),
        ("Nutrición", "Plan alimentario adaptado a cada persona.", "leaf"),
        ("Terapia ocupacional", "Actividades para la autonomía y la vida diaria.", "clipboard"),
    ],
    "horarios": [("Lunes a Viernes", "08:00 – 17:00"), ("Sábados", "con turno previo")],
},
# ---------------- TANDA 4: Decoración / Hogar / Artesanal (Comercio General) ----------------
"LEAD-IA-0487": {
    "template": "comercio", "accent": "#8B5A2B", "accent_dark": "#6E4520", "short": "Molduras Aguilera",
    "badge": "Marcos y molduras · Mar del Plata", "mark_icon": "frame",
    "tagline": "Marcos, molduras y enmarcados a medida en Mar del Plata. Pedí tu trabajo por WhatsApp y retiralo listo.",
    "ru_title": "Marcos y molduras", "prod_title": "Enmarcamos lo que más querés",
    "footer_desc": "Marcos, molduras y enmarcados a medida en Mar del Plata, con terminación profesional.",
    "productos": [
        ("Marco de madera 40x50", "Marco clásico de madera natural, listo para colgar.", "$18.900", "frame"),
        ("Marco de aluminio 30x40", "Diseño liviano y moderno, ideal para fotos y diplomas.", "$14.500", "frame"),
        ("Molduras decorativas x1m", "Molduras para cielorrasos y paredes, listas para pintar.", "$6.800", "home"),
        ("Cuadro enmarcado 50x70", "Lienzo enmarcado con paspartú, armado completo.", "$32.000", "palette"),
        ("Enmarcado de obra a medida", "Traé tu obra o afiche y lo dejamos listo para colgar.", "$24.000", "sparkles"),
        ("Espejo con marco", "Espejos con marco de madera o aluminio a elección.", "$28.500", "home"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 18:00"), ("Sábados", "09:30 – 13:30")],
},
"LEAD-IA-0512": {
    "template": "comercio", "accent": "#C0562F", "accent_dark": "#9C4526", "short": "Krearte",
    "badge": "Arte y regalos · Mar del Plata", "mark_icon": "palette",
    "tagline": "Arte, artesanías y regalos originales en Mar del Plata. Pedí por WhatsApp y retirá o recibí en tu casa.",
    "ru_title": "Arte y artesanías", "prod_title": "Piezas únicas hechas con amor",
    "footer_desc": "Arte, artesanías y regalos originales, con envíos y retiro en Mar del Plata.",
    "productos": [
        ("Cuadros personalizados", "Arte a medida según tu foto o idea, en varios tamaños.", "$25.000", "palette"),
        ("Artesanías en madera", "Objetos únicos tallados a mano para decorar.", "$9.800", "sparkles"),
        ("Velas artesanales", "Velas de soja con aromas, hechas en MDP.", "$7.500", "sparkles"),
        ("Decoración para eventos", "Detalles y centros de mesa para fechas especiales.", "$12.000", "store"),
        ("Sets de regalo", "Combos armados para regalar en cualquier ocasión.", "$15.000", "tag"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 18:00"), ("Sábados", "09:30 – 13:30")],
},
"LEAD-IA-0514": {
    "template": "comercio", "accent": "#38618C", "accent_dark": "#2C4D73", "short": "Patch & Flag",
    "badge": "Parches y banderas · Mar del Plata", "mark_icon": "flag",
    "tagline": "Parches bordados, banderas y sublimación textil en Mar del Plata. Pedilo por WhatsApp y lo tenés en pocos días.",
    "ru_title": "Parches y textiles", "prod_title": "Personalizado en tela, a tu medida",
    "footer_desc": "Parches bordados, banderas y sublimación textil personalizada en Mar del Plata.",
    "productos": [
        ("Parches bordados", "Parches con tu diseño, logo o escudo, en varios tamaños.", "$4.200", "flag"),
        ("Banderas personalizadas", "Banderas con estampado full color, varios formatos.", "$16.000", "flag"),
        ("Sublimación textil", "Estampamos tu diseño en tela, con entrega coordinada.", "$8.500", "shirt"),
        ("Remeras personalizadas", "Remeras con tu diseño, talle y color a elección.", "$11.000", "shirt"),
        ("Calcos y vinilos", "Sublimación rígida para mates, tazas y más.", "$5.800", "print"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 17:00"), ("Sábados", "09:30 – 13:00")],
},
"LEAD-IA-0515": {
    "template": "comercio", "accent": "#7A5FA0", "accent_dark": "#5F4A83", "short": "Dimensión Atelier",
    "badge": "Atelier de diseño · Mar del Plata", "mark_icon": "palette",
    "tagline": "Diseño, objetos y decoración hechos a mano en Mar del Plata. Consultá y pedí por WhatsApp.",
    "ru_title": "Diseño y decoración", "prod_title": "Objetos con diseño y personalidad",
    "footer_desc": "Diseño y objetos decorativos artesanales en Mar del Plata.",
    "productos": [
        ("Decoración de ambientes", "Asesoramiento y armado de espacios con estilo propio.", "$35.000", "home"),
        ("Objetos de diseño", "Piezas decorativas únicas hechas en el atelier.", "$18.000", "sparkles"),
        ("Cuadros de autor", "Arte original, ediciones limitadas firmadas.", "$42.000", "palette"),
        ("Accesorios para el hogar", "Detalles que cambian un ambiente.", "$9.500", "home"),
    ],
    "horarios": [("Lunes a Viernes", "10:00 – 18:00"), ("Sábados", "10:00 – 14:00")],
},
"LEAD-IA-0496": {
    "template": "comercio", "accent": "#4E7A3A", "accent_dark": "#3D6230", "short": "Humus",
    "badge": "Vivero y huerta · Mar del Plata", "mark_icon": "tree",
    "tagline": "Huerta, plantas y productos orgánicos en Mar del Plata. Pedí por WhatsApp y retirá tu kit listo.",
    "ru_title": "Huerta y vivero", "prod_title": "Verde todo el año",
    "footer_desc": "Huerta, plantas y productos orgánicos en Mar del Plata.",
    "productos": [
        ("Humus de lombriz 5kg", "Abono orgánico puro para huerta y macetas.", "$8.500", "leaf"),
        ("Compostera hogareña", "Armá tu propio compost, con guía incluida.", "$38.000", "tree"),
        ("Kit de huerta inicial", "Macetas, semillas y sustrato para empezar hoy.", "$22.000", "leaf"),
        ("Plantines de temporada", "Tomate, albahaca, lechuga y más, listos para trasplantar.", "$3.800", "tree"),
        ("Tierra abonada 10kg", "Sustrato preparado para huerta y macetones.", "$6.500", "leaf"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 17:00"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-0497": {
    "template": "comercio", "accent": "#3A7CA5", "accent_dark": "#2E6386", "short": "ALL DESIGN",
    "badge": "Diseño y decoración · Mar del Plata", "mark_icon": "palette",
    "tagline": "Diseño, vinilos y decoración para tu casa o negocio en Mar del Plata. Presupuestá por WhatsApp.",
    "ru_title": "Diseño a medida", "prod_title": "Tu espacio, con diseño propio",
    "footer_desc": "Diseño, vinilos decorativos y soluciones para espacios en Mar del Plata.",
    "productos": [
        ("Vinilos decorativos", "Vinilos para paredes, vidrios y vidrieras, instalados.", "$7.500", "print"),
        ("Cartelería y banners", "Diseño e impresión para tu local o evento.", "$14.000", "print"),
        ("Diseño de interiores", "Proyecto integral de decoración de ambientes.", "$45.000", "home"),
        ("Regalos corporativos", "Merchandising personalizado con tu marca.", "$16.000", "tag"),
    ],
    "horarios": [("Lunes a Viernes", "09:30 – 18:30"), ("Sábados", "10:00 – 14:00")],
},
"LEAD-IA-0508": {
    "template": "comercio", "accent": "#5D6D7E", "accent_dark": "#495765", "short": "Decor Wall",
    "badge": "Decoración de paredes · Mar del Plata", "mark_icon": "home",
    "tagline": "Vinilos, paneles y revestimientos de pared en Mar del Plata. Consultá por WhatsApp y transformá tu espacio.",
    "ru_title": "Paredes con estilo", "prod_title": "Cambiá tu espacio en un fin de semana",
    "footer_desc": "Vinilos, paneles y revestimientos decorativos de pared en Mar del Plata.",
    "productos": [
        ("Vinilos de pared", "Diseños listos o a medida, con instalación a domicilio.", "$8.200", "print"),
        ("Papel texturizado", "Revestimientos con relieve para dar profundidad.", "$21.000", "home"),
        ("Paneles decorativos", "Paneles 3D de PVC, livianos y fáciles de instalar.", "$18.000", "home"),
        ("Letras corpóreas", "Nombres y frases en 3D para paredes y vidrieras.", "$12.500", "sparkles"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 18:00"), ("Sábados", "09:30 – 13:30")],
},
"LEAD-IA-0509": {
    "template": "comercio", "accent": "#2C3E50", "accent_dark": "#233242", "short": "Monograf",
    "badge": "Imprenta y gráfica · Mar del Plata", "mark_icon": "print",
    "tagline": "Imprenta y servicio gráfico en Mar del Plata. Tarjetas, gigantografías y más, pedidas por WhatsApp.",
    "ru_title": "Servicio gráfico", "prod_title": "Impreso rápido y prolijo",
    "footer_desc": "Imprenta y servicio gráfico integral en Mar del Plata.",
    "productos": [
        ("Impresión digital", "Color y B/N, todos los tamaños, listo en el día.", "$3.500", "print"),
        ("Tarjetas personales", "Diseño + impresión de tarjetas para tu negocio.", "$6.800", "print"),
        ("Gigantografías", "Lonas y banners de gran formato para locales y eventos.", "$32.000", "flag"),
        ("Volantes y folletos", "Tirada mínima económica, con diseño incluido.", "$9.500", "print"),
        ("Encuadernación", "Tesinas, anillados y libros con acabado profesional.", "$4.800", "clipboard"),
    ],
    "horarios": [("Lunes a Viernes", "08:30 – 18:00"), ("Sábados", "09:00 – 13:00")],
},
"LEAD-IA-0565": {
    "template": "comercio", "accent": "#138D75", "accent_dark": "#0F7160", "short": "La Isla",
    "badge": "Regalería y decoración · Mar del Plata", "mark_icon": "sparkles",
    "tagline": "Regalos y decoración originales en el centro de Mar del Plata. Pedí por WhatsApp y lo tenés listo para regalar.",
    "ru_title": "Regalos y decoración", "prod_title": "El detalle que hace la diferencia",
    "footer_desc": "Regalería y artículos de decoración en Mar del Plata.",
    "productos": [
        ("Velas decorativas", "Velas aromáticas y decorativas, ideales para regalar.", "$6.800", "sparkles"),
        ("Souvenirs", "Recuerdos y detalles para eventos y regalos.", "$4.500", "tag"),
        ("Objetos de decoración", "Figuras, portarretratos y detalles para tu casa.", "$12.000", "home"),
        ("Sets de regalo", "Combos armados con envoltorio incluido.", "$14.500", "tag"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 19:00"), ("Sábados", "09:30 – 13:30")],
},
"LEAD-IA-0520": {
    "template": "comercio", "accent": "#C96A18", "accent_dark": "#A35512", "short": "Rompecabezas",
    "badge": "Juegos y puzzles · Mar del Plata", "mark_icon": "puzzle",
    "tagline": "Rompecabezas, juegos de mesa y puzzles personalizados en Mar del Plata. Pedí por WhatsApp.",
    "ru_title": "Juegos y puzzles", "prod_title": "Horas de diversión en familia",
    "footer_desc": "Rompecabezas, juegos de mesa y puzzles personalizados en Mar del Plata.",
    "productos": [
        ("Rompecabezas 1000 piezas", "Paisajes, arte y diseños exclusivos.", "$16.000", "puzzle"),
        ("Puzzles infantiles", "Para los más chicos, con piezas resistentes.", "$8.500", "puzzle"),
        ("Juegos de mesa", "Selección curada de juegos para toda la familia.", "$22.000", "sparkles"),
        ("Puzzle personalizado", "Tu foto convertida en puzzle, varios tamaños.", "$19.000", "puzzle"),
    ],
    "horarios": [("Lunes a Viernes", "10:00 – 19:00"), ("Sábados", "10:00 – 14:00")],
},
"LEAD-IA-0576": {
    "template": "comercio", "accent": "#8E6E53", "accent_dark": "#725743", "short": "Aderezo",
    "badge": "Accesorios para marroquinería · Mar del Plata", "mark_icon": "scissors",
    "tagline": "Accesorios, avíos y herramientas para marroquinería en Mar del Plata. Pedí por WhatsApp.",
    "ru_title": "Marroquinería", "prod_title": "Todo para tu taller",
    "footer_desc": "Accesorios, avíos y herramientas para marroquinería en Mar del Plata.",
    "productos": [
        ("Accesorios de marroquinería", "Hebillas, argollas y herrajes en varios acabados.", "$7.800", "scissors"),
        ("Herramientas del oficio", "Leznas, cutters y bancos de trabajo profesionales.", "$3.200", "scissors"),
        ("Cierres y avíos", "Cierres por metro, botones y remaches.", "$2.800", "tag"),
        ("Kit de costura de cuero", "Insumos básicos para empezar a trabajar.", "$9.500", "scissors"),
    ],
    "horarios": [("Lunes a Viernes", "09:00 – 18:00"), ("Sábados", "09:30 – 13:00")],
},
"LEAD-IA-0580": {
    "template": "comercio", "accent": "#2E8B57", "accent_dark": "#247046", "short": "Turmalina",
    "badge": "Bijou y accesorios · Mar del Plata", "mark_icon": "gem",
    "tagline": "Bijou artesanal y accesorios en Mar del Plata. Pedí tu favorito por WhatsApp y retiralo en el local.",
    "ru_title": "Bijou y accesorios", "prod_title": "Detalles que brillan",
    "footer_desc": "Bijou artesanal y accesorios en Mar del Plata.",
    "productos": [
        ("Bijou artesanal", "Aros, collares y pulseras hechos a mano.", "$6.800", "gem"),
        ("Aros de piedras", "Piedras naturales en diseños exclusivos.", "$5.500", "gem"),
        ("Collares personalizados", "Con tu nombre o inicial, estilo a elección.", "$9.800", "gem"),
        ("Sets de regalo", "Combos de bijou con caja para regalar.", "$12.000", "tag"),
    ],
    "horarios": [("Lunes a Viernes", "09:30 – 18:30"), ("Sábados", "09:30 – 13:30")],
},
"LEAD-IA-0587": {
    "template": "comercio", "accent": "#A9483D", "accent_dark": "#873A31", "short": "Casa Moda",
    "badge": "Moda y hogar · Mar del Plata", "mark_icon": "home",
    "tagline": "Moda, textil y artículos para el hogar en Mar del Plata. Pedí por WhatsApp y recibilo en tu casa.",
    "ru_title": "Moda y hogar", "prod_title": "Estilo para vos y tu casa",
    "footer_desc": "Moda, textil y artículos para el hogar en Mar del Plata.",
    "productos": [
        ("Textiles de hogar", "Cortinas, mantelería y toallas de calidad.", "$14.000", "home"),
        ("Cobertores y mantas", "Texturas cálidas para el invierno.", "$25.000", "home"),
        ("Cojines y acentos", "Detalles textiles para renovar ambientes.", "$9.800", "sparkles"),
        ("Cortinas a medida", "Confección y colocación en Mar del Plata.", "$18.500", "home"),
    ],
    "horarios": [("Lunes a Viernes", "09:30 – 18:30"), ("Sábados", "09:30 – 13:30")],
},
}

# ---------------------------------------------------------------- template HTML

TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{TITLE}}</title>
<meta name="description" content="{{DESC}}">
<meta name="robots" content="index, follow">
<meta name="geo.region" content="AR-B">
<meta name="geo.placename" content="Mar del Plata">
<meta name="geo.position" content="{{LAT}};{{LNG}}">
<meta name="ICBM" content="{{LAT}}, {{LNG}}">
<meta property="og:title" content="{{TITLE}}">
<meta property="og:description" content="{{DESC}}">
<meta property="og:type" content="website">
<meta property="og:locale" content="es_AR">
<link rel="icon" href="data:image/svg+xml,{{FAVICON}}">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Lora:wght@400;500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<script type="application/ld+json">
{
  "@context": "https://schema.org",
  "@type": "MedicalBusiness",
  "name": "{{SHORT}}",
  "description": "{{DESC}}",
  "image": "",
  "telephone": "+54 {{TEL_INTL}}",
  "address": {
    "@type": "PostalAddress",
    "streetAddress": "{{ADDR}}",
    "addressLocality": "Mar del Plata",
    "addressRegion": "Buenos Aires",
    "postalCode": "B7600",
    "addressCountry": "AR"
  },
  "geo": {"@type": "GeoCoordinates", "latitude": "{{LAT}}", "longitude": "{{LNG}}"},
  "openingHoursSpecification": [
    {"@type": "OpeningHoursSpecification", "dayOfWeek": ["Monday","Tuesday","Wednesday","Thursday","Friday"], "opens": "08:00", "closes": "20:00"}
  ],
  "areaServed": "Mar del Plata"
}
</script>
<style>
:root{
  --accent:{{ACCENT}};
  --accent-dark:{{ACCENT_DARK}};
  --accent-tint:{{ACCENT_TINT}};
  --ink:#15262B;
  --muted:#5C7075;
  --bg:#F6FAF9;
  --surface:#FFFFFF;
  --line:rgba(21,38,43,.1);
  --radius:18px;
  --ease:cubic-bezier(.22,1,.36,1);
  --serif:"Lora",Georgia,serif;
  --sans:"Inter",system-ui,sans-serif;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
@media (prefers-reduced-motion:reduce){
  html{scroll-behavior:auto}
  *,*::before,*::after{animation-duration:.01ms!important;animation-iteration-count:1!important;transition-duration:.01ms!important}
}
body{font-family:var(--sans);background:var(--bg);color:var(--ink);line-height:1.65;-webkit-font-smoothing:antialiased;overflow-x:hidden}
img{display:block;max-width:100%}
a{color:inherit;text-decoration:none}
ul{list-style:none}
h1,h2,h3{font-family:var(--serif);line-height:1.12;letter-spacing:-.015em}
.container{width:min(1120px,92%);margin-inline:auto}
section{padding:clamp(3.6rem,8vw,6.5rem) 0}
.eyebrow{display:inline-flex;align-items:center;gap:.55rem;font-size:.74rem;font-weight:700;letter-spacing:.22em;text-transform:uppercase;color:var(--accent)}
.eyebrow::before{content:"";width:30px;height:2px;background:var(--accent);border-radius:2px}
.section-title{font-size:clamp(1.8rem,4vw,3rem);margin:.65rem 0 1rem;font-weight:600}
.section-lead{color:var(--muted);max-width:60ch;font-size:clamp(.95rem,1.15vw,1.04rem)}
.section-head{padding-bottom:clamp(2rem,4.5vw,3.2rem)}

/* reveal */
.reveal{opacity:0;transform:translateY(30px);transition:opacity .85s var(--ease),transform .85s var(--ease)}
.reveal.is-in{opacity:1;transform:none}
.reveal[data-d="1"]{transition-delay:.12s}
.reveal[data-d="2"]{transition-delay:.24s}

/* nav */
.nav{position:fixed;inset:0 0 auto;z-index:100;padding:1.05rem 0;transition:background .4s,box-shadow .4s,padding .4s}
.nav.is-scrolled{background:rgba(246,250,249,.88);backdrop-filter:blur(14px);-webkit-backdrop-filter:blur(14px);box-shadow:0 1px 0 var(--line);padding:.65rem 0}
.nav-inner{display:flex;align-items:center;justify-content:space-between;gap:1rem}
.brand{display:flex;align-items:center;gap:.6rem;font-family:var(--serif);font-weight:700;font-size:1.22rem;letter-spacing:-.02em}
.brand .mark{width:38px;height:38px;flex:none;border-radius:12px;background:linear-gradient(145deg,var(--accent),var(--accent-dark));display:grid;place-items:center;color:#fff;box-shadow:0 8px 20px -8px var(--accent)}
.brand .mark svg{width:20px;height:20px}
.nav-links{display:flex;align-items:center;gap:1.7rem}
.nav-links a{font-size:.9rem;font-weight:500;color:var(--ink);position:relative;padding:.2rem 0}
.nav-links a::after{content:"";position:absolute;left:0;bottom:-2px;width:0;height:2px;background:var(--accent);transition:width .35s var(--ease)}
.nav-links a:hover::after{width:100%}
.btn{display:inline-flex;align-items:center;justify-content:center;gap:.55rem;font-family:var(--sans);font-weight:600;font-size:.9rem;padding:.9rem 1.55rem;border-radius:999px;border:none;cursor:pointer;transition:transform .3s var(--ease),box-shadow .3s var(--ease),background .3s var(--ease);white-space:nowrap}
.btn-wa{background:var(--accent);color:#fff;box-shadow:0 14px 30px -14px var(--accent)}
.btn-wa:hover{background:var(--accent-dark);transform:translateY(-2px)}
.btn-ghost{background:transparent;color:var(--ink);box-shadow:inset 0 0 0 1.5px var(--line)}
.btn-ghost:hover{box-shadow:inset 0 0 0 1.5px var(--accent);color:var(--accent);transform:translateY(-2px)}
.btn svg{width:18px;height:18px;flex:none}
.hamburger{display:none;background:none;border:none;cursor:pointer;padding:.4rem}
.hamburger span{display:block;width:26px;height:2px;background:var(--ink);margin:6px 0;transition:transform .35s var(--ease),opacity .2s}

/* hero */
.hero{min-height:100svh;display:flex;flex-direction:column;justify-content:center;position:relative;color:#fff;padding:6.5rem 0 3.2rem;overflow:hidden}
.hero-bg{position:absolute;inset:0;z-index:-2;background:radial-gradient(120% 90% at 85% 10%,{{ACCENT_TINT}} 0%,transparent 55%),linear-gradient(160deg,#0E2A2F 0%,#0A2025 60%,#0A1A1E 100%)}
.hero-bg::after{content:"";position:absolute;inset:0;background-image:radial-gradient(rgba(255,255,255,.06) 1px,transparent 1px);background-size:26px 26px;mask-image:radial-gradient(80% 70% at 50% 40%,#000 30%,transparent 100%);-webkit-mask-image:radial-gradient(80% 70% at 50% 40%,#000 30%,transparent 100%)}
.hero-bg .glow{position:absolute;width:560px;height:560px;border-radius:50%;background:radial-gradient(circle,{{ACCENT}}55 0%,transparent 65%);top:-180px;right:-120px;filter:blur(10px)}
.hero-tag{display:inline-flex;align-items:center;gap:.5rem;font-size:.74rem;font-weight:600;letter-spacing:.24em;text-transform:uppercase;border:1px solid rgba(255,255,255,.22);border-radius:999px;padding:.5rem 1.1rem;margin-bottom:1.6rem;background:rgba(255,255,255,.06);backdrop-filter:blur(8px)}
.hero-tag .dot{width:7px;height:7px;border-radius:50%;background:{{ACCENT}};box-shadow:0 0 0 4px {{ACCENT}}33;animation:pulse 2.2s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.4}}
.hero h1{font-size:clamp(2.5rem,7.5vw,4.8rem);font-weight:600;max-width:14ch;margin-bottom:1.3rem}
.hero p.sub{max-width:52ch;color:rgba(255,255,255,.82);font-size:clamp(1rem,1.35vw,1.12rem);margin-bottom:2.3rem}
.hero-ctas{display:flex;gap:.95rem;flex-wrap:wrap;margin-bottom:3rem}
.hero .btn-wa{background:#25D366;box-shadow:0 14px 30px -14px rgba(37,211,102,.55)}
.hero .btn-wa:hover{background:#1EBE5A}
.hero-pills{display:flex;gap:.7rem;flex-wrap:wrap}
.hero-pill{display:inline-flex;align-items:center;gap:.5rem;font-size:.8rem;font-weight:500;color:rgba(255,255,255,.85);border:1px solid rgba(255,255,255,.16);border-radius:999px;padding:.5rem 1rem;background:rgba(255,255,255,.05)}
.hero-pill svg{width:15px;height:15px;color:{{ACCENT}}}
.scroll-hint{position:absolute;bottom:1.8rem;left:50%;transform:translateX(-50%);display:flex;flex-direction:column;align-items:center;gap:.4rem;color:rgba(255,255,255,.5);font-size:.64rem;letter-spacing:.3em;text-transform:uppercase}
.scroll-hint .wheel{width:1px;height:42px;background:rgba(255,255,255,.2);position:relative;overflow:hidden}
.scroll-hint .wheel::after{content:"";position:absolute;left:0;top:-40%;width:1px;height:40%;background:#fff;animation:drop 1.8s var(--ease) infinite}
@keyframes drop{0%{top:-40%}60%,100%{top:110%}}

/* especialidades */
.spec-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:1.3rem}
.spec-card{background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:1.7rem;display:flex;flex-direction:column;gap:1rem;transition:transform .4s var(--ease),box-shadow .4s,border-color .4s}
.spec-card:hover{transform:translateY(-6px);box-shadow:0 24px 48px -26px rgba(21,38,43,.28);border-color:{{ACCENT}}55}
.spec-card .ic{width:52px;height:52px;flex:none;border-radius:15px;background:var(--accent-tint);color:var(--accent-dark);display:grid;place-items:center}
.spec-card .ic svg{width:26px;height:26px}
.spec-card h3{font-size:1.22rem;font-weight:600}
.spec-card p{font-size:.9rem;color:var(--muted);flex:1}
.spec-card a{display:inline-flex;align-items:center;gap:.45rem;font-weight:600;font-size:.88rem;color:var(--accent);transition:gap .3s var(--ease)}
.spec-card a:hover{gap:.7rem}
.spec-card a svg{width:16px;height:16px}

/* beneficios */
.benefits{background:var(--surface)}
.ben-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:1.3rem}
.ben{background:var(--bg);border:1px solid var(--line);border-radius:var(--radius);padding:1.6rem;transition:transform .4s var(--ease)}
.ben:hover{transform:translateY(-5px)}
.ben .ic{width:46px;height:46px;border-radius:13px;background:linear-gradient(145deg,var(--accent),var(--accent-dark));color:#fff;display:grid;place-items:center;margin-bottom:1rem}
.ben .ic svg{width:22px;height:22px}
.ben h3{font-size:1.05rem;font-weight:600;margin-bottom:.45rem}
.ben p{font-size:.87rem;color:var(--muted)}

/* pasos */
.steps{display:grid;grid-template-columns:repeat(3,1fr);gap:1.3rem}
.step{position:relative;background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:2rem 1.6rem 1.8rem}
.step .n{position:absolute;top:1.2rem;right:1.4rem;font-family:var(--serif);font-size:3.2rem;font-weight:700;color:var(--accent);opacity:.14}
.step .ic{width:48px;height:48px;border-radius:50%;background:var(--accent-tint);color:var(--accent-dark);display:grid;place-items:center;margin-bottom:1.1rem}
.step .ic svg{width:22px;height:22px}
.step h3{font-size:1.12rem;font-weight:600;margin-bottom:.45rem}
.step p{font-size:.88rem;color:var(--muted)}

/* form turno */
.turno{background:var(--surface)}
.turno-grid{display:grid;grid-template-columns:.9fr 1.1fr;gap:clamp(2.2rem,6vw,4.5rem);align-items:center}
.turno-copy .points{display:grid;gap:.9rem;margin-top:1.8rem}
.point{display:flex;gap:.95rem;align-items:flex-start}
.point .ic{width:40px;height:40px;flex:none;border-radius:11px;background:var(--accent-tint);color:var(--accent-dark);display:grid;place-items:center}
.point .ic svg{width:19px;height:19px}
.point b{font-size:.95rem;display:block}
.point p{font-size:.85rem;color:var(--muted)}
.form{background:var(--bg);border:1px solid var(--line);border-radius:calc(var(--radius) + 6px);padding:clamp(1.6rem,3vw,2.3rem);display:grid;gap:1.1rem;box-shadow:0 30px 60px -40px rgba(21,38,43,.3)}
.form .row{display:grid;grid-template-columns:1fr 1fr;gap:1rem}
.field{display:grid;gap:.4rem}
.field label{font-size:.72rem;font-weight:700;letter-spacing:.14em;text-transform:uppercase;color:var(--muted)}
.field input,.field select{font-family:var(--sans);font-size:1rem;color:var(--ink);background:var(--surface);border:1px solid var(--line);border-radius:11px;padding:.85rem 1rem;width:100%;transition:border-color .3s;-webkit-appearance:none;appearance:none}
.field input:focus,.field select:focus{outline:none;border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-tint)}
.form .btn{width:100%;justify-content:center;padding:1rem}
.form-note{font-size:.76rem;color:var(--muted);text-align:center}
.form-note b{color:var(--accent)}

/* contacto */
.contact-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:1.3rem}
.c-card{background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:1.7rem;transition:transform .4s var(--ease),box-shadow .4s}
.c-card:hover{transform:translateY(-5px);box-shadow:0 22px 45px -28px rgba(21,38,43,.3)}
.c-card .ic{width:46px;height:46px;border-radius:13px;background:var(--accent-tint);color:var(--accent-dark);display:grid;place-items:center;margin-bottom:1rem}
.c-card .ic svg{width:22px;height:22px}
.c-card h3{font-size:1.08rem;font-weight:600;margin-bottom:.55rem}
.c-card p{font-size:.9rem;color:var(--muted);line-height:1.75}
.c-card a{color:var(--accent);font-weight:600;font-size:.9rem}
.hours{width:100%;border-collapse:collapse;margin-top:.3rem}
.hours td{font-size:.88rem;color:var(--muted);padding:.2rem 0}
.hours td:last-child{text-align:right;font-weight:600;color:var(--ink)}

/* footer */
.footer{background:#0E1F23;color:#fff;padding:clamp(2.8rem,6vw,4rem) 0 2rem}
.footer-grid{display:grid;grid-template-columns:1.3fr 1fr;gap:2.2rem;margin-bottom:2.2rem;align-items:start}
.footer .brand .mark{width:34px;height:34px;border-radius:10px}
.footer-brand p{color:rgba(255,255,255,.62);font-size:.9rem;max-width:36ch;margin-top:.9rem}
.footer h4{font-size:.74rem;letter-spacing:.2em;text-transform:uppercase;color:{{ACCENT}};margin-bottom:.9rem}
.footer ul{display:grid;gap:.55rem}
.footer a{color:rgba(255,255,255,.72);font-size:.9rem;transition:color .3s}
.footer a:hover{color:#fff}
.footer-bottom{border-top:1px solid rgba(255,255,255,.1);padding-top:1.4rem;display:flex;justify-content:space-between;gap:1rem;flex-wrap:wrap;font-size:.78rem;color:rgba(255,255,255,.48)}

/* wa float */
.wa-float{position:fixed;right:1.4rem;bottom:1.4rem;z-index:90;width:58px;height:58px;border-radius:50%;background:#25D366;display:grid;place-items:center;box-shadow:0 14px 30px -10px rgba(37,211,102,.55);transition:transform .3s var(--ease)}
.wa-float:hover{transform:scale(1.1) rotate(6deg)}
.wa-float svg{width:30px;height:30px;color:#fff}
.wa-float::before{content:"";position:absolute;inset:0;border-radius:50%;background:#25D366;animation:ring 2.6s var(--ease) infinite;z-index:-1}
@keyframes ring{0%{transform:scale(1);opacity:.55}100%{transform:scale(1.7);opacity:0}}

/* responsive */
@media (max-width:900px){
  .turno-grid{grid-template-columns:1fr}
  .steps{grid-template-columns:1fr}
  .footer-grid{grid-template-columns:1fr}
}
@media (max-width:640px){
  .nav-links{position:fixed;inset:0 0 auto;z-index:99;flex-direction:column;align-items:flex-start;gap:1.3rem;background:var(--bg);padding:5.5rem 8% 2.2rem;transform:translateY(-100%);transition:transform .5s var(--ease);box-shadow:0 30px 60px -30px rgba(21,38,43,.4)}
  .nav-links.is-open{transform:none}
  .nav-links a{font-size:1.45rem;font-family:var(--serif);font-weight:600;width:100%}
  .nav-links .btn{display:none}
  .mobile-cta{display:inline-flex!important;margin-top:.5rem}
  .hamburger{display:block}
  .hamburger.is-open span:nth-child(1){transform:translateY(8px) rotate(45deg)}
  .hamburger.is-open span:nth-child(2){opacity:0}
  .hamburger.is-open span:nth-child(3){transform:translateY(-8px) rotate(-45deg)}
  .form .row{grid-template-columns:1fr}
  .hero h1{max-width:100%}
  .scroll-hint{display:none}
}
</style>
</head>
<body>

<nav class="nav" id="nav" aria-label="Navegación principal">
  <div class="container nav-inner">
    <a href="#inicio" class="brand"><span class="mark">{{MARK_ICON}}</span>{{SHORT}}</a>
    <ul class="nav-links" id="navLinks">
      <li><a href="#especialidades">Especialidades</a></li>
      <li><a href="#beneficios">Por qué elegirnos</a></li>
      <li><a href="#como">Cómo funciona</a></li>
      <li><a href="#contacto">Contacto</a></li>
      <li><a href="{{WA_HERO}}" target="_blank" rel="noopener" class="btn btn-wa mobile-cta">Turnos por WhatsApp</a></li>
    </ul>
    <a href="{{WA_HERO}}" target="_blank" rel="noopener" class="btn btn-wa">Turnos por WhatsApp</a>
    <button class="hamburger" id="hamburger" aria-label="Abrir menú" aria-expanded="false"><span></span><span></span><span></span></button>
  </div>
</nav>

<header class="hero" id="inicio">
  <div class="hero-bg" aria-hidden="true"><div class="glow"></div></div>
  <div class="container">
    <span class="hero-tag reveal"><span class="dot"></span>{{BADGE}}</span>
    <h1 class="reveal" data-d="1">{{SHORT}}</h1>
    <p class="sub reveal" data-d="2">{{TAGLINE}}</p>
    <div class="hero-ctas reveal" data-d="2">
      <a href="{{WA_HERO}}" target="_blank" rel="noopener" class="btn btn-wa">{{WA_ICON}}Sacar turno por WhatsApp</a>
      <a href="#especialidades" class="btn btn-ghost" style="color:#fff;box-shadow:inset 0 0 0 1.5px rgba(255,255,255,.28)">Ver especialidades</a>
    </div>
    <div class="hero-pills reveal" data-d="3">
      <span class="hero-pill">{{ICON_CHECK}}Turnos 24/7</span>
      <span class="hero-pill">{{ICON_CHECK}}Confirmación inmediata</span>
      <span class="hero-pill">{{ICON_CHECK}}Recordatorios automáticos</span>
    </div>
  </div>
  <div class="scroll-hint" aria-hidden="true"><span>Bajá</span><div class="wheel"></div></div>
</header>

<section id="especialidades">
  <div class="container">
    <div class="section-head">
      <span class="eyebrow reveal">Especialidades</span>
      <h2 class="section-title reveal" data-d="1">¿Qué necesitás hoy?</h2>
      <p class="section-lead reveal" data-d="2">Elegí la especialidad, tocá el botón y pedí tu turno directo por WhatsApp. Sin llamadas, sin formularios eternos.</p>
    </div>
    <div class="spec-grid">
      {{SPEC_CARDS}}
    </div>
  </div>
</section>

<section class="benefits" id="beneficios">
  <div class="container">
    <div class="section-head">
      <span class="eyebrow reveal">Por qué elegirnos</span>
      <h2 class="section-title reveal" data-d="1">{{BEN_TITLE}}</h2>
    </div>
    <div class="ben-grid">
      <div class="ben reveal">
        <span class="ic">{{ICON_ZAP}}</span>
        <h3>Reservá en 30 segundos</h3>
        <p>Agendá tu consulta desde WhatsApp en cualquier momento del día, incluso a las 2 de la mañana.</p>
      </div>
      <div class="ben reveal" data-d="1">
        <span class="ic">{{ICON_CLOCK}}</span>
        <h3>Recordatorios automáticos</h3>
        <p>Recibís avisos de tu turno por WhatsApp. Menos ausentismo, más pacientes atendidos.</p>
      </div>
      <div class="ben reveal" data-d="2">
        <span class="ic">{{ICON_SHIELD}}</span>
        <h3>Confirmación inmediata</h3>
        <p>Coordinamos la agenda al instante y te confirmamos el horario sin esperar respuesta.</p>
      </div>
      <div class="ben reveal">
        <span class="ic">{{ICON_USER}}</span>
        <h3>Atención cercana</h3>
        <p>Un equipo médico que te conoce, con seguimiento continuo de tu historia clínica.</p>
      </div>
    </div>
  </div>
</section>

<section id="como">
  <div class="container">
    <div class="section-head">
      <span class="eyebrow reveal">Así de simple</span>
      <h2 class="section-title reveal" data-d="1">Tu turno en 3 pasos</h2>
    </div>
    <div class="steps">
      <div class="step reveal">
        <span class="n">1</span>
        <span class="ic">{{ICON_PHONE}}</span>
        <h3>Escribinos</h3>
        <p>Te abrimos un chat de WhatsApp con el mensaje ya armado. Solo tocá el botón.</p>
      </div>
      <div class="step reveal" data-d="1">
        <span class="n">2</span>
        <span class="ic">{{ICON_CLOCK}}</span>
        <h3>Elegí día y hora</h3>
        <p>Contanos qué especialidad necesitás y coordinamos el horario que te quede cómodo.</p>
      </div>
      <div class="step reveal" data-d="2">
        <span class="n">3</span>
        <span class="ic">{{ICON_CHECK}}</span>
        <h3>Confirmado</h3>
        <p>Recibís la confirmación y un recordatorio el día previo. Llegás, te atienden, listo.</p>
      </div>
    </div>
  </div>
</section>

<section class="turno" id="turno">
  <div class="container turno-grid">
    <div class="turno-copy">
      <span class="eyebrow reveal">Reservá ahora</span>
      <h2 class="section-title reveal" data-d="1">Pedí tu turno en un toque</h2>
      <p class="section-lead reveal" data-d="2">Completá el formulario y te preparamos el mensaje listo para enviar. Confirmamos al instante.</p>
      <div class="points">
        <div class="point reveal" data-d="1">
          <span class="ic">{{ICON_CHECK}}</span>
          <div><b>Respuesta en minutos</b><p>En horario de atención coordinamos tu turno al momento.</p></div>
        </div>
        <div class="point reveal" data-d="2">
          <span class="ic">{{ICON_SHIELD}}</span>
          <div><b>Tus datos, seguros</b><p>Nada de papeles ni llamadas perdidas. Todo queda en tu chat.</p></div>
        </div>
      </div>
    </div>
    <form class="form reveal" data-d="1" id="turnoForm" aria-label="Formulario de turno">
      <div class="row">
        <div class="field">
          <label for="fNombre">Tu nombre</label>
          <input type="text" id="fNombre" name="nombre" required placeholder="María López">
        </div>
        <div class="field">
          <label for="fEspecialidad">Especialidad</label>
          <select id="fEspecialidad" name="especialidad">
            {{SPEC_OPTIONS}}
          </select>
        </div>
      </div>
      <div class="row">
        <div class="field">
          <label for="fFecha">Fecha preferida</label>
          <input type="date" id="fFecha" name="fecha" required>
        </div>
        <div class="field">
          <label for="fHora">Hora</label>
          <select id="fHora" name="hora">
            {{HOUR_OPTIONS}}
          </select>
        </div>
      </div>
      <button type="submit" class="btn btn-wa">{{WA_ICON}}Confirmar turno por WhatsApp</button>
      <p class="form-note">Te va a abrir el chat de <b>{{SHORT}}</b> con el mensaje listo. Solo presioná enviar.</p>
    </form>
  </div>
</section>

<section id="contacto">
  <div class="container">
    <div class="section-head">
      <span class="eyebrow reveal">Contacto</span>
      <h2 class="section-title reveal" data-d="1">Estamos acá</h2>
    </div>
    <div class="contact-grid">
      <div class="c-card reveal">
        <span class="ic">{{ICON_PIN}}</span>
        <h3>Dirección</h3>
        <p>{{ADDR}}<br>Mar del Plata, Buenos Aires</p>
        <a href="{{MAPS}}" target="_blank" rel="noopener">Cómo llegar →</a>
      </div>
      <div class="c-card reveal" data-d="1">
        <span class="ic">{{ICON_CLOCK}}</span>
        <h3>Horarios</h3>
        <table class="hours" aria-label="Horarios de atención">{{HOURS_ROWS}}</table>
      </div>
      <div class="c-card reveal" data-d="2">
        <span class="ic">{{ICON_PHONE}}</span>
        <h3>Turnos y consultas</h3>
        <p>{{PHONE_DISPLAY}}<br>WhatsApp directo, respondemos al momento.</p>
        <a href="{{WA_FLOAT}}" target="_blank" rel="noopener">Escribinos →</a>
      </div>
    </div>
  </div>
</section>

<footer class="footer">
  <div class="container">
    <div class="footer-grid">
      <div class="footer-brand">
        <a href="#inicio" class="brand"><span class="mark">{{MARK_ICON}}</span>{{SHORT}}</a>
        <p>Atención de salud de calidad en Mar del Plata, con turnos digitales por WhatsApp y atención cercana.</p>
      </div>
      <div>
        <h4>Accesos</h4>
        <ul>
          <li><a href="#especialidades">Especialidades</a></li>
          <li><a href="#turno">Pedir turno</a></li>
          <li><a href="{{WA_FLOAT}}" target="_blank" rel="noopener">WhatsApp</a></li>
          <li><a href="{{MAPS}}" target="_blank" rel="noopener">Cómo llegar</a></li>
        </ul>
      </div>
    </div>
    <div class="footer-bottom">
      <span>© <span id="year">2026</span> {{SHORT}} · Mar del Plata</span>
      <span>Turnos por WhatsApp · Atención 24/7</span>
    </div>
  </div>
</footer>

<a class="wa-float" href="{{WA_FLOAT}}" target="_blank" rel="noopener" aria-label="Escribinos por WhatsApp">
  <svg viewBox="0 0 24 24" fill="currentColor"><path d="M12 2a10 10 0 0 0-8.6 15L2 22l5.2-1.3A10 10 0 1 0 12 2zm4.8 13.5c-.2.6-1.2 1.1-1.7 1.2-.5.1-.9.2-2.6-.5-2-1-3.4-3.2-3.5-3.4-.1-.2-.9-1.2-.9-2.2s.6-1.6.8-1.8c.2-.2.5-.3.6-.3h.5c.2 0 .4-.1.6.4.2.6.8 1.9.8 2 .1.1.1.2 0 .4l-.4.5c-.1.2-.3.3-.1.6.2.3.8 1.4 1.8 2.2 1.2 1 2 1.3 2.3 1.4.3.1.4.1.6-.1.2-.2.8-.9 1-1.2.2-.3.4-.3.6-.2.3.1 1.7.8 2 1 .3.1.4.2.5.3 0 .1 0 .7-.3 1.3z"/></svg>
</a>

<script>
(function(){
  "use strict";
  var reduce = window.matchMedia("(prefers-reduced-motion: reduce)").matches;

  var nav = document.getElementById("nav");
  function onScroll(){ nav.classList.toggle("is-scrolled", window.scrollY > 40); }
  window.addEventListener("scroll", onScroll, {passive:true});
  onScroll();

  var reveals = document.querySelectorAll(".reveal");
  if (reduce || !("IntersectionObserver" in window)) {
    reveals.forEach(function(el){ el.classList.add("is-in"); });
  } else {
    var io = new IntersectionObserver(function(entries){
      entries.forEach(function(e){
        if (e.isIntersecting) { e.target.classList.add("is-in"); io.unobserve(e.target); }
      });
    }, {threshold:0.12, rootMargin:"0px 0px -6% 0px"});
    reveals.forEach(function(el){ io.observe(el); });
  }

  var burger = document.getElementById("hamburger");
  var links = document.getElementById("navLinks");
  function closeMenu(){ links.classList.remove("is-open"); burger.classList.remove("is-open"); burger.setAttribute("aria-expanded","false"); }
  burger.addEventListener("click", function(){
    var open = links.classList.toggle("is-open");
    burger.classList.toggle("is-open", open);
    burger.setAttribute("aria-expanded", open ? "true" : "false");
  });
  links.querySelectorAll("a").forEach(function(a){ a.addEventListener("click", closeMenu); });
  document.addEventListener("keydown", function(e){ if (e.key === "Escape") closeMenu(); });

  document.getElementById("year").textContent = new Date().getFullYear();

  var form = document.getElementById("turnoForm");
  var fecha = document.getElementById("fFecha");
  var hoy = new Date().toISOString().split("T")[0];
  fecha.min = hoy;
  fecha.value = hoy;

  form.addEventListener("submit", function(e){
    e.preventDefault();
    var nombre = document.getElementById("fNombre").value.trim() || "Paciente";
    var esp = document.getElementById("fEspecialidad").value;
    var texto = "Hola {{SHORT}}! Quiero reservar un turno:" +
      "\nNombre: " + nombre +
      "\nEspecialidad: " + esp +
      "\nFecha: " + fecha.value +
      "\nHora: " + document.getElementById("fHora").value;
    window.open("{{WA_BASE}}?text=" + encodeURIComponent(texto), "_blank", "noopener");
  });
})();
</script>
</body>
</html>
"""

TEMPLATE_COMMERCE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{TITLE}}</title>
<meta name="description" content="{{DESC}}">
<meta name="robots" content="index, follow">
<meta name="geo.region" content="AR-B">
<meta name="geo.placename" content="Mar del Plata">
<meta name="geo.position" content="{{LAT}};{{LNG}}">
<meta name="ICBM" content="{{LAT}}, {{LNG}}">
<meta property="og:title" content="{{TITLE}}">
<meta property="og:description" content="{{DESC}}">
<meta property="og:type" content="website">
<meta property="og:locale" content="es_AR">
<link rel="icon" href="data:image/svg+xml,{{FAVICON}}">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Lora:wght@400;500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<script type="application/ld+json">
{
  "@context": "https://schema.org",
  "@type": "Store",
  "name": "{{SHORT}}",
  "description": "{{DESC}}",
  "telephone": "+54 {{TEL_INTL}}",
  "address": {
    "@type": "PostalAddress",
    "streetAddress": "{{ADDR}}",
    "addressLocality": "Mar del Plata",
    "addressRegion": "Buenos Aires",
    "postalCode": "B7600",
    "addressCountry": "AR"
  },
  "geo": {"@type": "GeoCoordinates", "latitude": "{{LAT}}", "longitude": "{{LNG}}"},
  "areaServed": "Mar del Plata"
}
</script>
<style>
:root{
  --accent:{{ACCENT}};
  --accent-dark:{{ACCENT_DARK}};
  --accent-tint:{{ACCENT_TINT}};
  --ink:#211B16;
  --muted:#6E6257;
  --bg:#FAF6EF;
  --surface:#FFFFFF;
  --line:rgba(33,27,22,.1);
  --radius:18px;
  --ease:cubic-bezier(.22,1,.36,1);
  --serif:"Lora",Georgia,serif;
  --sans:"Inter",system-ui,sans-serif;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
@media (prefers-reduced-motion:reduce){
  html{scroll-behavior:auto}
  *,*::before,*::after{animation-duration:.01ms!important;animation-iteration-count:1!important;transition-duration:.01ms!important}
}
body{font-family:var(--sans);background:var(--bg);color:var(--ink);line-height:1.65;-webkit-font-smoothing:antialiased;overflow-x:hidden}
img{display:block;max-width:100%}
a{color:inherit;text-decoration:none}
ul{list-style:none}
h1,h2,h3{font-family:var(--serif);line-height:1.12;letter-spacing:-.015em}
.container{width:min(1120px,92%);margin-inline:auto}
section{padding:clamp(3.6rem,8vw,6.5rem) 0}
.eyebrow{display:inline-flex;align-items:center;gap:.55rem;font-size:.74rem;font-weight:700;letter-spacing:.22em;text-transform:uppercase;color:var(--accent)}
.eyebrow::before{content:"";width:30px;height:2px;background:var(--accent);border-radius:2px}
.section-title{font-size:clamp(1.8rem,4vw,3rem);margin:.65rem 0 1rem;font-weight:600}
.section-lead{color:var(--muted);max-width:60ch;font-size:clamp(.95rem,1.15vw,1.04rem)}
.section-head{padding-bottom:clamp(2rem,4.5vw,3.2rem)}

/* reveal */
.reveal{opacity:0;transform:translateY(30px);transition:opacity .85s var(--ease),transform .85s var(--ease)}
.reveal.is-in{opacity:1;transform:none}
.reveal[data-d="1"]{transition-delay:.12s}
.reveal[data-d="2"]{transition-delay:.24s}

/* nav */
.nav{position:fixed;inset:0 0 auto;z-index:100;padding:1.05rem 0;transition:background .4s,box-shadow .4s,padding .4s}
.nav.is-scrolled{background:rgba(250,246,239,.9);backdrop-filter:blur(14px);-webkit-backdrop-filter:blur(14px);box-shadow:0 1px 0 var(--line);padding:.65rem 0}
.nav-inner{display:flex;align-items:center;justify-content:space-between;gap:1rem}
.brand{display:flex;align-items:center;gap:.6rem;font-family:var(--serif);font-weight:700;font-size:1.22rem;letter-spacing:-.02em}
.brand .mark{width:38px;height:38px;flex:none;border-radius:12px;background:linear-gradient(145deg,var(--accent),var(--accent-dark));display:grid;place-items:center;color:#fff;box-shadow:0 8px 20px -8px var(--accent)}
.brand .mark svg{width:20px;height:20px}
.nav-links{display:flex;align-items:center;gap:1.7rem}
.nav-links a{font-size:.9rem;font-weight:500;color:var(--ink);position:relative;padding:.2rem 0}
.nav-links a::after{content:"";position:absolute;left:0;bottom:-2px;width:0;height:2px;background:var(--accent);transition:width .35s var(--ease)}
.nav-links a:hover::after{width:100%}
.btn{display:inline-flex;align-items:center;justify-content:center;gap:.55rem;font-family:var(--sans);font-weight:600;font-size:.9rem;padding:.9rem 1.55rem;border-radius:999px;border:none;cursor:pointer;transition:transform .3s var(--ease),box-shadow .3s var(--ease),background .3s var(--ease);white-space:nowrap}
.btn-wa{background:var(--accent);color:#fff;box-shadow:0 14px 30px -14px var(--accent)}
.btn-wa:hover{background:var(--accent-dark);transform:translateY(-2px)}
.btn-ghost{background:transparent;color:var(--ink);box-shadow:inset 0 0 0 1.5px var(--line)}
.btn-ghost:hover{box-shadow:inset 0 0 0 1.5px var(--accent);color:var(--accent);transform:translateY(-2px)}
.btn svg{width:18px;height:18px;flex:none}
.hamburger{display:none;background:none;border:none;cursor:pointer;padding:.4rem}
.hamburger span{display:block;width:26px;height:2px;background:var(--ink);margin:6px 0;transition:transform .35s var(--ease),opacity .2s}

/* hero */
.hero{min-height:100svh;display:flex;flex-direction:column;justify-content:center;position:relative;color:#fff;padding:6.5rem 0 3.2rem;overflow:hidden}
.hero-bg{position:absolute;inset:0;z-index:-2;background:radial-gradient(120% 90% at 85% 10%,{{ACCENT_TINT}} 0%,transparent 55%),linear-gradient(160deg,#241A12 0%,#1C1510 60%,#17110D 100%)}
.hero-bg::after{content:"";position:absolute;inset:0;background-image:radial-gradient(rgba(255,255,255,.05) 1px,transparent 1px);background-size:26px 26px;mask-image:radial-gradient(80% 70% at 50% 40%,#000 30%,transparent 100%);-webkit-mask-image:radial-gradient(80% 70% at 50% 40%,#000 30%,transparent 100%)}
.hero-bg .glow{position:absolute;width:560px;height:560px;border-radius:50%;background:radial-gradient(circle,{{ACCENT}}55 0%,transparent 65%);top:-180px;right:-120px;filter:blur(10px)}
.hero-tag{display:inline-flex;align-items:center;gap:.5rem;font-size:.74rem;font-weight:600;letter-spacing:.24em;text-transform:uppercase;border:1px solid rgba(255,255,255,.22);border-radius:999px;padding:.5rem 1.1rem;margin-bottom:1.6rem;background:rgba(255,255,255,.06);backdrop-filter:blur(8px)}
.hero-tag .dot{width:7px;height:7px;border-radius:50%;background:{{ACCENT}};box-shadow:0 0 0 4px {{ACCENT}}33;animation:pulse 2.2s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.4}}
.hero h1{font-size:clamp(2.5rem,7.5vw,4.8rem);font-weight:600;max-width:14ch;margin-bottom:1.3rem}
.hero p.sub{max-width:52ch;color:rgba(255,255,255,.82);font-size:clamp(1rem,1.35vw,1.12rem);margin-bottom:2.3rem}
.hero-ctas{display:flex;gap:.95rem;flex-wrap:wrap;margin-bottom:3rem}
.hero .btn-wa{background:#25D366;box-shadow:0 14px 30px -14px rgba(37,211,102,.55)}
.hero .btn-wa:hover{background:#1EBE5A}
.hero-pills{display:flex;gap:.7rem;flex-wrap:wrap}
.hero-pill{display:inline-flex;align-items:center;gap:.5rem;font-size:.8rem;font-weight:500;color:rgba(255,255,255,.85);border:1px solid rgba(255,255,255,.16);border-radius:999px;padding:.5rem 1rem;background:rgba(255,255,255,.05)}
.hero-pill svg{width:15px;height:15px;color:{{ACCENT}}}
.scroll-hint{position:absolute;bottom:1.8rem;left:50%;transform:translateX(-50%);display:flex;flex-direction:column;align-items:center;gap:.4rem;color:rgba(255,255,255,.5);font-size:.64rem;letter-spacing:.3em;text-transform:uppercase}
.scroll-hint .wheel{width:1px;height:42px;background:rgba(255,255,255,.2);position:relative;overflow:hidden}
.scroll-hint .wheel::after{content:"";position:absolute;left:0;top:-40%;width:1px;height:40%;background:#fff;animation:drop 1.8s var(--ease) infinite}
@keyframes drop{0%{top:-40%}60%,100%{top:110%}}

/* productos */
.prod-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:1.3rem}
.prod-card{background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:1.6rem;display:flex;flex-direction:column;gap:1rem;transition:transform .4s var(--ease),box-shadow .4s,border-color .4s}
.prod-card:hover{transform:translateY(-6px);box-shadow:0 24px 48px -26px rgba(33,27,22,.25);border-color:{{ACCENT}}55}
.prod-card .head{display:flex;justify-content:space-between;align-items:flex-start;gap:1rem}
.prod-card .ic{width:52px;height:52px;flex:none;border-radius:15px;background:var(--accent-tint);color:var(--accent-dark);display:grid;place-items:center}
.prod-card .ic svg{width:26px;height:26px}
.prod-card .price{font-family:var(--serif);font-size:1.35rem;font-weight:700;color:var(--accent);white-space:nowrap}
.prod-card h3{font-size:1.16rem;font-weight:600}
.prod-card p{font-size:.88rem;color:var(--muted);flex:1}
.prod-card .btn{width:100%;justify-content:center;margin-top:.2rem}
.prod-card .btn:focus-visible{outline:2px solid var(--accent);outline-offset:2px}

/* pasos */
.steps{display:grid;grid-template-columns:repeat(3,1fr);gap:1.3rem}
.step{position:relative;background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:2rem 1.6rem 1.8rem}
.step .n{position:absolute;top:1.2rem;right:1.4rem;font-family:var(--serif);font-size:3.2rem;font-weight:700;color:var(--accent);opacity:.14}
.step .ic{width:48px;height:48px;border-radius:50%;background:var(--accent-tint);color:var(--accent-dark);display:grid;place-items:center;margin-bottom:1.1rem}
.step .ic svg{width:22px;height:22px}
.step h3{font-size:1.12rem;font-weight:600;margin-bottom:.45rem}
.step p{font-size:.88rem;color:var(--muted)}

/* beneficios */
.benefits{background:var(--surface)}
.ben-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:1.3rem}
.ben{background:var(--bg);border:1px solid var(--line);border-radius:var(--radius);padding:1.6rem;transition:transform .4s var(--ease)}
.ben:hover{transform:translateY(-5px)}
.ben .ic{width:46px;height:46px;border-radius:13px;background:linear-gradient(145deg,var(--accent),var(--accent-dark));color:#fff;display:grid;place-items:center;margin-bottom:1rem}
.ben .ic svg{width:22px;height:22px}
.ben h3{font-size:1.05rem;font-weight:600;margin-bottom:.45rem}
.ben p{font-size:.87rem;color:var(--muted)}

/* form pedido */
.pedido{background:var(--surface)}
.pedido-grid{display:grid;grid-template-columns:.9fr 1.1fr;gap:clamp(2.2rem,6vw,4.5rem);align-items:center}
.pedido-copy .points{display:grid;gap:.9rem;margin-top:1.8rem}
.point{display:flex;gap:.95rem;align-items:flex-start}
.point .ic{width:40px;height:40px;flex:none;border-radius:11px;background:var(--accent-tint);color:var(--accent-dark);display:grid;place-items:center}
.point .ic svg{width:19px;height:19px}
.point b{font-size:.95rem;display:block}
.point p{font-size:.85rem;color:var(--muted)}
.form{background:var(--bg);border:1px solid var(--line);border-radius:calc(var(--radius) + 6px);padding:clamp(1.6rem,3vw,2.3rem);display:grid;gap:1.1rem;box-shadow:0 30px 60px -40px rgba(33,27,22,.3)}
.form .row{display:grid;grid-template-columns:1fr 1fr;gap:1rem}
.field{display:grid;gap:.4rem}
.field label{font-size:.72rem;font-weight:700;letter-spacing:.14em;text-transform:uppercase;color:var(--muted)}
.field input,.field select{font-family:var(--sans);font-size:1rem;color:var(--ink);background:var(--surface);border:1px solid var(--line);border-radius:11px;padding:.85rem 1rem;width:100%;transition:border-color .3s;-webkit-appearance:none;appearance:none}
.field input:focus,.field select:focus{outline:none;border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-tint)}
.form .btn{width:100%;justify-content:center;padding:1rem}
.form-note{font-size:.76rem;color:var(--muted);text-align:center}
.form-note b{color:var(--accent)}

/* contacto */
.contact-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:1.3rem}
.c-card{background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:1.7rem;transition:transform .4s var(--ease),box-shadow .4s}
.c-card:hover{transform:translateY(-5px);box-shadow:0 22px 45px -28px rgba(33,27,22,.3)}
.c-card .ic{width:46px;height:46px;border-radius:13px;background:var(--accent-tint);color:var(--accent-dark);display:grid;place-items:center;margin-bottom:1rem}
.c-card .ic svg{width:22px;height:22px}
.c-card h3{font-size:1.08rem;font-weight:600;margin-bottom:.55rem}
.c-card p{font-size:.9rem;color:var(--muted);line-height:1.75}
.c-card a{color:var(--accent);font-weight:600;font-size:.9rem}
.hours{width:100%;border-collapse:collapse;margin-top:.3rem}
.hours td{font-size:.88rem;color:var(--muted);padding:.2rem 0}
.hours td:last-child{text-align:right;font-weight:600;color:var(--ink)}

/* footer */
.footer{background:#1C1510;color:#fff;padding:clamp(2.8rem,6vw,4rem) 0 2rem}
.footer-grid{display:grid;grid-template-columns:1.3fr 1fr;gap:2.2rem;margin-bottom:2.2rem;align-items:start}
.footer .brand .mark{width:34px;height:34px;border-radius:10px}
.footer-brand p{color:rgba(255,255,255,.62);font-size:.9rem;max-width:36ch;margin-top:.9rem}
.footer h4{font-size:.74rem;letter-spacing:.2em;text-transform:uppercase;color:{{ACCENT}};margin-bottom:.9rem}
.footer ul{display:grid;gap:.55rem}
.footer a{color:rgba(255,255,255,.72);font-size:.9rem;transition:color .3s}
.footer a:hover{color:#fff}
.footer-bottom{border-top:1px solid rgba(255,255,255,.1);padding-top:1.4rem;display:flex;justify-content:space-between;gap:1rem;flex-wrap:wrap;font-size:.78rem;color:rgba(255,255,255,.48)}

/* wa float */
.wa-float{position:fixed;right:1.4rem;bottom:1.4rem;z-index:90;width:58px;height:58px;border-radius:50%;background:#25D366;display:grid;place-items:center;box-shadow:0 14px 30px -10px rgba(37,211,102,.55);transition:transform .3s var(--ease)}
.wa-float:hover{transform:scale(1.1) rotate(6deg)}
.wa-float svg{width:30px;height:30px;color:#fff}
.wa-float::before{content:"";position:absolute;inset:0;border-radius:50%;background:#25D366;animation:ring 2.6s var(--ease) infinite;z-index:-1}
@keyframes ring{0%{transform:scale(1);opacity:.55}100%{transform:scale(1.7);opacity:0}}

/* responsive */
@media (max-width:900px){
  .pedido-grid{grid-template-columns:1fr}
  .steps{grid-template-columns:1fr}
  .footer-grid{grid-template-columns:1fr}
}
@media (max-width:640px){
  .nav-links{position:fixed;inset:0 0 auto;z-index:99;flex-direction:column;align-items:flex-start;gap:1.3rem;background:var(--bg);padding:5.5rem 8% 2.2rem;transform:translateY(-100%);transition:transform .5s var(--ease);box-shadow:0 30px 60px -30px rgba(33,27,22,.4)}
  .nav-links.is-open{transform:none}
  .nav-links a{font-size:1.45rem;font-family:var(--serif);font-weight:600;width:100%}
  .nav-links .btn{display:none}
  .mobile-cta{display:inline-flex!important;margin-top:.5rem}
  .hamburger{display:block}
  .hamburger.is-open span:nth-child(1){transform:translateY(8px) rotate(45deg)}
  .hamburger.is-open span:nth-child(2){opacity:0}
  .hamburger.is-open span:nth-child(3){transform:translateY(-8px) rotate(-45deg)}
  .form .row{grid-template-columns:1fr}
  .hero h1{max-width:100%}
  .scroll-hint{display:none}
}
</style>
</head>
<body>

<nav class="nav" id="nav" aria-label="Navegación principal">
  <div class="container nav-inner">
    <a href="#inicio" class="brand"><span class="mark">{{MARK_ICON}}</span>{{SHORT}}</a>
    <ul class="nav-links" id="navLinks">
      <li><a href="#productos">Productos</a></li>
      <li><a href="#como">Cómo comprar</a></li>
      <li><a href="#beneficios">Por qué elegirnos</a></li>
      <li><a href="#contacto">Contacto</a></li>
      <li><a href="{{WA_HERO}}" target="_blank" rel="noopener" class="btn btn-wa mobile-cta">Pedir por WhatsApp</a></li>
    </ul>
    <a href="{{WA_HERO}}" target="_blank" rel="noopener" class="btn btn-wa">Pedir por WhatsApp</a>
    <button class="hamburger" id="hamburger" aria-label="Abrir menú" aria-expanded="false"><span></span><span></span><span></span></button>
  </div>
</nav>

<header class="hero" id="inicio">
  <div class="hero-bg" aria-hidden="true"><div class="glow"></div></div>
  <div class="container">
    <span class="hero-tag reveal"><span class="dot"></span>{{BADGE}}</span>
    <h1 class="reveal" data-d="1">{{SHORT}}</h1>
    <p class="sub reveal" data-d="2">{{TAGLINE}}</p>
    <div class="hero-ctas reveal" data-d="2">
      <a href="{{WA_HERO}}" target="_blank" rel="noopener" class="btn btn-wa">{{WA_ICON}}Pedir por WhatsApp</a>
      <a href="#productos" class="btn btn-ghost" style="color:#fff;box-shadow:inset 0 0 0 1.5px rgba(255,255,255,.28)">Ver productos</a>
    </div>
    <div class="hero-pills reveal" data-d="3">
      <span class="hero-pill">{{ICON_TRUCK}}Envío y retiro en local</span>
      <span class="hero-pill">{{ICON_CHECK}}Respuesta inmediata</span>
      <span class="hero-pill">{{ICON_TAG}}Precio directo</span>
    </div>
  </div>
  <div class="scroll-hint" aria-hidden="true"><span>Bajá</span><div class="wheel"></div></div>
</header>

<section id="productos">
  <div class="container">
    <div class="section-head">
      <span class="eyebrow reveal">{{RU_TITLE}}</span>
      <h2 class="section-title reveal" data-d="1">{{PROD_TITLE}}</h2>
      <p class="section-lead reveal" data-d="2">Tocá el botón de lo que te gusta y te armamos el mensaje listo. Confirmamos stock y precio al instante por WhatsApp.</p>
    </div>
    <div class="prod-grid">
      {{PROD_CARDS}}
    </div>
  </div>
</section>

<section id="como">
  <div class="container">
    <div class="section-head">
      <span class="eyebrow reveal">Así de simple</span>
      <h2 class="section-title reveal" data-d="1">Comprá en 3 pasos</h2>
    </div>
    <div class="steps">
      <div class="step reveal">
        <span class="n">1</span>
        <span class="ic">{{ICON_PHONE}}</span>
        <h3>Escribinos</h3>
        <p>Te abrimos el chat de WhatsApp con el mensaje ya armado. Solo tocá el botón.</p>
      </div>
      <div class="step reveal" data-d="1">
        <span class="n">2</span>
        <span class="ic">{{ICON_CHECK}}</span>
        <h3>Confirmá tu pedido</h3>
        <p>Te pasamos precio final, stock y formas de pago. Sin carritos ni pasarelas complicadas.</p>
      </div>
      <div class="step reveal" data-d="2">
        <span class="n">3</span>
        <span class="ic">{{ICON_TRUCK}}</span>
        <h3>Retirá o recibilo</h3>
        <p>Retirás por el local o coordinamos envío. Rápido, directo y sin sorpresas.</p>
      </div>
    </div>
  </div>
</section>

<section class="benefits" id="beneficios">
  <div class="container">
    <div class="section-head">
      <span class="eyebrow reveal">Por qué elegirnos</span>
      <h2 class="section-title reveal" data-d="1">{{BEN_TITLE}}</h2>
    </div>
    <div class="ben-grid">
      <div class="ben reveal">
        <span class="ic">{{ICON_ZAP}}</span>
        <h3>Atención al momento</h3>
        <p>Sin formularios ni esperas: tu consulta entra directo a nuestro WhatsApp.</p>
      </div>
      <div class="ben reveal" data-d="1">
        <span class="ic">{{ICON_TAG}}</span>
        <h3>Precio directo</h3>
        <p>El precio del local, sin intermediarios. Confirmamos todo antes de que pagues.</p>
      </div>
      <div class="ben reveal" data-d="2">
        <span class="ic">{{ICON_TRUCK}}</span>
        <h3>Retiro o envío</h3>
        <p>Retirá en {{SHORT}} o coordiná el envío a tu zona en el mismo chat.</p>
      </div>
      <div class="ben reveal">
        <span class="ic">{{ICON_USER}}</span>
        <h3>Atención cercana</h3>
        <p>Te atendemos personas reales, con asesoramiento pensado para vos.</p>
      </div>
    </div>
  </div>
</section>

<section class="pedido" id="pedido">
  <div class="container pedido-grid">
    <div class="pedido-copy">
      <span class="eyebrow reveal">Hacé tu pedido</span>
      <h2 class="section-title reveal" data-d="1">Pedilo en un toque</h2>
      <p class="section-lead reveal" data-d="2">Completá el formulario y te armamos el mensaje listo para enviar por WhatsApp.</p>
      <div class="points">
        <div class="point reveal" data-d="1">
          <span class="ic">{{ICON_CHECK}}</span>
          <div><b>Confirmación en minutos</b><p>En horario de atención respondemos al momento.</p></div>
        </div>
        <div class="point reveal" data-d="2">
          <span class="ic">{{ICON_TRUCK}}</span>
          <div><b>Retiro o envío</b><p>Elegís cómo recibir tu pedido en el mismo chat.</p></div>
        </div>
      </div>
    </div>
    <form class="form reveal" data-d="1" id="pedidoForm" aria-label="Formulario de pedido">
      <div class="row">
        <div class="field">
          <label for="fNombre">Tu nombre</label>
          <input type="text" id="fNombre" name="nombre" required placeholder="María López">
        </div>
        <div class="field">
          <label for="fProducto">Producto</label>
          <select id="fProducto" name="producto">
            {{PROD_OPTIONS}}
          </select>
        </div>
      </div>
      <div class="row">
        <div class="field">
          <label for="fCant">Cantidad</label>
          <select id="fCant" name="cantidad">
            {{CANT_OPTIONS}}
          </select>
        </div>
        <div class="field">
          <label for="fEntrega">Entrega</label>
          <select id="fEntrega" name="entrega">
            <option value="Retiro en local" selected>Retiro en local</option>
            <option value="Envío">Envío a coordinar</option>
          </select>
        </div>
      </div>
      <button type="submit" class="btn btn-wa">{{WA_ICON}}Enviar pedido por WhatsApp</button>
      <p class="form-note">Te va a abrir el chat de <b>{{SHORT}}</b> con tu pedido listo. Solo presioná enviar.</p>
    </form>
  </div>
</section>

<section id="contacto">
  <div class="container">
    <div class="section-head">
      <span class="eyebrow reveal">Contacto</span>
      <h2 class="section-title reveal" data-d="1">Estamos acá</h2>
    </div>
    <div class="contact-grid">
      <div class="c-card reveal">
        <span class="ic">{{ICON_PIN}}</span>
        <h3>Dirección</h3>
        <p>{{ADDR}}<br>Mar del Plata, Buenos Aires</p>
        <a href="{{MAPS}}" target="_blank" rel="noopener">Cómo llegar →</a>
      </div>
      <div class="c-card reveal" data-d="1">
        <span class="ic">{{ICON_CLOCK}}</span>
        <h3>Horarios</h3>
        <table class="hours" aria-label="Horarios de atención">{{HOURS_ROWS}}</table>
      </div>
      <div class="c-card reveal" data-d="2">
        <span class="ic">{{ICON_PHONE}}</span>
        <h3>Pedidos y consultas</h3>
        <p>{{PHONE_DISPLAY}}<br>WhatsApp directo, respondemos al momento.</p>
        <a href="{{WA_FLOAT}}" target="_blank" rel="noopener">Escribinos →</a>
      </div>
    </div>
  </div>
</section>

<footer class="footer">
  <div class="container">
    <div class="footer-grid">
      <div class="footer-brand">
        <a href="#inicio" class="brand"><span class="mark">{{MARK_ICON}}</span>{{SHORT}}</a>
        <p>{{FOOTER_DESC}}</p>
      </div>
      <div>
        <h4>Accesos</h4>
        <ul>
          <li><a href="#productos">Productos</a></li>
          <li><a href="#pedido">Hacer un pedido</a></li>
          <li><a href="{{WA_FLOAT}}" target="_blank" rel="noopener">WhatsApp</a></li>
          <li><a href="{{MAPS}}" target="_blank" rel="noopener">Cómo llegar</a></li>
        </ul>
      </div>
    </div>
    <div class="footer-bottom">
      <span>© <span id="year">2026</span> {{SHORT}} · Mar del Plata</span>
      <span>Pedidos por WhatsApp · Retiro en local</span>
    </div>
  </div>
</footer>

<a class="wa-float" href="{{WA_FLOAT}}" target="_blank" rel="noopener" aria-label="Escribinos por WhatsApp">
  <svg viewBox="0 0 24 24" fill="currentColor"><path d="M12 2a10 10 0 0 0-8.6 15L2 22l5.2-1.3A10 10 0 1 0 12 2zm4.8 13.5c-.2.6-1.2 1.1-1.7 1.2-.5.1-.9.2-2.6-.5-2-1-3.4-3.2-3.5-3.4-.1-.2-.9-1.2-.9-2.2s.6-1.6.8-1.8c.2-.2.5-.3.6-.3h.5c.2 0 .4-.1.6.4.2.6.8 1.9.8 2 .1.1.1.2 0 .4l-.4.5c-.1.2-.3.3-.1.6.2.3.8 1.4 1.8 2.2 1.2 1 2 1.3 2.3 1.4.3.1.4.1.6-.1.2-.2.8-.9 1-1.2.2-.3.4-.3.6-.2.3.1 1.7.8 2 1 .3.1.4.2.5.3 0 .1 0 .7-.3 1.3z"/></svg>
</a>

<script>
(function(){
  "use strict";
  var reduce = window.matchMedia("(prefers-reduced-motion: reduce)").matches;

  var nav = document.getElementById("nav");
  function onScroll(){ nav.classList.toggle("is-scrolled", window.scrollY > 40); }
  window.addEventListener("scroll", onScroll, {passive:true});
  onScroll();

  var reveals = document.querySelectorAll(".reveal");
  if (reduce || !("IntersectionObserver" in window)) {
    reveals.forEach(function(el){ el.classList.add("is-in"); });
  } else {
    var io = new IntersectionObserver(function(entries){
      entries.forEach(function(e){
        if (e.isIntersecting) { e.target.classList.add("is-in"); io.unobserve(e.target); }
      });
    }, {threshold:0.12, rootMargin:"0px 0px -6% 0px"});
    reveals.forEach(function(el){ io.observe(el); });
  }

  var burger = document.getElementById("hamburger");
  var links = document.getElementById("navLinks");
  function closeMenu(){ links.classList.remove("is-open"); burger.classList.remove("is-open"); burger.setAttribute("aria-expanded","false"); }
  burger.addEventListener("click", function(){
    var open = links.classList.toggle("is-open");
    burger.classList.toggle("is-open", open);
    burger.setAttribute("aria-expanded", open ? "true" : "false");
  });
  links.querySelectorAll("a").forEach(function(a){ a.addEventListener("click", closeMenu); });
  document.addEventListener("keydown", function(e){ if (e.key === "Escape") closeMenu(); });

  document.getElementById("year").textContent = new Date().getFullYear();

  var form = document.getElementById("pedidoForm");
  form.addEventListener("submit", function(e){
    e.preventDefault();
    var nombre = document.getElementById("fNombre").value.trim() || "Cliente";
    var prod = document.getElementById("fProducto").value;
    var cant = document.getElementById("fCant").value;
    var entrega = document.getElementById("fEntrega").value;
    var texto = "Hola {{SHORT}}! Quiero hacer un pedido:" +
      "\nNombre: " + nombre +
      "\nProducto: " + prod +
      "\nCantidad: " + cant +
      "\nEntrega: " + entrega;
    window.open("{{WA_BASE}}?text=" + encodeURIComponent(texto), "_blank", "noopener");
  });
})();
</script>
</body>
</html>
"""

def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def tint(accent, pct=0.1):
    r, g, b = hex_to_rgb(accent)
    r = int(r + (255 - r) * pct); g = int(g + (255 - g) * pct); b = int(b + (255 - b) * pct)
    return "#%02X%02X%02X" % (r, g, b)

def darken(accent, pct=0.22):
    r, g, b = hex_to_rgb(accent)
    r = int(r * (1 - pct)); g = int(g * (1 - pct)); b = int(b * (1 - pct))
    return "#%02X%02X%02X" % (r, g, b)

def build_page(lead, ov):
    name = str(lead["nombre"]).strip()
    short = ov.get("short", name)
    wa = str(lead["whatsapp"]).strip()
    addr = str(lead["direccion"] or "").strip()
    if not addr:
        addr = "Mar del Plata"
    maps_url = lead["maps"] or ""
    lat, lng = maps_geo(maps_url)

    msg_hero = "Hola %s! Quiero sacar un turno. ¿Qué especialidades tienen disponibles y en qué horarios?" % short
    wa_hero = wa_link(wa, msg_hero)
    wa_base = "https://wa.me/%s" % wa
    wa_float = wa_link(wa, "Hola %s! Quiero hacer una consulta." % short)

    spec_cards = []
    spec_options = []
    for sp_name, sp_desc, ic in ov.get("especialidades", []):
        l = wa_link(wa, "Hola %s! Quiero sacar un turno para %s." % (short, sp_name))
        spec_cards.append(
            '<article class="spec-card reveal">'
            '<span class="ic">%s</span><h3>%s</h3><p>%s</p>'
            '<a href="%s" target="_blank" rel="noopener">Pedir turno <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M5 12h14M13 6l6 6-6 6"/></svg></a>'
            "</article>" % (icon(ic), sp_name, sp_desc, l))
        spec_options.append('<option value="%s">%s</option>' % (sp_name, sp_name))

    hours_rows = "".join("<tr><td>%s</td><td>%s</td></tr>" % (d, h) for d, h in ov.get("horarios", []))

    hour_options = "".join('<option value="%s"%s>%s</option>' % (h, ' selected' if h == "09:00" else "", h)
                           for h in ["08:00","08:30","09:00","09:30","10:00","10:30","11:00","11:30","12:00","14:00","15:00","16:00","17:00","18:00","19:00"])

    accent = ov.get("accent", "#0E7C7B")
    title = "%s — Turnos por WhatsApp | Mar del Plata" % short
    desc = "Agendá tu turno en %s por WhatsApp, sin llamadas ni esperas. %s Mar del Plata." % (short, ov.get("tagline", "")[:110])

    mark_icon = icon(ov.get("mark_icon", "cross"))
    tel_intl = "549" + wa[-6:] if len(wa) > 6 else wa
    favicon = urllib.parse.quote("<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 64 64'><rect width='64' height='64' rx='16' fill='%s'/><path d='M26 18h12v8h8v12h-8v8H26v-8h-8V26h8z' fill='white'/></svg>" % accent)

    tpl = TEMPLATE
    for k, v in {
        "TITLE": title, "DESC": desc, "SHORT": short, "ADDR": addr, "BADGE": ov.get("badge", "Mar del Plata"),
        "TAGLINE": ov.get("tagline", ""), "SUB": ov.get("sub", ""), "ACCENT": accent,
        "ACCENT_DARK": ov.get("accent_dark", darken(accent)), "ACCENT_TINT": tint(accent, 0.1),
        "BEN_TITLE": ov.get("ben_title", "Tu salud, con menos vueltas"),
        "LAT": lat, "LNG": lng, "TEL_INTL": tel_intl, "PHONE_DISPLAY": phone_display(lead["telefono"]),
        "WA_HERO": wa_hero, "WA_BASE": wa_base, "WA_FLOAT": wa_float, "MAPS": maps_url,
        "MARK_ICON": mark_icon, "FAVICON": favicon,
        "SPEC_CARDS": "\n      ".join(spec_cards), "SPEC_OPTIONS": "\n            ".join(spec_options),
        "HOUR_OPTIONS": "\n            ".join(hour_options), "HOURS_ROWS": hours_rows,
        "WA_ICON": '<svg viewBox="0 0 24 24" fill="currentColor"><path d="M12 2a10 10 0 0 0-8.6 15L2 22l5.2-1.3A10 10 0 1 0 12 2zm4.8 13.5c-.2.6-1.2 1.1-1.7 1.2-.5.1-.9.2-2.6-.5-2-1-3.4-3.2-3.5-3.4-.1-.2-.9-1.2-.9-2.2s.6-1.6.8-1.8c.2-.2.5-.3.6-.3h.5c.2 0 .4-.1.6.4.2.6.8 1.9.8 2 .1.1.1.2 0 .4l-.4.5c-.1.2-.3.3-.1.6.2.3.8 1.4 1.8 2.2 1.2 1 2 1.3 2.3 1.4.3.1.4.1.6-.1.2-.2.8-.9 1-1.2.2-.3.4-.3.6-.2.3.1 1.7.8 2 1 .3.1.4.2.5.3 0 .1 0 .7-.3 1.3z"/></svg>',
        "ICON_CHECK": icon("check"), "ICON_ZAP": icon("zap"), "ICON_CLOCK": icon("clock"),
        "ICON_SHIELD": icon("shield"), "ICON_USER": icon("user"), "ICON_PHONE": icon("phone"),
        "ICON_PIN": icon("map-pin"),
    }.items():
        tpl = tpl.replace("{{%s}}" % k, v)

    slug = slugify(name)
    outdir = os.path.join(OUT_ROOT, "%s-%s" % (lead["id"], slug))
    os.makedirs(outdir, exist_ok=True)
    outfile = os.path.join(outdir, "index.html")
    with open(outfile, "w", encoding="utf-8") as f:
        f.write(tpl)
    return outfile, slug

def build_page_commerce(lead, ov):
    name = str(lead["nombre"]).strip()
    short = ov.get("short", name)
    wa = str(lead["whatsapp"]).strip()
    addr = str(lead["direccion"] or "").strip()
    if not addr:
        addr = "Mar del Plata"

    msg_hero = "Hola %s! Quiero hacer un pedido. ¿Qué tienen disponible?" % short
    wa_hero = wa_link(wa, msg_hero)
    wa_base = "https://wa.me/%s" % wa
    wa_float = wa_link(wa, "Hola %s! Quiero hacer una consulta." % short)

    prod_cards = []
    prod_options = []
    for p_name, p_desc, p_price, ic in ov.get("productos", []):
        l = wa_link(wa, "Hola %s! Quiero pedir: %s (%s)." % (short, p_name, p_price))
        prod_cards.append(
            '<article class="prod-card reveal">'
            '<div class="head"><span class="ic">%s</span><span class="price">%s</span></div>'
            "<h3>%s</h3><p>%s</p>"
            '<a href="%s" target="_blank" rel="noopener" class="btn btn-wa">%s</a>'
            "</article>" % (icon(ic), p_price, p_name, p_desc, l,
                            '<svg viewBox="0 0 24 24" fill="currentColor"><path d="M12 2a10 10 0 0 0-8.6 15L2 22l5.2-1.3A10 10 0 1 0 12 2zm4.8 13.5c-.2.6-1.2 1.1-1.7 1.2-.5.1-.9.2-2.6-.5-2-1-3.4-3.2-3.5-3.4-.1-.2-.9-1.2-.9-2.2s.6-1.6.8-1.8c.2-.2.5-.3.6-.3h.5c.2 0 .4-.1.6.4.2.6.8 1.9.8 2 .1.1.1.2 0 .4l-.4.5c-.1.2-.3.3-.1.6.2.3.8 1.4 1.8 2.2 1.2 1 2 1.3 2.3 1.4.3.1.4.1.6-.1.2-.2.8-.9 1-1.2.2-.3.4-.3.6-.2.3.1 1.7.8 2 1 .3.1.4.2.5.3 0 .1 0 .7-.3 1.3z"/></svg>Pedir por WhatsApp'))
        prod_options.append('<option value="%s">%s</option>' % (p_name, p_name))

    hours_rows = "".join("<tr><td>%s</td><td>%s</td></tr>" % (d, h) for d, h in ov.get("horarios", []))
    cant_options = "".join('<option value="%d"%s>%d</option>' % (n, " selected" if n == 1 else "", n) for n in range(1, 11))

    accent = ov.get("accent", "#8B5A2B")
    title = "%s — Pedidos por WhatsApp | Mar del Plata" % short
    desc = "Comprá en %s por WhatsApp: %s Mar del Plata." % (short, ov.get("tagline", "")[:110])

    mark_icon = icon(ov.get("mark_icon", "store"))
    favicon = urllib.parse.quote("<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 64 64'><rect width='64' height='64' rx='16' fill='%s'/><path d='M26 18h12v8h8v12h-8v8H26v-8h-8V26h8z' fill='white'/></svg>" % accent)

    tpl = TEMPLATE_COMMERCE
    for k, v in {
        "TITLE": title, "DESC": desc, "SHORT": short, "ADDR": addr, "BADGE": ov.get("badge", "Mar del Plata"),
        "TAGLINE": ov.get("tagline", ""), "ACCENT": accent,
        "ACCENT_DARK": ov.get("accent_dark", darken(accent)), "ACCENT_TINT": tint(accent, 0.12),
        "BEN_TITLE": ov.get("ben_title", "Comprar directo, sin vueltas"),
        "RU_TITLE": ov.get("ru_title", "Nuestros productos"),
        "PROD_TITLE": ov.get("prod_title", "Elegí lo que te gusta"),
        "FOOTER_DESC": ov.get("footer_desc", "Comprá directo por WhatsApp con atención personalizada y entrega en Mar del Plata."),
        "LAT": "", "LNG": "", "TEL_INTL": "549" + wa[-6:] if len(wa) > 6 else wa,
        "PHONE_DISPLAY": phone_display(lead["telefono"]),
        "WA_HERO": wa_hero, "WA_BASE": wa_base, "WA_FLOAT": wa_float, "MAPS": lead["maps"] or "",
        "MARK_ICON": mark_icon, "FAVICON": favicon,
        "PROD_CARDS": "\n      ".join(prod_cards), "PROD_OPTIONS": "\n            ".join(prod_options),
        "CANT_OPTIONS": "\n            ".join(cant_options), "HOURS_ROWS": hours_rows,
        "WA_ICON": '<svg viewBox="0 0 24 24" fill="currentColor"><path d="M12 2a10 10 0 0 0-8.6 15L2 22l5.2-1.3A10 10 0 1 0 12 2zm4.8 13.5c-.2.6-1.2 1.1-1.7 1.2-.5.1-.9.2-2.6-.5-2-1-3.4-3.2-3.5-3.4-.1-.2-.9-1.2-.9-2.2s.6-1.6.8-1.8c.2-.2.5-.3.6-.3h.5c.2 0 .4-.1.6.4.2.6.8 1.9.8 2 .1.1.1.2 0 .4l-.4.5c-.1.2-.3.3-.1.6.2.3.8 1.4 1.8 2.2 1.2 1 2 1.3 2.3 1.4.3.1.4.1.6-.1.2-.2.8-.9 1-1.2.2-.3.4-.3.6-.2.3.1 1.7.8 2 1 .3.1.4.2.5.3 0 .1 0 .7-.3 1.3z"/></svg>',
        "ICON_CHECK": icon("check"), "ICON_ZAP": icon("zap"), "ICON_CLOCK": icon("clock"),
        "ICON_USER": icon("user"), "ICON_PHONE": icon("phone"), "ICON_PIN": icon("map-pin"),
        "ICON_TRUCK": icon("truck"), "ICON_TAG": icon("tag"),
    }.items():
        tpl = tpl.replace("{{%s}}" % k, v)

    slug = slugify(name)
    outdir = os.path.join(OUT_ROOT, "%s-%s" % (lead["id"], slug))
    os.makedirs(outdir, exist_ok=True)
    outfile = os.path.join(outdir, "index.html")
    with open(outfile, "w", encoding="utf-8") as f:
        f.write(tpl)
    return outfile, slug

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ids", help="IDs separados por coma (default: todos los leads con OVERRIDES)")
    ap.add_argument("--max", type=int, default=100)
    args = ap.parse_args()

    rows = load_rows(max_row=args.max)
    wanted = set(x.strip() for x in args.ids.split(",")) if args.ids else set(OVERRIDES.keys())
    done = []
    for lead in rows:
        lid = lead["id"]
        if lid not in wanted:
            continue
        ov = OVERRIDES.get(lid)
        if not ov:
            print("  ! sin OVERRIDES para %s, salteado" % lid)
            continue
        if ov.get("template") == "comercio":
            outfile, slug = build_page_commerce(lead, ov)
        else:
            outfile, slug = build_page(lead, ov)
        done.append((lid, slug, outfile))
        print("  OK %s -> %s" % (lid, outfile.replace(BASE, "")))
    print("\n%d landings generadas en %s" % (len(done), OUT_ROOT.replace(BASE, "")))

if __name__ == "__main__":
    main()
