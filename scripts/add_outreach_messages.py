# -*- coding: utf-8 -*-
"""Genera el plan de mensajes de prospección para el Excel de 100 prompts.

Por cada lead real del Excel PROSPECCION_100_PROMPTS_OPEN_DESIGN.xlsx:
- Si ya tiene una web real construida (public/<slug>/), usa esa URL.
- Si no, asigna la demo del Demo Factory más cercana a su rubro,
  personalizada con su nombre, teléfono y dirección (?n=&t=&d=).
- Genera el mensaje según GUION_PROSPECCION_FRIO_HOOK y el link wa.me
  con el mensaje precargado.

Uso: python scripts/add_outreach_messages.py
"""
import re
import sys
from pathlib import Path
from urllib.parse import quote

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "PROSPECCION_100_PROMPTS_OPEN_DESIGN.xlsx"
SITE = "https://mis-clientes-html.pages.dev"

# Webs reales ya construidas (public/<slug>/): nombre-keyword -> slug
REAL_PAGES = [
    ("aderezo", "aderezo"),
    ("all design", "all-design"),
    ("antes muerta", "antes-muerta-que-sencilla-showroom"),
    ("bulgarie", "bulgarie"),
    ("cabanas entre los arboles", "cabanas-entre-los-arboles"),
    ("cabanas flemi", "cabanas-entre-los-arboles"),
    ("cardiologia", "cardiologia-dra-ana-gracia-alonso"),
    ("casa-moda", "casa-moda"),
    ("centro-de-dia", "centro-de-dia-lackar-del-sur"),
    ("centro-medico-edison", "centro-medico-edison"),
    ("clinica luro", "clinica-luro"),
    ("consultorio-edison", "consultorio-edison"),
    ("colinas", "consultorios-integrales-colinas"),
    ("medicos luro", "consultorios-medicos-luro"),
    ("consultorios-san-juan", "consultorios-san-juan"),
    ("cordoba-consultorios", "cordoba-consultorios-medicos"),
    ("decor-wall", "decor-wall"),
    ("dimension-atelier", "dimension-atelier"),
    ("espacio edison", "espacio-edison"),
    ("estetica n", "estetica-n"),
    ("humus", "humus-mar-del-plata"),
    ("kiva-cafe", "kiva-cafe"),
    ("krearte", "krearte"),
    ("kytos", "kytos-salud-integral"),
    ("la-isla", "la-isla"),
    ("molduras aguilera", "molduras-aguilera"),
    ("monograf", "monograf"),
    ("oceano mar", "oceano-mar-consultorios"),
    ("patchandflag", "patchandflag"),
    ("restaurante-la-marina", "restaurante-la-marina"),
    ("rompecabezas", "rompecabezas"),
    ("san lorenzo", "san-lorenzo-instituto-medico"),
    ("sanatorio avenida", "sanatorio-avenida"),
    ("terca", "terca"),
    ("turmalina", "turmalina"),
    ("urgencias", "urgencias-odontologicas"),
    ("verde limon", "verde-limon"),
]

# Rubro-keyword -> (slug demo, rubro para el mensaje)
DEMO_RULES = [
    ("contable", "contable-conta-co", "estudio contable"),
    ("repartos", "taller-motorbox", "comercio de repuestos"),
    ("gomeria", "gomeria-rodado-sur", "gomería"),
    ("neumatico", "gomeria-rodado-sur", "gomería"),
    ("lavadero", "lavadero-aquashine", "lavadero de autos"),
    ("lava autos", "lavadero-aquashine", "lavadero de autos"),
    ("repuestos", "taller-motorbox", "comercio de repuestos"),
    ("autopartes", "taller-motorbox", "comercio automotor"),
    ("accesorios para autos", "taller-motorbox", "comercio automotor"),
    ("autos", "taller-motorbox", "comercio automotor"),
    ("cabanas", "cabanas-aires-faro", "cabañas de alquiler"),
    ("cabaña", "cabanas-aires-faro", "cabañas de alquiler"),
    ("alquiler", "cabanas-aires-faro", "alquiler temporario"),
    ("inmobiliaria", "inmobiliaria-costa-real", "inmobiliaria"),
    ("bienes raices", "inmobiliaria-costa-real", "inmobiliaria"),
    ("propiedades", "inmobiliaria-costa-real", "inmobiliaria"),
    ("cerveceria", "cerveceria-punto-cebada", "cervecería"),
    ("cerveza", "cerveceria-punto-cebada", "cervecería"),
    ("beer", "cerveceria-punto-cebada", "cervecería"),
    ("bar ", "cerveceria-punto-cebada", "bar"),
    ("barley", "cerveceria-punto-cebada", "cervecería"),
    ("restaurante", "restaurante-rias", "restaurante"),
    ("cafe", "restaurante-rias", "café/restaurante"),
    ("cuisine", "restaurante-rias", "restaurante"),
    ("vegana", "viandas-sabores", "comida saludable"),
    ("tacc", "viandas-sabores", "comida saludable"),
    ("viandas", "viandas-sabores", "viandas"),
    ("sin gluten", "viandas-sabores", "comida saludable"),
    ("balanceado", "petshop-patitas", "distribuidora de balanceados"),
    ("distribuidora", "distribuidora-mdp", "distribuidora mayorista"),
    ("mayorista", "distribuidora-mdp", "distribuidora mayorista"),
    ("peluche", "regaleria-dulce-detalle", "regalería"),
    ("regaleria", "regaleria-dulce-detalle", "regalería"),
    ("souvenirs", "regaleria-dulce-detalle", "regalería"),
    ("showroom", "showroom-nube", "showroom"),
    ("sweaters", "sport-base9", "indumentaria"),
    ("indumentaria", "sport-base9", "indumentaria"),
    ("moda", "sport-base9", "indumentaria"),
    ("salud", "kinesio-movere", "salud integral"),
    ("medico", "kinesio-movere", "consultorio de salud"),
    ("estetica", "estetica-lumiere", "estética"),
    ("insolita", "estetica-lumiere", "estética"),
    ("meiojas", "estetica-lumiere", "estética"),
    ("consultorio", "kinesio-movere", "consultorio de salud"),
]

DEFAULT_DEMO = ("showroom-nube", "comercio local")

MSG = (
    "Hola {nombre}! Te escribo de Naro AI, hacemos webs para negocios de Mar del Plata \U0001F30A\n"
    "Vi tu {rubro} en {zona} y te prepare una demo en vivo de como se veria tu web, con tu marca y tu numero de WhatsApp:\n"
    "{url}\n"
    "Abrela en el celular y fijate que todo funciona: boton de WhatsApp, direccion y servicios.\n"
    "La activamos hoy? Tiene dominio propio y la podes actualizar cuando quieras. Cualquier cosa me escribis por aca \U0001F64C"
)


def norm_tel(raw: str) -> str:
    d = re.sub(r"\D", "", str(raw or ""))
    if not d:
        return "5492230000000"
    d = d.lstrip("0")
    if not d.startswith("54"):
        d = "54" + d
    return d


def norm(s):
    return (
        s.lower()
        .replace("á", "a").replace("é", "e").replace("í", "i")
        .replace("ó", "o").replace("ú", "u").replace("ñ", "n")
    )


def find_real(nombre: str):
    low = norm(nombre)
    for kw, slug in REAL_PAGES:
        if kw in low:
            return slug
    return None


def find_demo(nombre: str):
    low = norm(nombre)
    for kw, slug, rubro in DEMO_RULES:
        if kw in low:
            return slug, rubro
    return DEFAULT_DEMO


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    for col in range(len(headers) + 1, len(headers) + 6):
        ws.cell(row=1, column=col).value = None

    for i, name in enumerate(
        ("Rubro / Tipo", "URL Web", "Mensaje WhatsApp", "Link WhatsApp", "Demo asignada"), start=6
    ):
        ws.cell(row=1, column=i, value=name)

    n_ok = n_real = n_demo = 0
    for row in ws.iter_rows(min_row=2):
        nombre = row[0].value
        if not nombre or str(nombre).strip().lower().startswith("prompt test"):
            continue
        nombre = str(nombre).strip()
        tel = norm_tel(row[1].value)
        direccion = str(row[2].value or "").strip()
        zona = direccion.split(",")[0] if direccion else "Mar del Plata"

        slug_real = find_real(nombre)
        if slug_real:
            url = f"{SITE}/{slug_real}/"
            rubro_txt = "negocio"
            tipo = "WEB REAL"
            n_real += 1
        else:
            slug, rubro_txt = find_demo(nombre)
            q = quote(f"{nombre}, {direccion}")
            url = f"{SITE}/demos/{slug}.html?n={quote(nombre)}&t={tel}&d={q}"
            tipo = "DEMO"
            n_demo += 1

        mensaje = MSG.format(nombre=nombre, rubro=rubro_txt, zona=zona, url=url)
        link = f"https://wa.me/{tel}?text={quote(mensaje)}"

        row[5].value = f"{tipo} · {rubro_txt}"
        row[6].value = url
        row[7].value = mensaje
        row[8].value = link
        row[9].value = slug_real or slug
        n_ok += 1

    wb.save(XLSX)
    print(f"Leads procesados: {n_ok} (reales: {n_real}, demos: {n_demo})")


if __name__ == "__main__":
    sys.exit(main())
