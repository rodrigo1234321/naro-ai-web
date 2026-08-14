import os
import sys
import re
import csv
import pandas as pd
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")


def clean_phone_str(val):
    if pd.isna(val) or val is None:
        return ""
    val_s = str(val).strip()
    if not val_s or val_s.lower() == "nan":
        return ""
    val_s = re.sub(r"\.0$", "", val_s)
    if "e+" in val_s.lower():
        try:
            val_s = str(int(float(val_s)))
        except Exception:
            pass
    return val_s


def build_naro_official_script(nombre, sector, direccion):
    name_clean = nombre.split(".")[0].split("-")[0].strip() if pd.notna(nombre) else "Comercio"
    if not name_clean or name_clean.lower() == "nan":
        name_clean = "su comercio"

    # Personalización por rubro
    sec_str = str(sector).lower() if pd.notna(sector) else ""
    if "salud" in sec_str or "estética" in sec_str or "estetica" in sec_str:
        solucion = "agendamiento automático de turnos 24/7 y señas"
    elif "showroom" in sec_str or "indumentaria" in sec_str:
        solucion = "catálogo con stock por talle/color y cobro de MercadoPago"
    elif "delivery" in sec_str or "gastronomía" in sec_str or "gastronomia" in sec_str:
        solucion = "menú digital directo a cocina por WhatsApp"
    else:
        solucion = "catálogo web y sistema de gestión de stock/caja"

    script = (
        f"¡Hola! ¿Cómo estás?\n"
        f"Somos NARO AI (https://naro.ai), te escribo porque estamos trabajando con comercios y emprendimientos de Mar del Plata "
        f"ayudándolos a *mejorar su presencia en Internet y facilitar la atención de sus clientes*.\n\n"
        f"Vimos que hoy muchas personas buscan un negocio en Google, consultan por WhatsApp o quieren ver productos/servicios antes de decidirse a comprar. "
        f"Por eso estamos desarrollando soluciones simples para que {name_clean} pueda *recibir más consultas, mostrar lo que ofrece y automatizar parte de la atención* "
        f"(incluyendo {solucion}), sin que tengas que estar pendiente todo el tiempo.\n\n"
        f"Tenemos diferentes opciones según el tipo de negocio, desde una presencia profesional en Google hasta herramientas de *WhatsApp con IA, gestión de stock, pedidos y turnos*. "
        f"Podés ver nuestros trabajos y casos de éxito en nuestra web: https://naro.ai\n\n"
        f"Si te parece, puedo contarte brevemente *qué podríamos implementar en {name_clean} y qué cosas podrías mejorar*. Sin compromiso."
    )
    return script


def main():
    csv_file = "data/prospectos_seleccionados_con_demo.csv"
    if not os.path.exists(csv_file):
        print(f"No existe {csv_file}")
        return

    df = pd.read_csv(csv_file, dtype=str)

    scripts = []
    links = []

    for idx, row in df.iterrows():
        nombre = row.get("nombre", "")
        sector = row.get("sector", "")
        direccion = row.get("direccion", "")
        wa_num = clean_phone_str(row.get("whatsapp", ""))

        sc = build_naro_official_script(nombre, sector, direccion)
        scripts.append(sc)

        if wa_num:
            enc = urllib.parse.quote(sc)
            link = f"https://wa.me/{wa_num}?text={enc}"
        else:
            link = ""
        links.append(link)

    df["script_oficial_naro"] = scripts
    df["link_whatsapp_outreach"] = links

    df.to_csv(csv_file, index=False, encoding="utf-8-sig")
    print(f"[ÉXITO] Actualizado CSV oficial con script NARO AI y link https://naro.ai -> {csv_file}")

    # Regenerar el Excel PROSPECCION_LEADS_CON_DEMO_WHATSAPP.xlsx
    excel_path = "PROSPECCION_LEADS_CON_DEMO_WHATSAPP.xlsx"
    excel_data_path = "data/PROSPECCION_LEADS_CON_DEMO_WHATSAPP.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_hot_header = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    font_data = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_link = Font(name="Calibri", size=10, color="0563C1", underline="single")

    fill_even = PatternFill(start_color="F9FAFC", end_color="F9FAFC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    cols_order = [
        "ranking",
        "nombre",
        "sector",
        "prioridad",
        "puntaje_oportunidad",
        "demo_html_tipo",
        "demo_html_path",
        "script_oficial_naro",
        "apto_gestion_comercial",
        "whatsapp",
        "link_whatsapp_outreach",
        "direccion",
        "zona",
    ]

    def add_sheet(title, dataframe, is_hot=False):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True

        ws.append(cols_order)
        h_fill = fill_hot_header if is_hot else fill_header

        for c_idx in range(1, len(cols_order) + 1):
            cell = ws.cell(row=1, column=c_idx)
            cell.font = font_header
            cell.fill = h_fill
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[1].height = 26

        for r_idx, r_row in enumerate(dataframe[cols_order].values, start=2):
            ws.append(list(r_row))
            ws.row_dimensions[r_idx].height = 22
            r_fill = fill_even if r_idx % 2 == 0 else fill_odd

            for c_idx in range(1, len(cols_order) + 1):
                c = ws.cell(row=r_idx, column=c_idx)
                col_name = cols_order[c_idx - 1]
                c.fill = r_fill
                c.border = thin_border
                c.font = font_data
                c.alignment = align_left

                if col_name in ["whatsapp"]:
                    c.number_format = "@"
                    c.alignment = align_center

                if col_name in ["ranking", "puntaje_oportunidad", "prioridad", "demo_html_tipo"]:
                    c.alignment = align_center

                if col_name in ["link_whatsapp_outreach"] and c.value:
                    v_str = str(c.value)
                    if v_str.startswith("http"):
                        c.hyperlink = v_str
                        c.font = font_link

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    print("Generando solapa principal con Script Oficial NARO AI...")
    add_sheet("🚀 LEADS CON SCRIPT NARO AI", df, is_hot=True)

    df_top100 = df.head(100).copy()
    add_sheet("🔥 TOP 100 CON SCRIPT NARO", df_top100, is_hot=True)

    for p in [excel_path, excel_data_path, "PROSPECCION_LEADS_CON_DEMO_NARO_AI.xlsx"]:
        try:
            wb.save(p)
            print(f"[EXCEL OK] {p}")
        except Exception as e:
            print(f"[WARNING] No se pudo guardar en {p}: {e}")


if __name__ == "__main__":
    main()
