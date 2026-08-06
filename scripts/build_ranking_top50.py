import os
import re
import sys
import unicodedata

import pandas as pd
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
XLSX_PATH = os.path.join(DATA_DIR, "prospeccion_agencia_ia_mdp.xlsx")

SHEET_TODOS = "📊 Todos los Prospectos"

SECTOR_PRIORITY = {
    "Turismo y Alojamiento": 5,
    "Gastronomía": 5,
    "Salud y Estética": 4,
    "Inmobiliarias": 4,
    "Servicios Profesionales": 3,
    "Showrooms e Indumentaria (Instagram)": 2,
    "Comida a Domicilio y Delivery": 2,
    "Automotriz y Servicios": 1,
    "Industrial y Puerto": 1,
    "Comercio General": 1,
}

BAD_PATTERNS = [
    "subaru", "hyundai", "neumaticos", "gomer", "consorcio", "obra social",
    "veteranos", "talabarteria", "distribuidor", "concesionaria", "lifan",
    "faw", "jetour", "kyc", "motos", "repuestos", "transporte", "logistica",
    "fabrica", "peugeot", "ford", "renault", "garbarino", "grido", "freixo",
]


def nombre_polluido(name):
    n = str(name).strip()
    if "." in n:
        return True
    if re.search(r"(?i)^(alquiler|departamento|dpto|casa|depto|monoambiente|departamentos|casas)", n):
        return True
    return False


def normalize(s):
    s = unicodedata.normalize("NFKD", str(s))
    return s.encode("ascii", "ignore").decode("ascii").lower()


def rubro_real(name):
    n = normalize(name)
    if re.search(r"\b(cabana|camping|apart|hotel|hostel|suites|bungalow|hosteria|resort)\w*", n):
        return "Turismo y Alojamiento"
    if re.search(r"\b(parrill|restaurant|bistro|cafeteria|pizzeria|cervecer|helader|delivery|burger|sushi|kitchen|bodegon|comedor|panader)\w*", n):
        return "Gastronomía"
    if re.search(r"\b(clinica|consultor|sanatorio|medic|odonto|estetica|salud|kinesio|fisio|dental|podolog|depila|peluquer|barber|gym|spa|veterin|psicolog|nutricion|cardiologia|pediatria|laboratorio)\w*", n):
        return "Salud y Estética"
    if re.search(r"\b(inmobiliar|propiedades|bienes raices|raices|rental)\w*", n):
        return "Inmobiliarias"
    if re.search(r"\b(estudio contable|contador|abogad|juridic|arquitect|ingenier|asesor)\w*", n):
        return "Servicios Profesionales"
    return None


def digits(s):
    if pd.isna(s):
        return ""
    return re.sub(r"\D", "", str(s))


def build_pool():
    df = pd.read_excel(XLSX_PATH, sheet_name=SHEET_TODOS)
    pool = df[
        df["prioridad"].str.contains("ALTA", na=False)
        & (df["puntaje_oportunidad"] == 70)
        & df["whatsapp"].notna()
    ].copy()

    pool["pol"] = pool["nombre"].apply(nombre_polluido)
    pool["excl"] = (
        pool["nombre"].apply(lambda n: any(b in normalize(n) for b in BAD_PATTERNS))
        | pool["sector"].str.lower().isin(
            ["industrial y puerto", "automotriz y servicios", "comercio general"]
        )
    )
    pool["rubro_real"] = pool["nombre"].apply(rubro_real)
    pool["rubro_ok"] = pool["rubro_real"].notna()
    pool["digits"] = pool["whatsapp"].apply(digits)

    pool = pool[~pool["pol"] & ~pool["excl"] & pool["rubro_ok"]].copy()

    pool = pool.sort_values(
        ["rubro_ok", "rubro_real"],
        ascending=[False, False],
    )

    dedup_keep = []
    seen = set()
    for _, row in pool.iterrows():
        d = row["digits"]
        dup = False
        for k in seen:
            if d and (d in k or k in d):
                dup = True
                break
        if not dup:
            seen.add(d)
            dedup_keep.append(row["id_lead"])
    pool = pool[pool["id_lead"].isin(dedup_keep)].copy()

    pool["prio"] = pool["rubro_real"].map(SECTOR_PRIORITY)
    pool = pool.sort_values(
        ["prio", "rubro_real", "nombre"], ascending=[False, True, True]
    ).reset_index(drop=True)
    return pool


CURATED_TOP10 = [
    "Restaurante La Marina",
    "Centro Médico Edison",
    "La Tosana cabañas",
    "KYTOS Salud Integral",
    "Vistas del Mar Restaurante & Cafe",
    "Los Lirios cabañas",
    "Océano Mar Consultorios",
    "Maktub Cafetería & Bar",
    "Dental Studio Mar del Plata",
    "San Lorenzo Instituto Médico",
]


def order_curated(pool):
    order = {name: i for i, name in enumerate(CURATED_TOP10)}
    pool["cur"] = pool["nombre"].map(order).fillna(len(CURATED_TOP10))
    return pool.sort_values(["cur", "prio"], ascending=[True, False]).reset_index(drop=True)


def balanced_top50(pool):
    pool = order_curated(pool.copy())
    curated = pool[pool["cur"] < len(CURATED_TOP10)].sort_values("cur")
    rest = pool[pool["cur"] >= len(CURATED_TOP10)].copy()

    rest["has_dir"] = rest["direccion"].notna().astype(int)
    rest = rest.sort_values(["prio", "has_dir", "nombre"], ascending=[False, False, True])

    sector_order = ["Turismo y Alojamiento", "Gastronomía", "Salud y Estética", "Inmobiliarias", "Servicios Profesionales"]
    buckets = {s: rest[rest["rubro_real"] == s].copy() for s in sector_order}

    picks = []
    target = 50 - len(curated)
    i = 0
    while len(picks) < target:
        any_added = False
        for s in sector_order:
            b = buckets[s]
            if len(b) > 0:
                row = b.iloc[0]
                buckets[s] = b.iloc[1:]
                row["cur"] = len(CURATED_TOP10) + len(picks)
                picks.append(row)
                any_added = True
                if len(picks) >= target:
                    break
        if not any_added:
            break
    return pd.concat([curated, pd.DataFrame(picks)], ignore_index=True)


def main():
    pool = balanced_top50(build_pool())
    if len(pool) < 50:
        raise SystemExit(f"Pool insuficiente: {len(pool)}")

    top50 = pool.head(50).reset_index(drop=True)
    top10 = pool.head(10).reset_index(drop=True)
    top3 = pool.head(3).reset_index(drop=True)

    wb = openpyxl.load_workbook(XLSX_PATH)
    if "🎯 Ranking Top 50" in wb.sheetnames:
        del wb["🎯 Ranking Top 50"]
    ws = wb.create_sheet(title="🎯 Ranking Top 50")
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_section_top3 = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_section_top10 = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    fill_section_top50 = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    font_section = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_link = Font(name="Calibri", size=10, color="0563C1", underline="single")
    fill_even = PatternFill(start_color="F9FAFC", end_color="F9FAFC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    headers = [
        "Posición",
        "Tier",
        "Nombre",
        "Rubro",
        "Zona",
        "Teléfono",
        "WhatsApp",
        "Link Outreach (wa.me)",
        "Razones de Puntaje",
    ]

    ws.merge_cells("A1:I2")
    ws["A1"] = "🎯 RANKING DE PROSPECTOS PARA VENTA DE SITIO WEB - TOP 3 / 10 / 50"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = align_center

    ws.append([])

    def write_section(title, section_fill, df_sub):
        r_start = ws.max_row + 1
        ws.append([title] + [""] * (len(headers) - 1))
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_start, column=c_idx)
            cell.font = font_section
            cell.fill = section_fill
            cell.alignment = align_left if c_idx == 1 else align_center
        ws.append(headers)
        for c_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=r_start + 1, column=c_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        r = r_start + 2
        for pos, (_, row) in enumerate(df_sub.iterrows(), 1):
            wa = row["whatsapp"]
            wa_str = "" if pd.isna(wa) else str(int(wa))
            link = row["link_whatsapp_outreach"] if pd.notna(row["link_whatsapp_outreach"]) else ""
            vals = [
                pos,
                title,
                str(row["nombre"]),
                str(row["rubro_real"]),
                str(row["zona"]),
                str(row["telefono"])[:30] if pd.notna(row["telefono"]) else "",
                wa_str,
                link,
                str(row["razones_puntaje"]) if pd.notna(row["razones_puntaje"]) else "",
            ]
            ws.append(vals)
            row_fill = fill_even if r % 2 == 0 else fill_odd
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c_idx)
                cell.font = font_bold if pos <= 3 else font_data
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = align_center if c_idx in [1, 2, 4, 5, 6, 7] else align_left
                if c_idx == 8 and link:
                    cell.font = font_link
                    cell.hyperlink = link
            r += 1
        ws.append([])

    write_section("TOP 3 🏆", fill_section_top3, top3)
    write_section("TOP 10 🥇", fill_section_top10, top10)
    write_section("TOP 50 🎯", fill_section_top50, top50)

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 55)
    ws.column_dimensions["H"].width = 90
    ws.column_dimensions["I"].width = 60

    wb.save(XLSX_PATH)
    print(f"[OK] Hoja '🎯 Ranking Top 50' generada en {XLSX_PATH}")
    print("TOP 3:")
    for _, r in top3.iterrows():
        print(f"  - {r['nombre']} | {r['rubro_real']} | {r['telefono']}")
    print("TOP 10:", [str(n)[:30] for n in top10["nombre"]])
    print("Distribución rubro top50:", top50["rubro_real"].value_counts().to_dict())


if __name__ == "__main__":
    main()
