import json
import re
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Ensure UTF-8 output on Windows
sys.stdout.reconfigure(encoding='utf-8')

# Comprehensive templates mapping by sector with OKLCH palettes, font pairings, price ranges, and 6 canonical sections
SECTOR_TEMPLATES = {
    'Salud y Estética': {
        'tema': 'Bioluminiscente Médico & Wellness Esmeralda (Confianza, higiene pulcra, profesionalismo)',
        'paleta': 'Fondo Slate Profundo (oklch(0.15 0.03 240)), Acento Cyan Bioluminiscente (oklch(0.75 0.14 195)), Superficie Cristal Traslúcido con blur(16px), Texto Blanco Nieve (oklch(0.98 0.01 240))',
        'fuentes': 'Outfit (Títulos H1/H2) + Inter (Cuerpo y datos de contacto)',
        'precios': [
            'Consulta Médica / Evaluación Inicial: $18.000 - $25.000 ARS',
            'Tratamiento Especializado / Limpieza / Depilación: $25.000 - $45.000 ARS',
            'Estudios & Chequeos de Diagnóstico: $35.000 - $65.000 ARS'
        ],
        'features': [
            'Catálogo interactivo de tratamientos y especialidades clínicas',
            'Solicitud de turnos online exprés derivado a WhatsApp',
            'Perfil profesional del staff médico matriculado',
            'Mapa interactivo de ubicación y cobertura en Mar del Plata'
        ]
    },
    'Gastronomía': {
        'tema': 'Cálido Gastronómico & Brasas Gourmet (Apetitoso, sabor artesanal, cocina de autor)',
        'paleta': 'Fondo Azul Puerto Profundo (oklch(0.13 0.04 250)), Acento Ámbar Fuego (oklch(0.72 0.16 60)), Superficie Madera Oscura & Cristal Humeante, Texto Blanco Crema (oklch(0.98 0.01 60))',
        'fuentes': 'Playfair Display (Títulos Gastronómicos) + Plus Jakarta Sans (Cuerpo)',
        'precios': [
            'Plato Principal Gourmet / Cazuela / Parrillada: $18.000 - $28.000 ARS',
            'Tabla Completa de Mariscos / Picada para 2: $38.000 - $52.000 ARS',
            'Bebida / Vinos de Cava Selección Botella: $12.000 - $24.000 ARS'
        ],
        'features': [
            'Menú digital con fotos tentadoras en alta resolución',
            'Sistema de reservas de mesa y pedidos por WhatsApp',
            'Pizarra de recomendación del chef y maridaje',
            'Ubicación local y zona de envíos en Mar del Plata'
        ]
    },
    'Comida a Domicilio y Delivery': {
        'tema': 'Delivery Rápido & Urban Fast Food (Tentación inmediata, velocidad de entrega)',
        'paleta': 'Fondo Carbón Rotisería (oklch(0.14 0.02 20)), Acento Rojo Tomate Fuego (oklch(0.62 0.22 30)), Superficie Negra Mate con Acentos Amarillos, Texto Blanco Arroz (oklch(0.98 0.01 20))',
        'fuentes': 'Space Grotesk (Títulos Impactantes) + Plus Jakarta Sans (Cuerpo)',
        'precios': [
            'Milanesa XL a la Napolitana con Fritas (Para 2): $18.000 ARS',
            'Pizza Grande Muzzarella de Molde: $12.500 ARS',
            'Docena de Empanadas Cortadas a Cuchillo: $14.000 ARS'
        ],
        'features': [
            'Menú categorizado por pestañas con promos destacadas',
            'Botón de pedido exprés en 2 clics directo por WhatsApp',
            'Indicador claro de tiempos de entrega y cobertura',
            'Combos familiares económicos del día'
        ]
    },
    'Showrooms e Indumentaria (Instagram)': {
        'tema': 'Dark Luxury & High Fashion Glass (Exclusividad, tendencias de moda, estética instagramera)',
        'paleta': 'Fondo Negro Nube (oklch(0.14 0.02 350)), Acento Cuarzo Rosa Trendy (oklch(0.82 0.12 350)), Superficie Cristal Esmerilado (backdrop-filter blur(20px)), Texto Seda Pura (oklch(0.98 0.01 350))',
        'fuentes': 'Tenor Sans (Títulos Elegantes) + Plus Jakarta Sans (Cuerpo)',
        'precios': [
            'Vestidos / Sacos / Prendas de Noche: $45.000 - $85.000 ARS',
            'Blusas / Tops de Seda & Lino: $22.000 - $38.000 ARS',
            'Jeans Cargo / Calce Perfecto: $38.000 - $55.000 ARS'
        ],
        'features': [
            'Lookbook en formato bento grid estilo Instagram Stories',
            'Guía de talles, telas y calce personalizado',
            'Asesoramiento de imagen directo por WhatsApp',
            'Envíos locales en Mar del Plata y a todo el país'
        ]
    },
    'Inmobiliarias': {
        'tema': 'Arquitectura Obsidian & Geometría Dorada (Prestigio patrimonial, claridad de inversión)',
        'paleta': 'Fondo Obsidiana Patrimonial (oklch(0.14 0.03 240)), Acento Dorado Champán (oklch(0.78 0.12 85)), Superficie Cristal Arquitectónico, Texto Blanco Edificio (oklch(0.98 0.01 240))',
        'fuentes': 'Cinzel / Montserrat (Títulos) + Inter (Cuerpo)',
        'precios': [
            'Departamento 2 Ambientes Güemes / Macrocentro: USD 65.000 - USD 110.000',
            'Departamento 3 Ambientes Vista al Mar Playa Grande: USD 140.000 - USD 280.000',
            'Tasación Profesional de Inmuebles: En el día / Sin cargo'
        ],
        'features': [
            'Buscador de inmuebles con filtros por zona (Güemes, Playa Grande, Los Troncos)',
            'Fichas detalladas con ambientes, fotos y amenities',
            'Solicitud de tasación rápida por WhatsApp',
            'Equipo de corredores matriculados'
        ]
    },
    'Turismo y Alojamiento': {
        'tema': 'Turquesa Costero & Sol Salino (Vacaciones soñadas frente al mar y la sierra)',
        'paleta': 'Fondo Azul Noche Costero (oklch(0.15 0.04 230)), Acento Turquesa MDP (oklch(0.72 0.14 210)), Superficie Arena Dorada (oklch(0.97 0.02 85)), Texto Blanco Sal (oklch(0.99 0.01 230))',
        'fuentes': 'Outfit (Títulos) + DM Sans (Cuerpo)',
        'precios': [
            'Alquiler Diario Depto Frente al Mar (Temporada): $65.000 - $95.000 ARS / noche',
            'Cabaña Premium con Hidromasaje & Parrilla: $75.000 - $110.000 ARS / noche',
            'Habitación Hotel Boutique con Desayuno: $85.000 - $125.000 ARS / noche'
        ],
        'features': [
            'Galería de fotos inmersiva de departamentos y cabañas',
            'Amenities incluidos (Wi-Fi, Cochera, Vista al Mar, Parrilla)',
            'Calendario de disponibilidad y tarifas por noche',
            'Reserva de fechas en directo por WhatsApp'
        ]
    },
    'Automotriz y Servicios': {
        'tema': 'Cyber Metallic & High Performance (Mecánica de precisión, brillo cerámico y seguridad)',
        'paleta': 'Fondo Titanio Oscuro (oklch(0.14 0.03 240)), Acento Neón Cyan / Rojo Deportivo (oklch(0.62 0.22 25)), Superficie Textura Fibra Carbón, Texto Blanco Acero (oklch(0.98 0.01 240))',
        'fuentes': 'Space Grotesk / Outfit (Títulos) + Inter (Cuerpo)',
        'precios': [
            'Service Aceite Sintético & 4 Filtros: $48.000 - $75.000 ARS',
            'Alineación 3D + Balanceo x4 Ruedas: $28.000 - $38.000 ARS',
            'Tratamiento Cerámico / Detail Carrocería: $95.000 - $160.000 ARS'
        ],
        'features': [
            'Catálogo técnico de servicios de mecánica y lubricentro',
            'Diagnóstico computarizado de inyección con escáner',
            'Presupuesto exprés sin cargo enviado a WhatsApp',
            'Ubicación del taller y turnos de atención'
        ]
    },
    'Comercio General': {
        'tema': 'Modern Retail & Glassmorphism Pro (Variedad, atención cercana, confianza comercial)',
        'paleta': 'Fondo Oscuro Elegante (oklch(0.14 0.02 240)), Acento Azul Cobalto / Violeta (oklch(0.68 0.20 260)), Superficie Cristal Traslúcido, Texto Blanco Puro (oklch(0.98 0.01 240))',
        'fuentes': 'Plus Jakarta Sans (Títulos) + Inter (Cuerpo)',
        'precios': [
            'Productos Destacados de Selección: $8.500 - $25.000 ARS',
            'Artículos Especiales / Packs Comercial: $28.000 - $55.000 ARS',
            'Envíos a Domicilio en Mar del Plata: Sin cargo en compras seleccionadas'
        ],
        'features': [
            'Catálogo de productos destacados con fotos y precios',
            'Medios de pago, promociones y cuotas bancarias',
            'Atención inmediata por WhatsApp',
            'Ubicación del local comercial y horarios de apertura'
        ]
    },
    'Servicios Profesionales': {
        'tema': 'Corporate Tech & Tranquilidad Fiscal (Autoridad, soluciones efectivas, rigor profesional)',
        'paleta': 'Fondo Azul Ejecutivo (oklch(0.14 0.03 240)), Acento Verde Crecimiento (oklch(0.68 0.18 140)), Superficie Cristal Financiero, Texto Blanco Balance (oklch(0.98 0.01 240))',
        'fuentes': 'Outfit / Cormorant Garamond (Títulos) + Inter (Cuerpo)',
        'precios': [
            'Abono Mensual Asesoría / Monotributo / Pymes: $18.000 - $65.000 ARS / mes',
            'Servicios Específicos / Honorarios Profesionales: $35.000 - $95.000 ARS',
            'Diagnóstico Inicial de Situación: Sin cargo'
        ],
        'features': [
            'Áreas de práctica y servicios profesionales de consultoría',
            'Propuesta de valor y casos de éxito reales',
            'Agendamiento de diagnóstico previo vía WhatsApp',
            'Ubicación de oficinas y canales de atención'
        ]
    }
}

def clean_phone(phone_str):
    if not phone_str or str(phone_str).lower() in ['none', 'nan', '']:
        return '+54 9 223 555-0199'
    raw = re.sub(r'\D', '', str(phone_str))
    if not raw:
        return '+54 9 223 555-0199'
    if raw.startswith('549'):
        return f'+{raw}'
    elif raw.startswith('223') or raw.startswith('0223'):
        clean_num = raw.lstrip('0')
        return f'+54 9 {clean_num[:3]} {clean_num[3:6]}-{clean_num[6:]}' if len(clean_num) >= 9 else f'+54 9 {clean_num}'
    else:
        return f'+54 9 {raw}'

def build_ultra_opendesign_prompt(lead):
    nombre = lead.get('nombre', 'Comercio Local').strip()
    direccion = lead.get('direccion', 'Mar del Plata').strip()
    telefono = clean_phone(lead.get('telefono') or lead.get('whatsapp'))
    wa_num = re.sub(r'\D', '', telefono)
    wa_link = f"https://wa.me/{wa_num}?text=Hola%20{nombre.replace(' ', '%20')}%2C%20quisiera%20consultar%20por%20servicios"
    sector_raw = lead.get('sector') or 'Comercio General'
    
    # Match template
    matching_key = 'Comercio General'
    for k in SECTOR_TEMPLATES.keys():
        if k.lower() in sector_raw.lower() or sector_raw.lower() in k.lower():
            matching_key = k
            break
            
    tmpl = SECTOR_TEMPLATES[matching_key]

    prompt = (
        f"Usá la skill landing-web-opendesign para crear la landing page web profesional de alta conversión de {nombre}.\n\n"
        f"DATOS REALES DEL NEGOCIO (INNEGOCIABLES):\n"
        f"- Nombre Comercial: {nombre}\n"
        f"- Rubro / Sector: {sector_raw}\n"
        f"- Dirección Física: {direccion}, Mar del Plata, Buenos Aires [VERIFICADO]\n"
        f"- Contacto WhatsApp Directo: {telefono} (Link directo wa.me: {wa_link})\n\n"
        f"DIRECCIÓN ARTÍSTICA & ESTÉTICA VISUAL (ANTI-SLOP):\n"
        f"- Atmósfera Temática: {tmpl['tema']}\n"
        f"- Paleta de Colores OKLCH: {tmpl['paleta']}\n"
        f"- Regla Tipográfica: {tmpl['fuentes']} (Máximo 2 familias tipográficas en todo el proyecto)\n"
        f"- Tipografía H1 Contenida: Escala anti-gigantismo clamp(2.2rem, 3.8vw, 3.2rem)\n"
        f"- Paneles Glassmorphic: Tarjetas con cristal traslúcido (backdrop-filter: blur(14px - 20px)) y bordes sutiles de cristal\n"
        f"- Cero gradientes pastel flotantes generados por inercia\n\n"
        f"ESTRUCTURA DE SECCIONES OBLIGATORIAS (6 SECCIONES CANÓNICAS):\n"
        f"1. HERO SECTION: Título impactante para {nombre}, bajada comercial persuasiva, badge de ubicación ({direccion}) y botón principal 'Contactar por WhatsApp'.\n"
        f"2. CATÁLOGO / SERVICIOS DESTACADOS: Grid bento interactivo presentando soluciones clave ({', '.join(tmpl['features'][:2])}).\n"
        f"3. PROPUESTA DE VALOR / POR QUÉ ELEGIRNOS: 3 a 4 pilares diferenciales con micro-animaciones hover y elevación tonal.\n"
        f"4. PRECIOS & TARIFAS DE REFERENCIA EN MDP: Tarjetas limpias con precios del sector ({', '.join(tmpl['precios'][:2])}).\n"
        f"5. UBICACIÓN & HORARIOS: Mapa interactivo de llegada a {direccion} e información de atención al cliente.\n"
        f"6. CTA FLOTANTE Y FOOTER: Botón persistente de atención directa por WhatsApp con el texto exacto 'WhatsApp' (nunca abreviar como 'WA').\n\n"
        f"INSTRUCCIONES DE EJECUCIÓN MCP PARA OPEN DESIGN:\n"
        f"- Crear el proyecto con create_project('{nombre.lower().replace(' ', '-')}-landing')\n"
        f"- Ejecutar collect_brief y confirm_brief con este paquete de contexto enriquecido\n"
        f"- Lanzar start_run exigiendo diseño 100% responsive en smartphones, copywriting rioplatense profesional y animación fluida en scroll con Lenis y GSAP."
    )
    return prompt

def main():
    print("🚀 Generando los 100 Prompts ULTRA DETALLADOS sin pérdida de calidad...")
    
    catalog_path = os.path.join('data', 'leads_catalog.json')
    if not os.path.exists(catalog_path):
        print(f"Error: {catalog_path} no existe.")
        sys.exit(1)
        
    with open(catalog_path, 'r', encoding='utf-8') as f:
        catalog_data = json.load(f)
        
    by_id = catalog_data.get('by_id', {})
    
    all_leads = []
    seen_names = set()
    
    for lid, item in by_id.items():
        nombre = (item.get('nombre') or '').strip()
        direccion = (item.get('direccion') or '').strip()
        telefono = str(item.get('whatsapp') or item.get('telefono') or '').strip()
        sector = (item.get('sector') or item.get('rubro') or 'Comercio General').strip()
        
        if nombre and nombre.lower() not in seen_names and telefono not in ['None', 'nan', ''] and direccion not in ['None', 'nan', '']:
            if len(nombre) > 2 and len(direccion) > 4:
                seen_names.add(nombre.lower())
                all_leads.append({
                    'id': lid,
                    'nombre': nombre,
                    'direccion': direccion,
                    'telefono': telefono,
                    'whatsapp': item.get('whatsapp') or telefono,
                    'sector': sector,
                    'puntaje': item.get('puntaje_oportunidad') or 80
                })

    print(f"Total de leads válidos encontrados: {len(all_leads)}")
    
    all_leads.sort(key=lambda x: float(x.get('puntaje') or 0), reverse=True)
    
    selected_leads = []
    by_sector = {}
    for lead in all_leads:
        sec = lead['sector']
        by_sector.setdefault(sec, []).append(lead)
        
    sector_keys = list(by_sector.keys())
    while len(selected_leads) < 100 and any(by_sector.values()):
        for sec in sector_keys:
            if len(selected_leads) >= 100:
                break
            if by_sector[sec]:
                selected_leads.append(by_sector[sec].pop(0))

    print(f"Seleccionados exactamente {len(selected_leads)} leads.")
    
    excel_rows = []
    md_content = "# Catálogo Maestro de 100 Prompts Ultra Detallados para Open Design\n\n"
    
    for idx, lead in enumerate(selected_leads, start=1):
        prompt_text = build_ultra_opendesign_prompt(lead)
        num_contacto = clean_phone(lead.get('whatsapp') or lead.get('telefono'))
        ubicacion = f"{lead['direccion']}, Mar del Plata"
        
        excel_rows.append({
            'Local': lead['nombre'],
            'Número': num_contacto,
            'Dónde está': ubicacion,
            'Prompt Open Design': prompt_text
        })
        
        md_content += f"### {idx}. {lead['nombre']}\n"
        md_content += f"- **Número**: `{num_contacto}`\n"
        md_content += f"- **Dónde está**: {ubicacion}\n"
        md_content += f"- **Prompt Open Design**:\n```text\n{prompt_text}\n```\n\n---\n\n"
        
    # Write 100_PROMPTS_COMPLETOS.md
    with open('100_PROMPTS_COMPLETOS.md', 'w', encoding='utf-8') as f:
        f.write(md_content)
    print("✅ Creado: 100_PROMPTS_COMPLETOS.md con 100 prompts idénticos en máxima calidad.")
    
    # Save Excel files
    data_output_path = os.path.join('data', 'PROSPECCION_100_PROMPTS_OPEN_DESIGN.xlsx')
    root_output_path = 'PROSPECCION_100_PROMPTS_OPEN_DESIGN.xlsx'
    
    for target_file in [data_output_path, root_output_path]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "100 Prompts Open Design"
        
        ws.views.sheetView[0].showGridLines = True
        
        headers = ['Local', 'Número', 'Dónde está', 'Prompt Open Design']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        for col_num, h_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            
        ws.row_dimensions[1].height = 28
        
        font_regular = Font(name="Arial", size=10, color="0F172A")
        font_bold = Font(name="Arial", size=10, bold=True, color="0F172A")
        font_phone = Font(name="Consolas", size=10, color="0369A1")
        align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        align_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
        
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for row_idx, row_data in enumerate(excel_rows, start=2):
            c_local = ws.cell(row=row_idx, column=1, value=row_data['Local'])
            c_num = ws.cell(row=row_idx, column=2, value=row_data['Número'])
            c_ubic = ws.cell(row=row_idx, column=3, value=row_data['Dónde está'])
            c_prompt = ws.cell(row=row_idx, column=4, value=row_data['Prompt Open Design'])
            
            c_local.font = font_bold
            c_local.alignment = align_top_left
            
            c_num.font = font_phone
            c_num.alignment = align_top_center
            
            c_ubic.font = font_regular
            c_ubic.alignment = align_top_left
            
            c_prompt.font = font_regular
            c_prompt.alignment = align_top_left
            
            fill = zebra_fill if row_idx % 2 == 0 else white_fill
            for c in [c_local, c_num, c_ubic, c_prompt]:
                c.fill = fill
                c.border = thin_border
                
            ws.row_dimensions[row_idx].height = 180  # Generous height for reading full ultra prompts
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 90
        
        wb.save(target_file)
        print(f"✅ Archivo Excel ULTRA guardado en: {target_file}")

if __name__ == '__main__':
    main()
