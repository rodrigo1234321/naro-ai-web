import json
import re
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Ensure stdout handles UTF-8 on Windows
sys.stdout.reconfigure(encoding='utf-8')

# Sector visual themes & instructions mapping
SECTOR_TEMPLATES = {
    'Gastronomía': {
        'tema': 'Cálido Gastronómico & Brasas',
        'paleta': 'Oscuro rústico con acentos ámbar fuego (oklch(0.68 0.18 45)), superficie madera oscura y cristal humeante',
        'sensacion': 'Apetitoso, artesanal, alta conversión de pedidos y reservas',
        'icono_hero': '🍔 / 🍕 / ☕',
        'features': ['Menú digital interactivo con fotos', 'Sistema de pedidos por WhatsApp con cálculo automático', 'Horarios de cocina y zona de delivery', 'Mapa interactivo de ubicación local']
    },
    'Comida a Domicilio y Delivery': {
        'tema': 'Delivery Rápido & Urban Fast Food',
        'paleta': 'Fondo ultra oscuro con acentos naranja quemado u oro (oklch(0.72 0.22 50)) y superficies traslúcidas',
        'sensacion': 'Velocidad, frescura, tentación inmediata',
        'icono_hero': '🛵 / 🍕 / 🌮',
        'features': ['Combos destacados y promos del día', 'Buscador/filtrado rápido de menú', 'Botón directo de pedido exprés por WhatsApp', 'Información clara de tiempos de envío']
    },
    'Salud y Estética': {
        'tema': 'Bioluminiscente Médico & Wellness Esmeralda',
        'paleta': 'Fondo slate profundo con acentos cyan médico (oklch(0.75 0.14 195)) y verde esmeralda traslúcido',
        'sensacion': 'Confianza, higiene impecable, bienestar, profesionalismo',
        'icono_hero': '🩺 / ✨ / 💆‍♀️',
        'features': ['Catálogo de tratamientos / especialidades clínicas', 'Solicitud de turnos online vía WhatsApp', 'Perfiles del equipo profesional', 'Preguntas frecuentes y preparación previa']
    },
    'Showrooms e Indumentaria (Instagram)': {
        'tema': 'Dark Luxury & High Fashion Glass',
        'paleta': 'Negro azabache con acentos oro rosa / cuarzo (oklch(0.82 0.12 350)) y paneles de cristal esmerilado',
        'sensacion': 'Exclusividad, tendencia, estética visual instagramera',
        'icono_hero': '👗 / 👠 / ✨',
        'features': ['Lookbook / Galería de nueva colección', 'Tabla de talles y envíos a todo el país', 'Atención directa por WhatsApp para asesoramiento personalizado', 'Grid bento con tendencias']
    },
    'Inmobiliarias': {
        'tema': 'Arquitectura Obsidian & Geometría Dorada',
        'paleta': 'Obsidiana profunda con acentos dorado champán (oklch(0.78 0.12 85)) e interacciones limpias',
        'sensacion': 'Solidez, prestigio patrimonial, claridad de inversión',
        'icono_hero': '🏢 / 🔑 / 🏘️',
        'features': ['Buscador de propiedades destacadas (Venta / Alquiler)', 'Fichas detalladas con ambientes y amenities', 'Consulta directa de tasación / visita por WhatsApp', 'Calculadora / formulario de contacto']
    },
    'Turismo y Alojamiento': {
        'tema': 'Ecovía Cálida & Naturaleza Turística',
        'paleta': 'Marrón bosque y acentos luz solar filtrada (oklch(0.75 0.15 130)) con cristal sereno',
        'sensacion': 'Desconexión, confort, experiencia costera/sierra inolvidable',
        'icono_hero': '🏖️ / 🌲 / 🏡',
        'features': ['Galería inmersiva de instalaciones y habitaciones', 'Disponibilidad y reservas directas por WhatsApp', 'Guía de experiencias y atractivos cercanos', 'Mapa de llegada y servicios incluidos']
    },
    'Comercio General': {
        'tema': 'Modern Retail & Glassmorphism Pro',
        'paleta': 'Modo oscuro elegante con acento azul cobalto / violeta (oklch(0.68 0.20 260))',
        'sensacion': 'Variedad, atención cercana, confianza comercial',
        'icono_hero': '🛍️ / 🏬 / 📦',
        'features': ['Catálogo de productos destacados con precios', 'Atención inmediata por WhatsApp', 'Medios de pago y promociones bancarias', 'Ubicación del local y horarios comerciales']
    },
    'Automotriz y Servicios': {
        'tema': 'Cyber Metallic & High Performance',
        'paleta': 'Gris titanio con acentos rojo deportivo o azul eléctrico (oklch(0.65 0.22 25))',
        'sensacion': 'Potencia, precisión técnica, respuesta garantizada',
        'icono_hero': '🚗 / 🔧 / 🛞',
        'features': ['Listado de servicios de taller / repuestos / vehículos', 'Presupuesto rápido en 1 clic por WhatsApp', 'Garantía de servicio y testimonios de clientes', 'Ubicación y turnos de atención']
    },
    'Servicios Profesionales': {
        'tema': 'Corporate Tech & Minimalist Precision',
        'paleta': 'Azul noche con acentos blanco puro y cian (oklch(0.70 0.16 220))',
        'sensacion': 'Autoridad, soluciones efectivas, rigor profesional',
        'icono_hero': '⚖️ / 📊 / 💼',
        'features': ['Áreas de práctica / servicios de consultoría', 'Casos de éxito y propuesta de valor', 'Agendamiento de diagnóstico previo vía WhatsApp', 'Ubicación de oficinas']
    }
}

def clean_phone(phone_str):
    if not phone_str or str(phone_str).lower() in ['none', 'nan', '']:
        return 'Consulte por WhatsApp'
    raw = re.sub(r'\D', '', str(phone_str))
    if not raw:
        return str(phone_str)
    if raw.startswith('549'):
        return f'+{raw}'
    elif raw.startswith('223') or raw.startswith('0223'):
        clean_num = raw.lstrip('0')
        return f'+54 9 {clean_num[:3]} {clean_num[3:6]}-{clean_num[6:]}' if len(clean_num) >= 9 else f'+54 9 {clean_num}'
    else:
        return f'+54 9 {raw}'

def build_opendesign_prompt(lead):
    nombre = lead.get('nombre', 'Comercio Local').strip()
    direccion = lead.get('direccion', 'Mar del Plata').strip()
    telefono = clean_phone(lead.get('telefono') or lead.get('whatsapp'))
    wa_num = re.sub(r'\D', '', telefono)
    wa_link = f"https://wa.me/{wa_num}" if wa_num else "#"
    sector_raw = lead.get('sector') or 'Comercio General'
    
    matching_key = 'Comercio General'
    for k in SECTOR_TEMPLATES.keys():
        if k.lower() in sector_raw.lower() or sector_raw.lower() in k.lower():
            matching_key = k
            break
            
    tmpl = SECTOR_TEMPLATES[matching_key]

    prompt = (
        f"Usá la skill landing-web-opendesign para crear la landing page web profesional de alta conversión de {nombre}.\n\n"
        f"DATOS REALES DEL NEGOCIO (INNEGOCIABLES):\n"
        f"- Nombre Comercial: {nombre}\n"
        f"- Rubro / Sector: {sector_raw}\n"
        f"- Dirección Física: {direccion}, Mar del Plata, Buenos Aires\n"
        f"- Contacto WhatsApp Directo: {telefono} (Link directo wa.me: {wa_link})\n\n"
        f"DIRECCIÓN ARTÍSTICA & ESTÉTICA VISUAL (ANTI-SLOP):\n"
        f"- Atmósfera Temática: {tmpl['tema']} ({tmpl['sensacion']})\n"
        f"- Paleta de Colores: {tmpl['paleta']}\n"
        f"- Tipografía: Máximo 2 familias tipográficas (Inter/Outfit para cuerpo, Montserrat/Playfair para títulos)\n"
        f"- Tipografía H1 Contenida: Usar escala anti-gigantismo clamp(2.2rem, 3.8vw, 3.2rem)\n"
        f"- Paneles Glassmorphic: Tarjetas con cristal traslúcido (backdrop-filter: blur(14px)) y bordes sutiles\n"
        f"- Cero gradientes pastel flotantes generados por inercia\n\n"
        f"ESTRUCTURA DE SECCIONES OBLIGATORIAS:\n"
        f"1. HERO SECTION: Título impactante para {nombre}, bajada comercial, badge de ubicación ({direccion}) y botón principal de conversión 'Contactar por WhatsApp'.\n"
        f"2. CATÁLOGO / SERVICIOS DESTACADOS: Grid interactivo presentando las principales soluciones del local ({', '.join(tmpl['features'][:2])}).\n"
        f"3. PROPUESTA DE VALOR / POR QUÉ ELEGIRNOS: 3 a 4 pilares diferenciales de la marca con micro-animaciones hover.\n"
        f"4. UBICACIÓN & ATENCIÓN: Mapa interactivo / indicador claro de cómo llegar a {direccion} e información de horarios.\n"
        f"5. CTA FLOTANTE Y FOOTER: Botón persistente de atención directa por WhatsApp con el texto exacto 'WhatsApp' (nunca abreviar como 'WA').\n\n"
        f"INSTRUCCIONES DE EJECUCIÓN MCP PARA OPEN DESIGN:\n"
        f"- Crear el proyecto con create_project('{nombre.lower().replace(' ', '-')}-landing')\n"
        f"- Ejecutar collect_brief y confirm_brief con este paquete de contexto\n"
        f"- Lanzar start_run exigiendo diseño 100% responsive, copywriting rioplatense profesional y animación fluida en scroll."
    )
    return prompt

def main():
    print("Iniciando extracción de 100 leads y generación de prompts...")
    
    catalog_path = os.path.join('data', 'leads_catalog.json')
    if not os.path.exists(catalog_path):
        print(f"Error: {catalog_path} no existe.")
        sys.exit(1)
        
    with open(catalog_path, 'r', encoding='utf-8') as f:
        catalog_data = json.load(f)
        
    by_id = catalog_data.get('by_id', {})
    
    all_leads = []
    seen_names = set()
    
    for lid, item in by_id.items():
        nombre = (item.get('nombre') or '').strip()
        direccion = (item.get('direccion') or '').strip()
        telefono = str(item.get('whatsapp') or item.get('telefono') or '').strip()
        sector = (item.get('sector') or item.get('rubro') or 'Comercio General').strip()
        
        if nombre and nombre.lower() not in seen_names and telefono not in ['None', 'nan', ''] and direccion not in ['None', 'nan', '']:
            if len(nombre) > 2 and len(direccion) > 4:
                seen_names.add(nombre.lower())
                all_leads.append({
                    'id': lid,
                    'nombre': nombre,
                    'direccion': direccion,
                    'telefono': telefono,
                    'whatsapp': item.get('whatsapp') or telefono,
                    'sector': sector,
                    'puntaje': item.get('puntaje_oportunidad') or 80
                })

    print(f"Total de leads válidos encontrados: {len(all_leads)}")
    
    all_leads.sort(key=lambda x: float(x.get('puntaje') or 0), reverse=True)
    
    selected_leads = []
    by_sector = {}
    for lead in all_leads:
        sec = lead['sector']
        by_sector.setdefault(sec, []).append(lead)
        
    sector_keys = list(by_sector.keys())
    while len(selected_leads) < 100 and any(by_sector.values()):
        for sec in sector_keys:
            if len(selected_leads) >= 100:
                break
            if by_sector[sec]:
                selected_leads.append(by_sector[sec].pop(0))

    print(f"Seleccionados exactamente {len(selected_leads)} leads para la hoja de cálculo.")
    
    excel_rows = []
    for lead in selected_leads:
        prompt_text = build_opendesign_prompt(lead)
        num_contacto = clean_phone(lead.get('whatsapp') or lead.get('telefono'))
        ubicacion = f"{lead['direccion']}, Mar del Plata"
        
        excel_rows.append({
            'Local': lead['nombre'],
            'Número': num_contacto,
            'Dónde está': ubicacion,
            'Prompt Open Design': prompt_text
        })
        
    df = pd.DataFrame(excel_rows)
    
    data_output_path = os.path.join('data', 'PROSPECCION_100_PROMPTS_OPEN_DESIGN.xlsx')
    root_output_path = 'PROSPECCION_100_PROMPTS_OPEN_DESIGN.xlsx'
    
    for target_file in [data_output_path, root_output_path]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "100 Prompts Open Design"
        
        ws.views.sheetView[0].showGridLines = True
        
        headers = ['Local', 'Número', 'Dónde está', 'Prompt Open Design']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        for col_num, h_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            
        ws.row_dimensions[1].height = 28
        
        font_regular = Font(name="Arial", size=10, color="0F172A")
        font_bold = Font(name="Arial", size=10, bold=True, color="0F172A")
        font_phone = Font(name="Consolas", size=10, color="0369A1")
        align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        align_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
        
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for row_idx, row_data in enumerate(excel_rows, start=2):
            c_local = ws.cell(row=row_idx, column=1, value=row_data['Local'])
            c_num = ws.cell(row=row_idx, column=2, value=row_data['Número'])
            c_ubic = ws.cell(row=row_idx, column=3, value=row_data['Dónde está'])
            c_prompt = ws.cell(row=row_idx, column=4, value=row_data['Prompt Open Design'])
            
            c_local.font = font_bold
            c_local.alignment = align_top_left
            
            c_num.font = font_phone
            c_num.alignment = align_top_center
            
            c_ubic.font = font_regular
            c_ubic.alignment = align_top_left
            
            c_prompt.font = font_regular
            c_prompt.alignment = align_top_left
            
            fill = zebra_fill if row_idx % 2 == 0 else white_fill
            for c in [c_local, c_num, c_ubic, c_prompt]:
                c.fill = fill
                c.border = thin_border
                
            ws.row_dimensions[row_idx].height = 160
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 85
        
        wb.save(target_file)
        print(f"✅ Archivo Excel guardado con éxito en: {target_file}")

if __name__ == '__main__':
    main()
