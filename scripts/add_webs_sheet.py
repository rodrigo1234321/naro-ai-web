# -*- coding: utf-8 -*-
"""
add_webs_sheet.py
Agrega (o actualiza) la hoja "🌐 Webs Construidas" al Excel principal
PROSPECCION_MAR_DEL_PLATA_COMPLETA.xlsx con todas las webs/HTMLs publicados,
ordenadas por calidad: Premium custom -> Landings personalizadas -> Previews.

Re-ejecutable: cada corrida regenera la hoja completa.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "PROSPECCION_MAR_DEL_PLATA_COMPLETA.xlsx"
RAW = ROOT / "raw_html"
BASE_URL = "https://mis-clientes-html.pages.dev"

# slug -> lead id para webs que no vienen de webs/leads
EXTRA_IDS = {
    "terca": "LEAD-IA-0251",
    "verde-limon": "LEAD-IA-0267",
    "kiva-cafe": "LEAD-IA-5678",
    "restaurante-la-marina": "LEAD-IA-0838",
    "antes-muerta-que-sencilla-showroom": "LEAD-IA-0002",
    "bulgarie": "LEAD-IA-0029",
    "urgencias-odontologicas": "LEAD-IA-1364",
    "cabanas-entre-los-arboles": "LEAD-IA-0654",
}

TIER = {
    "terca": "Premium Custom",
    "verde-limon": "Premium Custom",
    "kiva-cafe": "Premium Custom",
    "restaurante-la-marina": "Premium Custom",
    "antes-muerta-que-sencilla-showroom": "Preview",
    "bulgarie": "Preview",
    "urgencias-odontologicas": "Preview",
    "cabanas-entre-los-arboles": "Preview",
}


def site_total_kb(slug: str) -> float:
    """KB totales del slug (html + assets)."""
    src = RAW / slug
    if not src.exists():
        return 0.0
    total = 0
    for f in src.rglob("*"):
        if f.is_file():
            total += f.stat().st_size
    return total / 1024


def html_kb(slug: str) -> float:
    p = RAW / f"{slug}.html"
    return p.stat().st_size / 1024 if p.exists() else 0.0


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    ws_all = wb["📊 Todos los Prospectos"]
    rows = list(ws_all.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {h: i for i, h in enumerate(hdr[:17])}
    by_id = {}
    for r in rows[1:]:
        if r[idx["id_lead"]]:
            by_id[r[idx["id_lead"]]] = r

    entries = []
    for f in sorted(RAW.glob("*.html")):
        slug = f.stem
        lead_id = EXTRA_IDS.get(slug)
        if not lead_id:
            # buscar en webs/leads/LEAD-IA-XXXX-<slug>
            for folder in (ROOT / "webs" / "leads").iterdir():
                if folder.is_dir() and folder.name.endswith("-" + slug):
                    lead_id = folder.name.split("-", 2)[0] + "-" + folder.name.split("-", 3)[1] + "-" + "-".join(folder.name.split("-")[2:3])
                    if folder.name.startswith("LEAD-IA-"):
                        parts = folder.name.split("-")
                        lead_id = "-".join(parts[:3])
                    break
        r = by_id.get(lead_id)
        nombre = r[idx["nombre"]] if r else slug.replace("-", " ").title()
        rubro = r[idx["sector"]] if r else "-"
        score = r[idx["puntaje_oportunidad"]] if r else "-"
        telefono = r[idx["telefono"]] if r else "-"
        direccion = r[idx["direccion"]] if r else "-"
        wa = r[idx["link_whatsapp_outreach"]] if r else "-"
        entries.append({
            "slug": slug,
            "lead_id": lead_id or "-",
            "nombre": nombre,
            "rubro": rubro,
            "score": score,
            "telefono": telefono or "-",
            "direccion": direccion or "-",
            "wa": wa,
            "html_kb": html_kb(slug),
            "total_kb": site_total_kb(slug),
            "tipo": TIER.get(slug, "Landing Personalizada"),
        })

    # orden: Premium (por tamaño total desc) -> Landings (html kb desc) -> Preview
    orden = {"Premium Custom": 0, "Landing Personalizada": 1, "Preview": 2}
    entries.sort(key=lambda e: (orden[e["tipo"]], -(e["total_kb"] if e["tipo"] == "Premium Custom" else e["html_kb"])))

    if "🌐 Webs Construidas" in wb.sheetnames:
        del wb["🌐 Webs Construidas"]
    ws = wb.create_sheet(title="🌐 Webs Construidas")

    headers = [
        "#", "Tipo", "Nombre", "Rubro", "Score", "Web (HTML KB)",
        "Total con assets (KB)", "URL Pública", "Teléfono", "WhatsApp", "Dirección", "ID Lead",
    ]
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "Premium Custom": PatternFill("solid", fgColor="D1FAE5"),
        "Landing Personalizada": PatternFill("solid", fgColor="FEF3C7"),
        "Preview": PatternFill("solid", fgColor="E5E7EB"),
    }

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for i, e in enumerate(entries, 1):
        url = f"{BASE_URL}/{e['slug']}/"
        row = [
            i, e["tipo"], e["nombre"], e["rubro"], e["score"],
            round(e["html_kb"], 1), round(e["total_kb"], 1),
            url, e["telefono"], e["wa"], e["direccion"], e["lead_id"],
        ]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.border = border
            if c == 8:
                cell.hyperlink = url
                cell.font = Font(color="2563EB", underline="single")
            cell.fill = fills.get(e["tipo"], PatternFill())
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 3))

    widths = [5, 20, 34, 26, 7, 12, 20, 46, 16, 50, 34, 14]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(entries) + 1}"

    wb.save(XLSX)
    print(f"Hoja '🌐 Webs Construidas' actualizada: {len(entries)} webs")


if __name__ == "__main__":
    main()
